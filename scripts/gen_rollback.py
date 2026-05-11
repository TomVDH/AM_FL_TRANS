import openpyxl
import glob
import os
import json

targets = {
    "Muilengemmers": "Muilegemmers",
    "Muilengem": "Muilegem",
    "MUILENGEM": "MUILEGEM"
}

corrections = []

for file_path in glob.glob("excels/*.xlsx"):
    file_name = os.path.basename(file_path)
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if len(row) > 9:
                    val = row[9]
                    if val is None: continue
                    val_str = str(val)
                    new_val = val_str
                    has_change = False
                    
                    for target, replacement in targets.items():
                        if target in new_val:
                            new_val = new_val.replace(target, replacement)
                            has_change = True
                    
                    if has_change:
                        corrections.append({
                            "file": file_name,
                            "sheet": sheet_name,
                            "cell": f"J{row_idx}",
                            "english": str(row[2]) if len(row) > 2 else "",
                            "current_nl": val_str,
                            "proposed_nl": new_val
                        })
    except Exception as e:
        print(f"Error reading {file_name}: {e}")

output_path = "data/editorial/rollback-muilengem.json"
with open(output_path, "w", encoding="utf-8") as f:
    json.dump({"corrections": corrections}, f, indent=2, ensure_ascii=False)

print(f"Generated {len(corrections)} rollback corrections in {output_path}")
