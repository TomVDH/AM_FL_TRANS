#!/usr/bin/env python3
"""
generate-character-profiles.py

Reads ALL Excel translation files, extracts every NL dialogue line per speaker,
and generates a comprehensive linguistic voice profile for each character.

Outputs:
  data/analysis/character-profiles.json  — full structured profiles
  stdout                                 — human-readable summary

Usage:
  python3 scripts/analysis/generate-character-profiles.py
  python3 scripts/analysis/generate-character-profiles.py --speaker="Bad Ass"
  python3 scripts/analysis/generate-character-profiles.py --top=20
  python3 scripts/analysis/generate-character-profiles.py --min-lines=5
"""

import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_DIR = PROJECT_ROOT / "excels"
CODEX_PATH = PROJECT_ROOT / "data" / "json" / "codex_translations.json"
VERIFIED_PATH = PROJECT_ROOT / "data" / "json" / "codex_verified.json"
OUT_DIR = PROJECT_ROOT / "data" / "analysis"
OUT_FILE = OUT_DIR / "character-profiles.json"

# ---------------------------------------------------------------------------
# CLI args
# ---------------------------------------------------------------------------
args = sys.argv[1:]
speaker_filter = None
top_n = 0
min_lines = 1

for arg in args:
    if arg.startswith("--speaker="):
        speaker_filter = arg.split("=", 1)[1]
    elif arg.startswith("--top="):
        top_n = int(arg.split("=", 1)[1])
    elif arg.startswith("--min-lines="):
        min_lines = int(arg.split("=", 1)[1])

# ---------------------------------------------------------------------------
# Linguistic patterns (ported from analyze-nl-corpus.js)
# ---------------------------------------------------------------------------
PRONOUN_PATTERNS = {
    "ge/gij": re.compile(r"\b(ge|gij|gijlie)\b", re.I),
    "je/jij": re.compile(r"\b(je|jij|jullie)\b", re.I),
    "u": re.compile(r"\b(u|uw)\b", re.I),
}

FLEMISH_MARKERS = {
    "nie": re.compile(r"\bnie\b", re.I),
    "da": re.compile(r"\bda\b", re.I),
    "allez": re.compile(r"\ballez\b", re.I),
    "amai": re.compile(r"\bamai\b", re.I),
    "zenne": re.compile(r"\bzenne\b", re.I),
    "nen/ne": re.compile(r"\b(nen|ne)\b", re.I),
    "manneke": re.compile(r"\bmanneke\b", re.I),
    "goesting": re.compile(r"\bgoesting\b", re.I),
    "pansen": re.compile(r"\bpansen\b", re.I),
    "zever": re.compile(r"\bzever\b", re.I),
    "-ke diminutive": re.compile(r"\w+ke\b", re.I),
    "godver": re.compile(r"godver\w*", re.I),
}

CONTRACTION_PATTERNS = {
    "'k": re.compile(r"['\u2018\u2019]\s*k\b", re.I),
    "'t": re.compile(r"\b['\u2018\u2019]\s*t\b", re.I),
    "'s": re.compile(r"\b['\u2018\u2019]\s*s\b", re.I),
    "da's": re.compile(r"\bda['\u2018\u2019]\s*s\b", re.I),
}

NEGATION_PATTERNS = {
    "nie": re.compile(r"\bnie\b", re.I),
    "niet": re.compile(r"\bniet\b", re.I),
}

# Address terms
ADDRESS_PATTERNS = {
    "kameraad": re.compile(r"\bkameraad\b", re.I),
    "manneke": re.compile(r"\bmanneke\b", re.I),
    "nonkel": re.compile(r"\bnonkel\b", re.I),
    "veulentje": re.compile(r"\bveulentje\b", re.I),
    "kleintje": re.compile(r"\bkleintje\b", re.I),
    "snotje": re.compile(r"\bsnotje\b", re.I),
}

# Exclamations / interjections
EXCLAMATION_PATTERNS = {
    "oesje": re.compile(r"\boesje\b", re.I),
    "jakkes": re.compile(r"\bjakkes\b", re.I),
    "jemig": re.compile(r"\bjemig\b", re.I),
    "amai": re.compile(r"\bamai\b", re.I),
    "godverdomme": re.compile(r"\bgodverdomme\b", re.I),
    "verdikke": re.compile(r"\bverdikke\b", re.I),
    "potvolkoffie": re.compile(r"\bpotvolkoffie\b", re.I),
    "hee": re.compile(r"\bh[eé]{2,}\b", re.I),
    "oef": re.compile(r"\boef\b", re.I),
    "zucht": re.compile(r"\bzucht\b", re.I),
    "snik": re.compile(r"\bsnik\b", re.I),
}


# ---------------------------------------------------------------------------
# Excel extraction
# ---------------------------------------------------------------------------
def extract_all_dialogue():
    """Read all Excel files and extract dialogue rows with speaker, EN, NL."""
    files = sorted(
        [f for f in os.listdir(EXCEL_DIR) if f.endswith(".xlsx") and re.match(r"^\d+_asses", f)],
        key=lambda f: int(re.match(r"^(\d+)", f).group(1)),
    )

    all_rows = []

    for filename in files:
        ep_match = re.match(r"^(\d+)_", filename)
        episode = f"E{ep_match.group(1)}" if ep_match else filename

        wb = openpyxl.load_workbook(EXCEL_DIR / filename, read_only=True, data_only=True)

        for ws in wb.worksheets:
            # Find columns by header
            header = [str(cell.value or "").strip().upper() for cell in ws[1]]
            nl_col = next((i for i, h in enumerate(header) if h == "NL"), None)
            en_col = next((i for i, h in enumerate(header) if h in ("STANDARD", "ENGLISH")), None)
            key_col = next((i for i, h in enumerate(header) if h == "KEY"), None)

            if nl_col is None or en_col is None:
                continue

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) <= max(nl_col, en_col):
                    continue

                key = str(row[key_col] if key_col is not None and key_col < len(row) else "").strip()
                english = str(row[en_col] or "").strip()
                dutch = str(row[nl_col] or "").strip()

                if not key or not english or not dutch:
                    continue
                if key == "SPACER":
                    continue

                # Extract speaker from key: SAY.Dialog:XXX.NN.Speaker Name
                parts = key.split(".")
                speaker = None
                if len(parts) >= 4:
                    speaker = ".".join(parts[3:]).strip()

                if speaker and len(speaker) > 1:
                    all_rows.append({
                        "speaker": speaker,
                        "episode": episode,
                        "sheet": ws.title,
                        "english": english,
                        "dutch": dutch,
                        "key": key,
                    })

        wb.close()

    return all_rows


# ---------------------------------------------------------------------------
# Analysis functions
# ---------------------------------------------------------------------------
def count_matches(text, pattern):
    return len(pattern.findall(text))


def find_repeated_phrases(lines, min_len=3, max_len=6, min_occurrences=2):
    """Find phrases (3-6 words) that appear across multiple lines."""
    phrase_counts = Counter()

    for line in lines:
        words = re.sub(r'[.,!?;:\"""\'()]', "", line.lower()).split()
        seen = set()
        for length in range(min_len, min(max_len + 1, len(words) + 1)):
            for i in range(len(words) - length + 1):
                phrase = " ".join(words[i : i + length])
                if phrase not in seen:
                    seen.add(phrase)
                    phrase_counts[phrase] += 1

    return [
        {"phrase": phrase, "count": count}
        for phrase, count in phrase_counts.most_common(20)
        if count >= min_occurrences
    ]


def score_line_flemish(dutch):
    """Score a single line for Flemish marker density."""
    score = 0
    for pattern in FLEMISH_MARKERS.values():
        score += count_matches(dutch, pattern)
    for pattern in CONTRACTION_PATTERNS.values():
        score += count_matches(dutch, pattern)
    return score


def select_sample_lines(rows, repeated_phrases):
    """Pick up to 5 representative lines for a character."""
    if not rows:
        return []

    samples = []
    used_indices = set()
    used_episodes = set()

    # Score all lines for Flemish density
    scored = [(i, score_line_flemish(r["dutch"]), r) for i, r in enumerate(rows)]

    # 1. Line with most Flemish markers
    scored_sorted = sorted(scored, key=lambda x: x[1], reverse=True)
    for idx, score, row in scored_sorted:
        if score > 0 and idx not in used_indices:
            samples.append(row)
            used_indices.add(idx)
            used_episodes.add(row["episode"])
            break

    # 2. Line containing a top repeated phrase (verbal tic)
    if repeated_phrases:
        top_phrase = repeated_phrases[0]["phrase"]
        for idx, _, row in scored:
            if idx not in used_indices and top_phrase in row["dutch"].lower():
                samples.append(row)
                used_indices.add(idx)
                used_episodes.add(row["episode"])
                break

    # 3. Longest line (shows full speech pattern)
    by_length = sorted(scored, key=lambda x: len(x[2]["dutch"]), reverse=True)
    for idx, _, row in by_length:
        if idx not in used_indices:
            samples.append(row)
            used_indices.add(idx)
            used_episodes.add(row["episode"])
            break

    # 4. Line with contractions
    for idx, _, row in scored:
        if idx not in used_indices:
            has_contraction = any(
                count_matches(row["dutch"], p) > 0 for p in CONTRACTION_PATTERNS.values()
            )
            if has_contraction:
                samples.append(row)
                used_indices.add(idx)
                used_episodes.add(row["episode"])
                break

    # 5. Line from a different episode
    all_episodes = {r["episode"] for r in rows}
    missing_eps = all_episodes - used_episodes
    if missing_eps:
        target_ep = sorted(missing_eps)[0]
        for idx, _, row in scored:
            if idx not in used_indices and row["episode"] == target_ep:
                samples.append(row)
                used_indices.add(idx)
                break
    elif len(samples) < 5:
        # Just grab the next unused line
        for idx, _, row in scored:
            if idx not in used_indices:
                samples.append(row)
                used_indices.add(idx)
                break

    return [
        {"english": s["english"], "dutch": s["dutch"], "episode": s["episode"], "sheet": s["sheet"]}
        for s in samples[:5]
    ]


def infer_register(flemish_density, dominant_pronoun, flemish_total, line_count):
    """Infer sociolinguistic register from markers."""
    if flemish_density == "heavy":
        return "plat Vlaams"
    if flemish_density in ("medium", "light"):
        return "tussentaal"
    # zero or trace
    if dominant_pronoun in ("ge/gij",):
        return "tussentaal"
    return "ABN"


def format_contractions(contraction_counts):
    """Format contraction counts into a summary string."""
    active = {k: v for k, v in contraction_counts.items() if v > 0}
    if not active:
        return "none"
    return ", ".join(f"{k} ({v}x)" for k, v in sorted(active.items(), key=lambda x: -x[1]))


def format_verbal_tics(repeated_phrases, exclamation_counts, address_counts):
    """Build a verbal tics summary from repeated phrases and exclamations."""
    parts = []

    # Top repeated phrases
    for rp in repeated_phrases[:5]:
        parts.append(f'"{rp["phrase"]}" (x{rp["count"]})')

    # Exclamations with counts
    for name, count in sorted(exclamation_counts.items(), key=lambda x: -x[1]):
        if count >= 2:
            parts.append(f"{name} (x{count})")

    # Address terms
    for name, count in sorted(address_counts.items(), key=lambda x: -x[1]):
        if count >= 2:
            parts.append(f"{name} (x{count})")

    return ", ".join(parts[:8]) if parts else "none detected"


def analyze_speaker(rows):
    """Full linguistic analysis for one speaker's NL dialogue lines."""
    dutch_lines = [r["dutch"] for r in rows]
    all_text = " ".join(dutch_lines)
    line_count = len(dutch_lines)
    episodes = sorted(set(r["episode"] for r in rows))
    sheets = sorted(set(r["sheet"] for r in rows))

    # Pronoun counts
    pronouns = {}
    for label, pattern in PRONOUN_PATTERNS.items():
        pronouns[label] = count_matches(all_text, pattern)

    # Dominant pronoun
    pronoun_total = sum(pronouns.values())
    if pronoun_total == 0:
        dominant_pronoun = "none"
    else:
        sorted_pronouns = sorted(pronouns.items(), key=lambda x: -x[1])
        dominant_pronoun = sorted_pronouns[0][0]
        if len(sorted_pronouns) > 1 and sorted_pronouns[1][1] > sorted_pronouns[0][1] * 0.25:
            dominant_pronoun = f"mixed ({sorted_pronouns[0][0]}: {sorted_pronouns[0][1]}, {sorted_pronouns[1][0]}: {sorted_pronouns[1][1]})"

    # Flemish markers
    flemish_markers = {}
    flemish_total = 0
    for label, pattern in FLEMISH_MARKERS.items():
        c = count_matches(all_text, pattern)
        if c > 0:
            flemish_markers[label] = c
            flemish_total += c

    # Flemish density
    flemish_rate = flemish_total / line_count if line_count > 0 else 0
    if flemish_rate > 1.5:
        flemish_density = "heavy"
    elif flemish_rate > 0.5:
        flemish_density = "medium"
    elif flemish_rate > 0.1:
        flemish_density = "light"
    elif flemish_total > 0:
        flemish_density = "trace"
    else:
        flemish_density = "zero"

    # Contractions
    contractions = {}
    for label, pattern in CONTRACTION_PATTERNS.items():
        c = count_matches(all_text, pattern)
        if c > 0:
            contractions[label] = c

    # Negation
    negation = {}
    for label, pattern in NEGATION_PATTERNS.items():
        negation[label] = count_matches(all_text, pattern)

    neg_dominant = "niet"
    if negation.get("nie", 0) > negation.get("niet", 0):
        neg_dominant = "nie"
    elif negation.get("nie", 0) > 0 and negation.get("niet", 0) > 0:
        neg_dominant = "mixed"

    # Exclamations
    exclamation_counts = {}
    for label, pattern in EXCLAMATION_PATTERNS.items():
        c = count_matches(all_text, pattern)
        if c > 0:
            exclamation_counts[label] = c

    # Address terms
    address_counts = {}
    for label, pattern in ADDRESS_PATTERNS.items():
        c = count_matches(all_text, pattern)
        if c > 0:
            address_counts[label] = c

    # Repeated phrases
    repeated_phrases = find_repeated_phrases(dutch_lines)

    # Inconsistencies
    inconsistencies = []
    if pronouns.get("ge/gij", 0) > 2 and pronouns.get("je/jij", 0) > 2:
        ratio = min(pronouns["ge/gij"], pronouns["je/jij"]) / max(pronouns["ge/gij"], pronouns["je/jij"])
        if ratio > 0.2:
            inconsistencies.append({
                "type": "mixed-pronouns",
                "detail": f"ge/gij: {pronouns['ge/gij']}, je/jij: {pronouns['je/jij']} — ratio {ratio * 100:.0f}%",
            })

    if negation.get("nie", 0) > 0 and negation.get("niet", 0) > 0:
        inconsistencies.append({
            "type": "mixed-negation",
            "detail": f"nie: {negation['nie']}, niet: {negation['niet']}",
        })

    # Sample lines
    sample_lines = select_sample_lines(rows, repeated_phrases)

    # Derive voice profile
    clean_dominant = dominant_pronoun.split(" ")[0] if dominant_pronoun.startswith("mixed") else dominant_pronoun
    register = infer_register(flemish_density, clean_dominant, flemish_total, line_count)

    # Build note
    note_parts = []
    if contractions:
        top_contr = sorted(contractions.items(), key=lambda x: -x[1])[:3]
        note_parts.append(", ".join(f"{k}:{v}" for k, v in top_contr))
    if flemish_markers:
        top_markers = sorted(flemish_markers.items(), key=lambda x: -x[1])[:3]
        note_parts.append("markers: " + ", ".join(f"{k}:{v}" for k, v in top_markers))
    if negation.get("nie", 0) > 0 and negation.get("niet", 0) > 0:
        note_parts.append(f"negation: nie:{negation['nie']} niet:{negation['niet']}")
    if address_counts:
        top_addr = sorted(address_counts.items(), key=lambda x: -x[1])[:2]
        note_parts.append("address: " + ", ".join(f"{k}:{v}" for k, v in top_addr))

    voice_profile = {
        "flemishDensity": flemish_density,
        "register": register,
        "pronounForm": clean_dominant if not dominant_pronoun.startswith("mixed") else "mixed",
        "contractions": format_contractions(contractions),
        "verbalTics": format_verbal_tics(repeated_phrases, exclamation_counts, address_counts),
        "note": "; ".join(note_parts) if note_parts else "insufficient data for detailed notes",
    }

    return {
        "lineCount": line_count,
        "episodes": episodes,
        "sheets": sheets,
        "pronouns": pronouns,
        "dominantPronoun": dominant_pronoun,
        "flemishMarkers": flemish_markers,
        "flemishTotal": flemish_total,
        "flemishDensity": flemish_density,
        "contractions": contractions,
        "negation": negation,
        "negationDominant": neg_dominant,
        "exclamations": exclamation_counts,
        "addressTerms": address_counts,
        "repeatedPhrases": repeated_phrases,
        "inconsistencies": inconsistencies,
        "sampleLines": sample_lines,
        "voiceProfile": voice_profile,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Extracting NL dialogue from all Excel sheets...\n")

    all_rows = extract_all_dialogue()
    print(f"Total dialogue rows with NL: {len(all_rows)}")

    # Group by speaker
    by_speaker = defaultdict(list)
    for row in all_rows:
        by_speaker[row["speaker"]].append(row)

    print(f"Unique speakers: {len(by_speaker)}\n")

    # Load codex for metadata
    codex_entries = {}
    if CODEX_PATH.exists():
        codex = json.loads(CODEX_PATH.read_text())
        for entry in codex.get("entries", []):
            codex_entries[entry["name"]] = entry

    # Load verified profiles
    verified_profiles = {}
    if VERIFIED_PATH.exists():
        verified = json.loads(VERIFIED_PATH.read_text())
        for entry in verified.get("entries", []):
            verified_profiles[entry["name"]] = entry

    # Filter
    if speaker_filter:
        matches = {k: v for k, v in by_speaker.items() if k.lower() == speaker_filter.lower()}
        if not matches:
            # Fuzzy
            fuzzy = [k for k in by_speaker if speaker_filter.lower() in k.lower()]
            if fuzzy:
                print(f'Speaker "{speaker_filter}" not found. Did you mean: {", ".join(fuzzy)}?')
            else:
                print(f'Speaker "{speaker_filter}" not found.')
            return
        by_speaker = matches

    # Sort by line count
    speakers_sorted = sorted(by_speaker.items(), key=lambda x: len(x[1]), reverse=True)

    # Apply filters
    if min_lines > 1:
        speakers_sorted = [(s, rows) for s, rows in speakers_sorted if len(rows) >= min_lines]
    if top_n > 0:
        speakers_sorted = speakers_sorted[:top_n]

    # Analyze each speaker
    profiles = {}

    for speaker, rows in speakers_sorted:
        analysis = analyze_speaker(rows)

        # Merge codex metadata
        codex_entry = codex_entries.get(speaker, {})
        dutch_name = codex_entry.get("dutch", "")
        gender = codex_entry.get("gender", "")
        bio = codex_entry.get("bio", "")
        nicknames = codex_entry.get("nicknames", [])

        # Check if verified — preserve hand-curated profile
        is_verified = speaker in verified_profiles
        if is_verified:
            vp = verified_profiles[speaker]
            analysis["voiceProfile"] = {
                "flemishDensity": vp.get("flemishDensity", analysis["voiceProfile"]["flemishDensity"]),
                "register": vp.get("register", analysis["voiceProfile"]["register"]),
                "pronounForm": vp.get("pronounForm", analysis["voiceProfile"]["pronounForm"]),
                "contractions": vp.get("contractions", analysis["voiceProfile"]["contractions"]),
                "verbalTics": vp.get("verbalTics", analysis["voiceProfile"]["verbalTics"]),
                "dynamics": vp.get("dynamics", ""),
                "relationships": vp.get("relationships", ""),
                "note": vp.get("note", analysis["voiceProfile"]["note"]),
            }

        profile = {
            "name": speaker,
            "dutch": dutch_name,
            "gender": gender,
            "bio": bio,
            "nicknames": nicknames,
            "isVerified": is_verified,
            **analysis,
        }

        profiles[speaker] = profile

        # Print summary
        v_marker = " [VERIFIED]" if is_verified else ""
        print(f"{'━' * 80}")
        print(f"{speaker} ({dutch_name or '?'}){v_marker}")
        print(f"  {analysis['lineCount']} lines across {', '.join(analysis['episodes'])}")
        print(f"  Sheets: {', '.join(analysis['sheets'][:5])}{'...' if len(analysis['sheets']) > 5 else ''}")
        print()

        vp = analysis["voiceProfile"]
        print(f"  VOICE PROFILE:")
        print(f"    Flemish density:  {vp['flemishDensity']}")
        print(f"    Register:         {vp['register']}")
        print(f"    Pronoun form:     {vp.get('pronounForm', 'unknown')}")
        print(f"    Contractions:     {vp['contractions']}")
        print(f"    Verbal tics:      {vp['verbalTics']}")
        if is_verified:
            if vp.get("dynamics"):
                print(f"    Dynamics:         {vp['dynamics']}")
            if vp.get("relationships"):
                print(f"    Relationships:    {vp['relationships']}")
        print(f"    Note:             {vp['note']}")

        # Pronouns
        p = analysis["pronouns"]
        print(f"\n  PRONOUNS: ge/gij:{p.get('ge/gij',0)}  je/jij:{p.get('je/jij',0)}  u:{p.get('u',0)}  → {analysis['dominantPronoun']}")

        # Negation
        neg = analysis["negation"]
        print(f"  NEGATION: nie:{neg.get('nie',0)}  niet:{neg.get('niet',0)}  → {analysis['negationDominant']}")

        # Flemish markers
        if analysis["flemishMarkers"]:
            markers = sorted(analysis["flemishMarkers"].items(), key=lambda x: -x[1])
            print(f"  FLEMISH MARKERS ({analysis['flemishTotal']} total): {', '.join(f'{k}:{v}' for k, v in markers)}")

        # Exclamations
        if analysis["exclamations"]:
            excl = sorted(analysis["exclamations"].items(), key=lambda x: -x[1])
            print(f"  EXCLAMATIONS: {', '.join(f'{k}:{v}' for k, v in excl)}")

        # Address terms
        if analysis["addressTerms"]:
            addr = sorted(analysis["addressTerms"].items(), key=lambda x: -x[1])
            print(f"  ADDRESS TERMS: {', '.join(f'{k}:{v}' for k, v in addr)}")

        # Repeated phrases
        if analysis["repeatedPhrases"]:
            phrases = analysis["repeatedPhrases"][:5]
            formatted = ", ".join(f'"{rp["phrase"]}" x{rp["count"]}' for rp in phrases)
            print(f"  REPEATED PHRASES: {formatted}")

        # Inconsistencies
        for inc in analysis["inconsistencies"]:
            print(f"  ⚠ INCONSISTENCY [{inc['type']}]: {inc['detail']}")

        # Sample lines
        if analysis["sampleLines"]:
            print(f"\n  SAMPLE LINES:")
            for sl in analysis["sampleLines"]:
                print(f"    [{sl['episode']}] EN: {sl['english'][:80]}{'...' if len(sl['english']) > 80 else ''}")
                print(f"          NL: {sl['dutch'][:80]}{'...' if len(sl['dutch']) > 80 else ''}")

        print()

    # Write output
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(json.dumps(profiles, indent=2, ensure_ascii=False) + "\n")

    print(f"{'━' * 80}")
    print(f"Profiles written to: {OUT_FILE}")
    print(f"Total speakers profiled: {len(profiles)}")
    verified_count = sum(1 for p in profiles.values() if p["isVerified"])
    generated_count = len(profiles) - verified_count
    print(f"  Verified (hand-curated): {verified_count}")
    print(f"  Generated (auto):        {generated_count}")


if __name__ == "__main__":
    main()
