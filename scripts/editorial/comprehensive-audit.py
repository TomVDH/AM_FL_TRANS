#!/usr/bin/env python3
"""Comprehensive canon-adherence audit across E0-E10.

Applies the most-comprehensive rule set (e10_sweep_scan overlay) + register-drift
check (from e6_sweep_scan) + alignment heuristics (e10_alignment_scan) to every
workbook, producing per-episode markdown reports and a master INDEX.

Speaker-column convention:
  E0-E4: column B (row[1]) holds speaker (descriptive cell)
  E5-E10: column A (row[0]) holds dialog key; speaker = last '.'-separated segment

Output: data/editorial/audit-2026-05-12-E{N}.md per episode + INDEX.
LOGGING ONLY — no xlsx edits, no commits.
"""
import re
import sys
from pathlib import Path
from difflib import SequenceMatcher
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
AUDIT_DATE = '2026-05-12'
CANON_SHA = '24c0df0'
PUSH_LOG = REPO / 'data' / 'editorial' / '_PUSH-LOG.md'

EXCELS = [
    (0, 'Manager+Intermissions+E0Proxy', 'B'),
    (1, 'E1Proxy', 'B'),
    (2, 'E2Proxy', 'B'),
    (3, 'E3Proxy', 'B'),
    (4, 'E4Proxy', 'B'),
    (5, 'E5Proxy', 'A'),
    (6, 'E6Proxy', 'A'),
    (7, 'E7Proxy', 'A'),
    (8, 'E8Proxy', 'A'),
    (9, 'E9Proxy', 'A'),
    (10, 'E10Proxy', 'A'),
]

# Speaker sets — canon §5 + §5.0 update (Helpful Ass → ge/gij)
GE_GIJ_SPEAKERS = {
    'Bad Ass', 'Big Ass', 'Cole-Machine', 'Golden Ass', 'Hard Ass',
    'Haw', 'Hee', 'Kick Ass', 'Old Ass', 'Thirsty Ass', 'Miner Jenny',
    'Helpful Ass',
}
JE_JIJ_SPEAKERS = {
    'Nice Ass', 'Sad Ass', 'Slow Ass', 'Smart Ass', 'Sturdy Ass',
    'Trusty Ass', 'Lazy Ass', 'Mme. Derriere', 'Mme. Derrière',
    'Ringmaster Rico', 'Zookeeper Rose', 'Radio Host Marcos',
    'Butte', 'Child Joey', 'Grandma Kulan', 'Melvin',
}
# §17 Q19 exception — Resentful Ass keeps gij+uw
RESENTFUL_WHITELIST = {'Resentful Ass'}
# §1 mixed/mute — exclude from drift check
MIXED_OR_MUTE = {'Foal', 'Sick Ass'}

RETIRED_MONIKERS = ['Felle Gast', 'Betweter', 'Bikkeharde', 'Snotezel',
                    'Stampgast', 'Sloom']

MACHINE_COMPOUNDS_UNHYPHENATED = [
    r'\bBoormachine[ns]?\b',
    r'\bCamion[ ]Machine[ns]?\b',
    r'\bTractor[ ]Machine[ns]?\b',
    r'\bAuto[ ]Machine[ns]?\b',
    r'\bMysterie[ ]Machine[ns]?\b',
    r'\bTank[ ]Machine[ns]?\b',
    r'\bVlieg[ ]Machine[ns]?\b',
]

SIGN_TYPOS = ['NIEW', 'gesubsidiëerde', 'regios']

DIARY_SHEETS = {'E3_100_localization', 'E3_200_localization', 'E3_300_localization'}

HARD_LOCK_RULES = {'§2', '§3.1', '§3.4', '§3.4.1', '§3.6', '§4.3', '§4.4',
                   '§6.1', '§6.11', '§8.1', '§13.5', '§14.1'}
BUG_RULES = {'LEN-RATIO', 'EMPTY-NL', 'EMPTY-EN', 'CROSS-ROW', '§9.2', '§15.6'}


def extract_speaker_from_key(key):
    if not key:
        return ''
    parts = str(key).split('.')
    return parts[-1].strip() if len(parts) >= 2 else str(key).strip()


def get_speaker(row, convention):
    if convention == 'B':
        return (row[1] or '').strip() if len(row) > 1 else ''
    return extract_speaker_from_key(row[0] or '') if row else ''


def severity_tier(rule):
    if rule in HARD_LOCK_RULES:
        return 'HARD-LOCK'
    if rule in BUG_RULES:
        return 'BUG'
    return 'CAP/STYLE'


def parse_push_log(path):
    """Parse _PUSH-LOG.md → {(episode, sheet, row): [push_entry, ...]}.

    Handles 5 cell-entry formats observed in the log:
      A. E0-E2 standard: **[ID] J<row> · <speaker>** with multi-line EN/Was/Pushed/Rule
      B. E3+ inline-sheet: **[ID] <Sheet> J<row> · <speaker>** with multi-line fields
      C. E3+ compact: **[ID] J<row> · §<rule>** <pushed-text-inline>
      D. E3+ compact w/speaker: **[ID] J<row> · §<rule> · <speaker>** <text>
      E. E3+ arrow-bullets: header then `- <was> → <pushed>` bullets
    """
    if not path.exists():
        return {}
    text = path.read_text()
    pushed = {}
    current_event = None
    current_episode = None
    current_sheet = None
    current_cell = None

    def commit_current():
        if current_cell:
            key = (current_cell['episode'], current_cell['sheet'], current_cell['row'])
            pushed.setdefault(key, []).append(current_cell)

    cell_header_re = re.compile(
        r'^\*\*\[([^\]]+)\]\s+'
        r'(?:(E\d+_[A-Za-z0-9_]+?)\s+)?'
        r'J(\d+)\s*·\s*'
        r'([^*]+?)\*\*\s*(.*)$'
    )

    in_table = False
    table_columns = []  # list of normalized column names

    def strip_md_inline(s):
        s = s.strip()
        if s.startswith('`') and s.endswith('`'):
            s = s[1:-1]
        return s

    for line in text.splitlines():
        # Push event header
        if line.startswith('## '):
            commit_current()
            current_cell = None
            current_sheet = None
            in_table = False
            table_columns = []
            m = re.search(r'E(\d+)', line)
            if m:
                current_event = line.lstrip('# ').strip()
                current_episode = int(m.group(1))
            else:
                current_event = None
                current_episode = None
            continue

        # Table parsing (Format G — E6+)
        if line.startswith('|') and current_episode is not None:
            cells = [c.strip() for c in line.strip().strip('|').split('|')]
            if not in_table:
                header_lower = [c.lower() for c in cells]
                has_sheet = 'sheet' in header_lower
                has_cell = 'cell' in header_lower
                has_was = any(h in header_lower for h in ('was', 'before', 'test push'))
                has_pushed = any(h in header_lower for h in ('pushed', 'after', 'final state (after revert)'))
                if has_sheet and has_cell and has_was and has_pushed:
                    table_columns = header_lower
                    in_table = True
                    continue
            else:
                if all(set(c) <= set('- :') for c in cells if c):
                    continue
                if len(cells) >= len(table_columns):
                    row_data = dict(zip(table_columns, cells))
                    sheet_raw = row_data.get('sheet', '').strip()
                    cell_ref = row_data.get('cell', '').strip()
                    was_val = strip_md_inline(
                        row_data.get('was', '')
                        or row_data.get('before', '')
                        or row_data.get('test push', '')
                    )
                    pushed_val = strip_md_inline(
                        row_data.get('pushed', '')
                        or row_data.get('after', '')
                        or row_data.get('final state (after revert)', '')
                    )
                    rule_val = row_data.get('rule', '').strip()
                    speaker_val = row_data.get('speaker', '').strip()
                    m = re.match(r'^J(\d+)$', cell_ref)
                    if m and sheet_raw:
                        sheet = sheet_raw if sheet_raw.endswith('_localization') else sheet_raw + '_localization'
                        commit_current()
                        current_cell = {
                            'id': f'E{current_episode}-T{m.group(1)}',
                            'row': int(m.group(1)),
                            'speaker': speaker_val,
                            'episode': current_episode,
                            'sheet': sheet,
                            'event': current_event,
                            'en': '',
                            'original': '',
                            'was': was_val,
                            'pushed': pushed_val,
                            'rule': rule_val,
                        }
                        commit_current()
                        current_cell = None
                continue

        if in_table and not line.startswith('|'):
            in_table = False
            table_columns = []

        # Sheet / sub-push header (level 3)
        m_h3 = re.match(r'^###\s+(.+?)(?:\s+\(.*\))?\s*$', line)
        if m_h3 and current_episode is not None:
            commit_current()
            current_cell = None
            heading = m_h3.group(1).strip()
            if heading.endswith('_localization'):
                current_sheet = heading
            # else: sub-push heading (e.g. "Push 1 (4 staged cells)") — don't reset sheet,
            # rely on inline-sheet in cell headers instead.
            continue

        # Cell header (formats A–D)
        m = cell_header_re.match(line)
        if m and current_episode is not None:
            commit_current()
            id_, sheet_inline, row_s, descriptor, trailing = m.group(1, 2, 3, 4, 5)
            sheet = current_sheet
            if sheet_inline:
                sheet = sheet_inline if sheet_inline.endswith('_localization') else sheet_inline + '_localization'
            if not sheet:
                current_cell = None
                continue
            descriptor = descriptor.strip()
            parts = [p.strip() for p in descriptor.split('·') if p.strip()]
            speaker = ''
            rule_in_header = ''
            for p in parts:
                if p.startswith('§'):
                    rule_in_header = (rule_in_header + ' + ' + p) if rule_in_header else p
                elif not speaker:
                    speaker = p
            was_val = ''
            pushed_val = ''
            trailing = trailing.strip()
            if trailing:
                arrow = re.search(r'(.+?)\s*→\s*(.+)$', trailing)
                if arrow:
                    was_val = arrow.group(1).strip()
                    pushed_val = arrow.group(2).strip()
                else:
                    pushed_val = trailing
            current_cell = {
                'id': id_,
                'row': int(row_s),
                'speaker': speaker,
                'episode': current_episode,
                'sheet': sheet,
                'event': current_event,
                'en': '',
                'original': '',
                'was': was_val,
                'pushed': pushed_val,
                'rule': rule_in_header,
            }
            continue

        # Multi-line field extraction (formats A, B, E)
        if current_cell:
            m = re.match(r'^-\s+\*\*EN:\*\*\s+(.*)$', line)
            if m:
                current_cell['en'] = m.group(1).strip()
                continue
            m = re.match(r'^-\s+\*\*Original[^*]*:\*\*\s+(.*)$', line)
            if m:
                current_cell['original'] = m.group(1).strip()
                continue
            m = re.match(r'^-\s+\*\*Was[^*]*:\*\*\s+(.*)$', line)
            if m:
                current_cell['was'] = m.group(1).strip()
                continue
            m = re.match(r'^-\s+\*\*Pushed:\*\*\s+(.*)$', line)
            if m:
                current_cell['pushed'] = m.group(1).strip()
                continue
            m = re.match(r'^-\s+\*Rule:\*\s+(.*)$', line)
            if m:
                if current_cell['rule']:
                    current_cell['rule'] += ' / ' + m.group(1).strip()
                else:
                    current_cell['rule'] = m.group(1).strip()
                continue
            # Format E: bullet arrow changes (capture latest)
            m = re.match(r'^-\s+(.+?)\s*→\s*(.+)$', line)
            if m and not current_cell['pushed']:
                current_cell['was'] = m.group(1).strip()
                current_cell['pushed'] = m.group(2).strip()
                continue

    commit_current()
    return pushed


def latest_pushed_value(history):
    """Return the most-recent non-empty 'pushed' value from a list of push entries."""
    for entry in reversed(history):
        if entry.get('pushed'):
            return entry['pushed']
    return None


def is_fragment(pushed):
    """Detect ellipsis-bracketed fragments (table-format push entries)."""
    if not pushed:
        return False
    if '…' in pushed:
        return True
    if pushed.startswith('...') or pushed.endswith('...'):
        return True
    # Mid-string `...` used as etc.-marker (whitespace on at least one side)
    if re.search(r'(?:\s\.\.\.|\.\.\.\s)', pushed):
        return True
    return False


META_ANNOTATIONS = ('kept', '(no push)', 'no push')


def diverges(current, pushed):
    """True if current NL doesn't reflect the pushed value.

    Strategy:
    - Filter meta-annotations ("kept", "(no push)") — not actual pushes.
    - Strip surrounding markdown (** and `) from pushed.
    - Normalize literal "\n" in pushed to real newlines.
    - For ellipsis-bracketed values: every non-empty segment must appear in current.
    - For non-fragment values: pushed must equal OR be a substring of current
      (substring catches diff-style table push values that are fragments without ellipsis).
    """
    if not pushed:
        return False
    if not current:
        return True
    p = pushed.strip()
    c = current.strip()
    # Pattern: **kept `<value>`** or kept `<value>` → extract <value>
    kept_match = re.match(r'\**\s*kept\s+`([^`]+)`\s*\**\s*$', p, re.IGNORECASE)
    if kept_match:
        p = kept_match.group(1)
    p_lower = p.lower()
    if p_lower in META_ANNOTATIONS:
        return False
    if 'no push' in p_lower or p_lower.startswith('kept '):
        return False
    # Strip surrounding ** emphasis and ` code markers (may have residual whitespace)
    p = p.strip('*').strip().strip('`').strip()
    # Normalize literal "\n" sequences (from inline-code values in tables) to real newlines
    p_normalized = p.replace('\\n', '\n')
    c_normalized = c.replace('\\n', '\n')
    if is_fragment(p_normalized):
        segments = [s.strip() for s in re.split(r'…|\.\.\.', p_normalized) if s.strip()]
        if not segments:
            return False
        return not all(seg in c_normalized for seg in segments)
    if c_normalized == p_normalized:
        return False
    if p_normalized in c_normalized:
        return False
    # Loose whitespace match (collapse all whitespace to single spaces)
    p_ws = re.sub(r'\s+', ' ', p_normalized)
    c_ws = re.sub(r'\s+', ' ', c_normalized)
    if p_ws in c_ws or p_ws == c_ws:
        return False
    return True


def is_template_row(speaker, en, nl):
    """Identify template/spacer/placeholder rows that aren't real translations."""
    if 'Lorem ipsum' in (speaker or '') or 'Lorem ipsum' in (nl or ''):
        return True
    if (speaker or '').strip() == 'SPACER':
        return True
    # E0 CharacterProfiles header rows
    if en in ('Character Name',) and not nl:
        return True
    if (en or '').startswith('***') and not nl:
        return True
    return False


def buffer_rows(ws, convention):
    """Buffer (r, speaker, en, nl) tuples from a sheet, skipping template rows."""
    rows = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) < 10:
            continue
        speaker = get_speaker(row, convention)
        en = (row[2] or '').strip() if row[2] is not None else ''
        nl = (row[9] or '').strip() if row[9] is not None else ''
        if is_template_row(speaker, en, nl):
            continue
        rows.append((r, speaker, en, nl))
    return rows


def scan_rows_canon(rows, sheet_name):
    """Apply unified canon rule set to buffered rows."""
    findings = []
    diary = sheet_name in DIARY_SHEETS

    for r, speaker, en, nl in rows:
        if not nl:
            continue

        def add(rule, note):
            findings.append((r, rule, note, speaker, en, nl))

        # §2 nie standalone (HARD LOCK)
        if re.search(r"\bnie\b", nl, flags=re.IGNORECASE) or re.search(r"\bnie'", nl):
            add('§2', 'nie standalone → niet (HARD LOCK corpus-wide)')

        # §7.1 Ezel cap mid-sentence
        for m in re.finditer(r"\bezel(s|innekes|tje|tjes|en|kes)?\b", nl):
            if m.start() == 0:
                continue
            prev = nl[max(0, m.start()-2):m.start()]
            if re.search(r"[.!?]\s$", prev + ' '):
                continue
            add('§7.1', f'lowercase "{m.group(0)}" mid-sentence — cap?')
            break

        # §7.2 Circusdirecteur / Tournee cap-always
        if re.search(r"\bcircusdirecteur\b", nl):
            add('§7.2', 'lowercase circusdirecteur — cap-always per §7.2')
        if re.search(r"\btournee\b", nl):
            add('§7.2', 'lowercase tournee — cap-always per §7.2')

        # §3.6 Boerderij → Hoeve
        if re.search(r"\bboerderij\b", nl, flags=re.IGNORECASE):
            add('§3.6', 'Boerderij → Hoeve')

        # §3.1 Muilebeek
        if re.search(r"\bMuilebeek\b", nl):
            add('§3.1', 'Muilebeek → Muilenbeek')

        # §3.4 Mecha / Mechalen
        if re.search(r"\bMecha(len)?\b", nl):
            if 'Ziekenhuis' in nl:
                add('§3.4.1', 'Mecha + Ziekenhuis → Imechelda Algemeen Ziekenhuis')
            else:
                add('§3.4', 'Mecha/Mechalen → Technopolis')
        if 'MECHA' in nl and 'MECHALEN' not in nl:
            add('§3.4', 'MECHA → TECHNOPOLIS (all-caps)')

        # §6.16 Beroep / job — only flag when EN context clearly indicates game-system Job
        if re.search(r"\bBeroep(en)?\b", nl) and 'beroep doen op' not in nl.lower():
            add('§6.16', 'Beroep(en) → Job(s) — verify game-system context')
        if re.search(r"\bjob[s]?\b", nl) and not re.search(r"\bJob[s]?\b", nl):
            if re.search(r"\bJob[s]?\b", en):
                add('§6.16-lc', 'lowercase job/jobs but EN has capitalized Job — game-system cap')

        # §6.1 Hudo / Sekreet / Privaat
        if re.search(r"\b[Hh]udo\b", nl) or 'HUDO' in nl:
            add('§6.1', 'Hudo → Plee')
        if re.search(r"\bSekreet\b|\bPrivaat\b", nl):
            add('§6.1', 'Sekreet/Privaat → Plee')

        # §6.2 Oom (non-Smart Ass)
        if re.search(r"\bOom\b", nl) and 'Smart' not in speaker:
            add('§6.2', f'Oom (speaker={speaker!r}) → Nonkel?')

        # §6.10 Acte/Actes/Acten
        if re.search(r"\b[Aa]cte[sn]?\b", nl):
            add('§6.10', 'Acte/Actes/Acten → Nummer/Nummers')

        # §6.12 Nijg
        if re.search(r"\b[Nn]ijg(t|en|de)?\b", nl):
            add('§6.12', 'Nijg → fel')

        # §6.11 jansen — extinct
        if re.search(r"\bjansen\b", nl, flags=re.IGNORECASE):
            add('§6.11', f'jansen — EXTINCT (Thirsty→gulle/u, Bad→Kleine)')

        # §10.1 doekjes rond winden
        if 'doekjes rond winden' in nl.lower():
            add('§10.1', 'doekjes rond winden → doekjes om winden')

        # §10.7 slokkie/slokje → slokske
        if re.search(r"\bslokk?[ij]e\b", nl, flags=re.IGNORECASE):
            add('§10.7', 'slokkie/slokje → slokske (Flemish diminutive)')

        # §18 Wereldtournee cap-always
        if re.search(r"\bwereldtournee\b", nl):
            add('§18', 'lowercase wereldtournee — cap-always (retcon REVERSED)')

        # §7.4 / §5.2 Gods lowercase u/uw
        god_speakers = {'Haw', 'Hee', 'Golden Ass', 'THE GODS'}
        if any(g in speaker for g in god_speakers):
            if re.search(r"\bu\b|\buw\b", nl):
                add('§7.4', f'God speaker uses lowercase u/uw — should be U/Uw')

        # §12.2 Sturdy motto fragment
        if 'Sturdy' in speaker:
            if re.search(r"\bwerk[- ]afpakkende\b|\bkind[- ]dodende\b|\bzielloze\b|\bslechte\b",
                         nl, flags=re.IGNORECASE):
                add('§12.2', 'Sturdy motto fragment — verify full canonical 4-adj form')

        # §8.1 Machien
        if re.search(r"\bMachien(en)?\b", nl):
            add('§8.1', 'Machien → Machine')

        # §8.2 Machine compound hyphenation
        for pat in MACHINE_COMPOUNDS_UNHYPHENATED:
            if re.search(pat, nl):
                add('§8.2', f'machine compound needs hyphen ({pat})')

        # §4.3 retired monikers
        for mon in RETIRED_MONIKERS:
            if mon in nl:
                add('§4.3', f'retired moniker {mon} — EXTINCT')

        # §4.4 Schoon Beest — preserve only if Thirsty→Nice
        if 'Schoon Beest' in nl:
            thirsty_speaks = 'Thirsty' in speaker
            nice_in_en = 'Nice Ass' in en or 'Nice' in en
            if thirsty_speaks and nice_in_en:
                pass  # canonical Thirsty→Nice — preserve
            else:
                add('§4.4', 'Schoon Beest — verify Thirsty→Nice context (canon §4.4 preserve)')

        # §12.1 Kameraad (EN-driven)
        if re.search(r",\s*Kameraad\b", nl):
            if 'Comrade' not in en:
                add('§12.1', ', Kameraad without EN Comrade — strip?')

        # §14.1 retired slogans
        for slog in ['EZELSKRACHT', 'EZEL MACHT', 'EZELKRACHT', 'EZELS-KRACHT']:
            if slog in nl:
                add('§14.1', f'retired slogan {slog} → EZELS EERST')

        # §5.4 Stop imperative — only flag when register is ge/gij (speaker or NL context)
        if re.search(r"(^|[.!?]\s+)Stop\b", nl):
            is_gegij_speaker = any(g in speaker for g in GE_GIJ_SPEAKERS) if speaker else False
            has_gegij_in_nl = bool(re.search(r'\b(ge|gij|zijt|zijde)\b', nl))
            if is_gegij_speaker or has_gegij_in_nl:
                add('§5.4', f'Stop imperative — ge/gij register wants "Stopt" (stem+t)')

        # §9.1 sentence-start apostrophe missing
        if re.search(r"(^|[.!?]\s+)k\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: k → 'k")
        if re.search(r"(^|[.!?]\s+)t\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: t → 't")

        # §9.6 diary uncontracted
        if diary and re.search(r"'[kt]\b", nl):
            add('§9.6', "diary: contracted 'k/'t → uncontracted Ik/Het")

        # §9.2 mid-sentence Ik
        if re.search(r",\s+Ik\s+[A-Z]", nl):
            add('§9.2', "mid-sentence ', Ik [Cap]' — lowercase verb (Patrick artifact)")

        # §13.5 Reassignment lock
        if 'Reassignment' in en or 'reassign' in en.lower():
            if re.search(r'\bherverkoren\b|\bherrijz', nl, flags=re.IGNORECASE):
                add('§13.5', 'Reassignment → Heraanstelling/heraangesteld (not herverkoren/herrijzen)')

        # §15.6 sign typos
        for typo in SIGN_TYPOS:
            if typo in nl:
                add('§15.6', f'sign typo: {typo}')

        # §5 register drift
        sp_norm = speaker.replace('Mme. Derrière', 'Mme. Derriere')
        in_resentful = sp_norm in RESENTFUL_WHITELIST or 'Resentful' in sp_norm
        in_mixed_mute = any(m in sp_norm for m in MIXED_OR_MUTE)
        if in_resentful or in_mixed_mute:
            pass
        elif sp_norm in GE_GIJ_SPEAKERS:
            jj_hits = re.findall(r'\b(je|jij|jou|jouw)\b', nl, re.IGNORECASE)
            if jj_hits:
                add('§5-drift', f'ge/gij speaker has je/jij form(s): {jj_hits}')
        elif sp_norm in JE_JIJ_SPEAKERS:
            gg_hits = re.findall(r'\b(ge|gij|zijt|zijde)\b', nl, re.IGNORECASE)
            if gg_hits:
                add('§5-drift', f'je/jij speaker has ge/gij form(s): {gg_hits}')

    return findings


def _norm(s):
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def _sim(a, b):
    return SequenceMatcher(None, a, b).ratio()


def scan_rows_alignment(rows, sheet_name):
    LEN_RATIO_LO = 0.4
    LEN_RATIO_HI = 2.5
    MIN_EN_CHARS = 8
    NL_SIM_THRESHOLD = 0.75
    EN_DIFF_THRESHOLD = 0.55
    MIN_LEN_CROSS = 25

    cell_flags = []
    for r, sp, en, nl in rows:
        if en and not nl:
            cell_flags.append((r, 'EMPTY-NL', 'EN non-empty but NL empty', sp, en, nl))
            continue
        if not en and nl:
            cell_flags.append((r, 'EMPTY-EN', 'NL non-empty but EN empty', sp, en, nl))
            continue
        if len(en) >= MIN_EN_CHARS:
            ratio = len(nl) / len(en) if len(en) else 0
            if ratio < LEN_RATIO_LO or ratio > LEN_RATIO_HI:
                cell_flags.append((r, 'LEN-RATIO',
                                   f'NL/EN char-ratio {ratio:.2f} outside [{LEN_RATIO_LO},{LEN_RATIO_HI}]',
                                   sp, en, nl))

    cross_flags = []
    nl_rows = [(r, sp, en, nl) for (r, sp, en, nl) in rows
               if len(nl) >= MIN_LEN_CROSS and en]
    nl_n = {r: _norm(nl) for (r, _, _, nl) in nl_rows}
    en_n = {r: _norm(en) for (r, _, en, _) in nl_rows}
    flagged_pairs = set()
    for i, (r_i, sp_i, en_i, nl_i) in enumerate(nl_rows):
        for r_j, sp_j, en_j, nl_j in nl_rows[i+1:]:
            pair = (r_i, r_j)
            if pair in flagged_pairs:
                continue
            nl_s = _sim(nl_n[r_i], nl_n[r_j])
            if nl_s < NL_SIM_THRESHOLD:
                continue
            en_s = _sim(en_n[r_i], en_n[r_j])
            if en_s > EN_DIFF_THRESHOLD:
                continue
            cross_flags.append((r_i, r_j, sp_i, sp_j, en_i, en_j, nl_i, nl_j, nl_s, en_s))
            flagged_pairs.add(pair)

    return cell_flags, cross_flags


def dedupe(findings):
    by_key = {}
    for f in findings:
        r, rule, note, sp, en, nl = f
        key = (r, rule)
        if key not in by_key or len(note) > len(by_key[key][2]):
            by_key[key] = f
    return sorted(by_key.values(), key=lambda x: (x[0], x[1]))


def md_inline(s):
    if s is None:
        return '∅'
    s = str(s)
    if not s:
        return '∅'
    return s.replace('`', '\\`').replace('|', '\\|').replace('\n', ' ⏎ ')


def _emit_push_history(L, history):
    if not history:
        return
    L.append(f'- Push history ({len(history)} push event(s)):')
    for entry in history:
        L.append(f'  - `[{entry["id"]}]` {md_inline(entry["event"])}')
        if entry.get('rule'):
            L.append(f'    - Rule: `{md_inline(entry["rule"])}`')
        if entry.get('pushed'):
            L.append(f'    - Pushed: `{md_inline(entry["pushed"])}`')


def write_episode_report(ep_num, ep_label, sheets_data, total_cells, out_path):
    L = []
    L.append(f'# Audit — E{ep_num} ({ep_label})')
    L.append('')
    L.append(f'_Generated: {AUDIT_DATE} — canon SHA `{CANON_SHA}` — workbook: `excels/{ep_num}_asses.masses_{ep_label}.xlsx`_')
    L.append('')

    total_findings = sum(len(s['findings']) for s in sheets_data)
    total_align_cell = sum(len(s['align_cell']) for s in sheets_data)
    total_align_cross = sum(len(s['align_cross']) for s in sheets_data)
    total_divergences = sum(len(s['divergences']) for s in sheets_data)
    clean_sheets = sum(1 for s in sheets_data
                       if not s['findings'] and not s['align_cell']
                       and not s['align_cross'] and not s['divergences'])

    L.append('## Summary')
    L.append('')
    L.append(f'- Sheets scanned: **{len(sheets_data)}**')
    L.append(f'- NL cells scanned: **{total_cells}**')
    L.append(f'- Canon findings: **{total_findings}**')
    L.append(f'- Alignment per-cell flags: **{total_align_cell}**')
    L.append(f'- Cross-row paste candidates: **{total_align_cross}**')
    L.append(f'- Push-divergence cells (current NL ≠ most-recent pushed): **{total_divergences}**')
    L.append(f'- Clean sheets: **{clean_sheets}/{len(sheets_data)}**')
    L.append('')

    for sheet in sheets_data:
        sn = sheet['name']
        findings_with_push = sheet['findings_with_push']
        align_cell_with_push = sheet['align_cell_with_push']
        align_cross = sheet['align_cross']
        divergences = sheet['divergences']

        if not findings_with_push and not align_cell_with_push and not align_cross and not divergences:
            L.append(f'### {sn} — 0 findings ✓')
            L.append('')
            continue

        L.append(f'### {sn}')
        L.append('')
        L.append(f'_NL cells: {sheet["cell_count"]} — {len(findings_with_push)} canon | {len(align_cell_with_push)} per-cell align | {len(align_cross)} cross-row | {len(divergences)} push-divergence_')
        L.append('')

        if findings_with_push:
            L.append('#### Canon findings')
            L.append('')
            for (r, rule, note, sp, en, nl), history in findings_with_push:
                tier = severity_tier(rule)
                pushed_tag = ' `[PUSHED-BEFORE]`' if history else ''
                L.append(f'**J{r}** `[{rule}]` `[{tier}]`{pushed_tag} — {md_inline(note)}')
                L.append(f'- Speaker: `{md_inline(sp)}`')
                L.append(f'- EN: `{md_inline(en)}`')
                L.append(f'- NL: `{md_inline(nl)}`')
                _emit_push_history(L, history)
                L.append('')

        if align_cell_with_push:
            L.append('#### Alignment — per-cell flags')
            L.append('')
            for (r, kind, note, sp, en, nl), history in align_cell_with_push:
                pushed_tag = ' `[PUSHED-BEFORE]`' if history else ''
                L.append(f'**J{r}** `[{kind}]` `[BUG]`{pushed_tag} — {md_inline(note)}')
                L.append(f'- Speaker: `{md_inline(sp)}`')
                L.append(f'- EN: `{md_inline(en)}`')
                L.append(f'- NL: `{md_inline(nl)}`')
                _emit_push_history(L, history)
                L.append('')

        if align_cross:
            L.append('#### Alignment — cross-row paste candidates (review queue)')
            L.append('')
            for r_i, r_j, sp_i, sp_j, en_i, en_j, nl_i, nl_j, nl_s, en_s in align_cross:
                L.append(f'**J{r_i} ↔ J{r_j}** `[CROSS-ROW]` `[BUG]` — NL-sim={nl_s:.2f}, EN-sim={en_s:.2f}')
                L.append(f'- J{r_i} speaker `{md_inline(sp_i)}`')
                L.append(f'  - EN: `{md_inline(en_i)}`')
                L.append(f'  - NL: `{md_inline(nl_i)}`')
                L.append(f'- J{r_j} speaker `{md_inline(sp_j)}`')
                L.append(f'  - EN: `{md_inline(en_j)}`')
                L.append(f'  - NL: `{md_inline(nl_j)}`')
                L.append('')

        if divergences:
            L.append('#### Push-divergence — current NL differs from most-recent pushed value')
            L.append('')
            L.append('_These cells were pushed editorially but the current local NL no longer matches what was pushed. Could be Patrick re-edited, or a later push superseded._')
            L.append('')
            for d in divergences:
                L.append(f'**J{d["row"]}** `[DIVERGED-FROM-PUSH]` — push event: {md_inline(d["event"])}')
                L.append(f'- Speaker: `{md_inline(d["speaker"])}`')
                L.append(f'- EN: `{md_inline(d["en"])}`')
                L.append(f'- Pushed: `{md_inline(d["pushed"])}`')
                L.append(f'- Current: `{md_inline(d["current"])}`')
                if d['rule']:
                    L.append(f'- Push rule: `{md_inline(d["rule"])}`')
                L.append('')

    out_path.write_text('\n'.join(L))


def audit_workbook(ep_num, ep_label, convention, push_log):
    xlsx_path = REPO / 'excels' / f'{ep_num}_asses.masses_{ep_label}.xlsx'
    if not xlsx_path.exists():
        return None
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sheets_data = []
    total_cells = 0
    divergences = []  # (sheet, row, speaker, en, pushed, current_nl, push_event, rule)
    print(f'  Sheets: {len(wb.sheetnames)}')
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = buffer_rows(ws, convention)
        cell_count = sum(1 for (_, _, _, nl) in rows if nl)
        # Capture current NL for every row (including template rows) for push-divergence check
        all_rows_map = {}
        # Re-buffer without template filter for divergence check
        ws_full = wb[sheet_name]
        for r, row in enumerate(ws_full.iter_rows(min_row=2, values_only=True), 2):
            if len(row) < 10:
                continue
            nl = (row[9] or '').strip() if row[9] is not None else ''
            all_rows_map[r] = nl

        findings = dedupe(scan_rows_canon(rows, sheet_name))
        align_cell, align_cross = scan_rows_alignment(rows, sheet_name)

        # Attach push history to each finding
        findings_with_push = []
        for f in findings:
            r, rule, note, sp, en, nl = f
            history = push_log.get((ep_num, sheet_name, r), [])
            findings_with_push.append((f, history))

        align_cell_with_push = []
        for f in align_cell:
            r, kind, note, sp, en, nl = f
            history = push_log.get((ep_num, sheet_name, r), [])
            align_cell_with_push.append((f, history))

        # Push-divergence scan for this sheet
        sheet_divergences = []
        for (ep_k, sh_k, row_k), history in push_log.items():
            if ep_k != ep_num or sh_k != sheet_name:
                continue
            latest = latest_pushed_value(history)
            if latest is None:
                continue
            current = all_rows_map.get(row_k, '')
            if diverges(current, latest):
                sheet_divergences.append({
                    'row': row_k,
                    'speaker': history[-1].get('speaker', ''),
                    'en': history[-1].get('en', ''),
                    'pushed': latest,
                    'pushed_is_fragment': is_fragment(latest),
                    'current': current,
                    'event': history[-1].get('event', ''),
                    'rule': history[-1].get('rule', ''),
                })
                divergences.append((sheet_name, row_k, history[-1].get('speaker', ''),
                                    history[-1].get('en', ''), latest, current,
                                    history[-1].get('event', ''), history[-1].get('rule', '')))

        sheets_data.append({
            'name': sheet_name,
            'cell_count': cell_count,
            'findings': findings,
            'findings_with_push': findings_with_push,
            'align_cell': align_cell,
            'align_cell_with_push': align_cell_with_push,
            'align_cross': align_cross,
            'divergences': sheet_divergences,
        })
        total_cells += cell_count
        div_note = f' | {len(sheet_divergences)} push-divergence' if sheet_divergences else ''
        print(f'    {sheet_name}: {cell_count} NL cells → {len(findings)} canon | {len(align_cell)} cell-align | {len(align_cross)} cross-row{div_note}')
    wb.close()
    return sheets_data, total_cells, divergences


def write_index(episode_stats, out_dir, push_log):
    L = []
    L.append(f'# Comprehensive Canon-Adherence Audit — {AUDIT_DATE}')
    L.append('')
    L.append(f'_Canon SHA: `{CANON_SHA}` — Codex: `data/json/codex_verified.json` v3.4 (v3.5 sync pending) — Push log: `data/editorial/_PUSH-LOG.md` ({len(push_log)} unique pushed cells)_')
    L.append('')
    L.append('## Scope')
    L.append('')
    L.append('- Source: local `excels/` (no fresh pull performed; reflects committed/staged local state)')
    L.append('- Episodes: E0–E10 (E11 deferred per canon §19/Q5)')
    L.append('- Cells scanned: column J (Dutch) where non-empty')
    L.append('- Tooling: unified rule scanner (e10_sweep_scan ruleset + §5 register drift from e6 + §13.5/§15.6 additions) + alignment heuristics (length-ratio, empty-NL, cross-row paste) + push-log cross-reference (per-finding push history + current-vs-pushed divergence detection)')
    L.append('- This is a **logging pass**; no xlsx edits, no commits.')
    L.append('')

    total_sheets = sum(len(e['sheets']) for e in episode_stats)
    total_cells = sum(e['total_cells'] for e in episode_stats)
    total_findings = sum(e['total_findings'] for e in episode_stats)
    total_align_cell = sum(e['total_align_cell'] for e in episode_stats)
    total_align_cross = sum(e['total_align_cross'] for e in episode_stats)
    total_divergences = sum(e['total_divergences'] for e in episode_stats)
    total_clean = sum(e['clean_sheets'] for e in episode_stats)

    L.append('## Totals')
    L.append('')
    L.append(f'- Episodes scanned: **{len(episode_stats)}**')
    L.append(f'- Sheets scanned: **{total_sheets}**')
    L.append(f'- NL cells scanned: **{total_cells}**')
    L.append(f'- Canon findings: **{total_findings}**')
    L.append(f'- Alignment per-cell flags: **{total_align_cell}**')
    L.append(f'- Cross-row paste candidates: **{total_align_cross}**')
    L.append(f'- Push-divergence cells (current ≠ pushed): **{total_divergences}**')
    L.append(f'- Pushed cells in log: **{len(push_log)}**')
    L.append(f'- Clean sheets: **{total_clean}/{total_sheets}**')
    L.append('')

    L.append('## Per-episode summary')
    L.append('')
    L.append('| Episode | Sheets | Cells | Canon | Align (cell) | Cross-row | Push-diverge | Clean | Report |')
    L.append('|---|---|---|---|---|---|---|---|---|')
    for e in episode_stats:
        link = f'audit-{AUDIT_DATE}-E{e["ep"]}.md'
        L.append(f'| E{e["ep"]} ({e["label"]}) | {len(e["sheets"])} | {e["total_cells"]} | {e["total_findings"]} | {e["total_align_cell"]} | {e["total_align_cross"]} | {e["total_divergences"]} | {e["clean_sheets"]}/{len(e["sheets"])} | [{link}]({link}) |')
    L.append('')

    triage = []
    rule_tally = {}
    findings_on_pushed = []
    for e in episode_stats:
        for sheet in e['sheets']:
            for (r, rule, note, sp, en, nl), history in sheet['findings_with_push']:
                tier = severity_tier(rule)
                rule_tally[rule] = rule_tally.get(rule, 0) + 1
                if history:
                    findings_on_pushed.append((tier, e['ep'], sheet['name'], r, rule, note, sp, en, nl, history))
                if tier in ('HARD-LOCK', 'BUG'):
                    triage.append((tier, e['ep'], sheet['name'], r, rule, note, sp, en, nl, history))
            for (r, kind, note, sp, en, nl), history in sheet['align_cell_with_push']:
                rule_tally[kind] = rule_tally.get(kind, 0) + 1
                if history:
                    findings_on_pushed.append(('BUG', e['ep'], sheet['name'], r, kind, note, sp, en, nl, history))
                triage.append(('BUG', e['ep'], sheet['name'], r, kind, note, sp, en, nl, history))

    order = {'HARD-LOCK': 0, 'BUG': 1, 'CAP/STYLE': 2}
    triage.sort(key=lambda x: (order[x[0]], x[1], x[2], x[3]))
    findings_on_pushed.sort(key=lambda x: (order[x[0]], x[1], x[2], x[3]))

    L.append('## Findings by rule (counts)')
    L.append('')
    L.append('| Rule | Count | Tier |')
    L.append('|---|---|---|')
    for rule, cnt in sorted(rule_tally.items(), key=lambda x: (-x[1], x[0])):
        L.append(f'| `{rule}` | {cnt} | {severity_tier(rule)} |')
    L.append('')

    L.append('## Triage — HARD-LOCK and BUG findings (ordered)')
    L.append('')
    if not triage:
        L.append('_No hard-lock or bug findings across corpus._')
    else:
        for tier, ep, sn, r, rule, note, sp, en, nl, history in triage:
            pushed_tag = ' `[PUSHED-BEFORE]`' if history else ''
            L.append(f'- **E{ep} / {sn} / J{r}** `[{rule}]` `[{tier}]`{pushed_tag} — {md_inline(note)}')
            L.append(f'  - Speaker: `{md_inline(sp)}` | EN: `{md_inline(en)}` | NL: `{md_inline(nl)}`')
            if history:
                latest = history[-1]
                L.append(f'  - Push: `[{latest["id"]}]` {md_inline(latest["event"])} | rule: `{md_inline(latest["rule"])}`')
                L.append(f'  - Pushed value: `{md_inline(latest["pushed"])}`')
    L.append('')

    L.append('## Findings on previously-pushed cells')
    L.append('')
    L.append('_These cells were pushed in a prior editorial sweep (per `_PUSH-LOG.md`) and the audit re-flags them now. Likely either (a) a canon rule was added after the push and re-push needed, or (b) the push intentionally locked a non-canonical value as an exception — check push rule._')
    L.append('')
    if not findings_on_pushed:
        L.append('_No findings on pushed cells — all flagged drift is on cells that have never been editorially pushed._')
    else:
        for tier, ep, sn, r, rule, note, sp, en, nl, history in findings_on_pushed:
            latest = history[-1]
            L.append(f'- **E{ep} / {sn} / J{r}** `[{rule}]` `[{tier}]` — {md_inline(note)}')
            L.append(f'  - Current NL: `{md_inline(nl)}`')
            L.append(f'  - Pushed: `{md_inline(latest["pushed"])}` (push `[{latest["id"]}]` {md_inline(latest["event"])})')
            L.append(f'  - Push rule: `{md_inline(latest["rule"])}`')
    L.append('')

    L.append('## Push-divergence — current NL diverges from most-recent pushed value')
    L.append('')
    L.append('_Cells where the live xlsx no longer matches what we pushed. Could be Patrick re-edits or later-push supersession. Verify before re-pushing._')
    L.append('')
    all_divergences = []
    for e in episode_stats:
        for sheet in e['sheets']:
            for d in sheet['divergences']:
                all_divergences.append((e['ep'], sheet['name'], d))
    if not all_divergences:
        L.append('_No divergences. All pushed cells in the log still match the local xlsx values._')
    else:
        for ep, sn, d in all_divergences:
            L.append(f'- **E{ep} / {sn} / J{d["row"]}** — push event: {md_inline(d["event"])}')
            L.append(f'  - Speaker: `{md_inline(d["speaker"])}` | EN: `{md_inline(d["en"])}`')
            L.append(f'  - Pushed: `{md_inline(d["pushed"])}`')
            L.append(f'  - Current: `{md_inline(d["current"])}`')
            if d['rule']:
                L.append(f'  - Push rule: `{md_inline(d["rule"])}`')
    L.append('')

    L.append('## Scanner blind spots')
    L.append('')
    L.append('The following canon sections are NOT programmatically enforced; manual eyeball passes needed:')
    L.append('')
    L.append('- **§1 codex voice** — per-character pronoun/contraction/article rules beyond §5. Codex v3.4→v3.5 sync pending: Miner Jenny ge/gij flip, Helpful Ass new entry, Sick Ass mute.')
    L.append('- **§5.1 register exceptions** — Sad/Nice/Slow/Grandma/Foal formal-context u/uw. Resentful Ass `gij+uw` exception (§17 Q19) IS whitelisted; Foal mixed register IS whitelisted; Sick Ass mute IS whitelisted.')
    L.append('- **§6.7 cross-sheet battle-verb consistency** — needs sheet-pivot, no scanner.')
    L.append('- **§6.9 Plan vs afspraak** — EN-driven both-sides check, no scanner.')
    L.append('- **§7.3.1 cap/lc collisions** — `de Mijn` (cap) vs `een mijn` (lc) — context-only.')
    L.append('- **§9.3 terminal punctuation** — matching EN final punct.')
    L.append('- **§12.3 Slow Ass stutter pattern** — Dutch consonants only, not coded.')
    L.append('- **§12.4 English bleed** — no English-word allowlist scanner.')
    L.append('- **§13 broader mistranslations** — only §13.5 Reassignment-Heraanstelling is checked.')
    L.append('')
    L.append('## Constraints')
    L.append('')
    L.append('- EN + NL printed verbatim per cell. No paraphrase, no language invention.')
    L.append('- Fix proposals come from canon text only; ambiguous cases left as flag.')
    L.append('- All findings are **proposals for review**. No automated push.')
    L.append('')

    (out_dir / f'audit-{AUDIT_DATE}-comprehensive-INDEX.md').write_text('\n'.join(L))


def main():
    out_dir = REPO / 'data' / 'editorial'
    target_eps = sys.argv[1:]

    print(f'Loading push log from {PUSH_LOG} ...')
    push_log = parse_push_log(PUSH_LOG)
    push_log_episodes = {k[0] for k in push_log.keys()}
    print(f'  {len(push_log)} unique pushed cells across episodes: {sorted(push_log_episodes)}')

    all_stats = []
    for ep_num, ep_label, convention in EXCELS:
        if target_eps and str(ep_num) not in target_eps:
            continue
        print(f'\n=== AUDITING E{ep_num} ({ep_label}) — speaker col: {convention} ===')
        result = audit_workbook(ep_num, ep_label, convention, push_log)
        if result is None:
            print(f'  !! workbook not found')
            continue
        sheets_data, total_cells, divergences = result
        out_path = out_dir / f'audit-{AUDIT_DATE}-E{ep_num}.md'
        write_episode_report(ep_num, ep_label, sheets_data, total_cells, out_path)
        tf = sum(len(s['findings']) for s in sheets_data)
        tac = sum(len(s['align_cell']) for s in sheets_data)
        tax = sum(len(s['align_cross']) for s in sheets_data)
        td = sum(len(s['divergences']) for s in sheets_data)
        cs = sum(1 for s in sheets_data
                 if not s['findings'] and not s['align_cell']
                 and not s['align_cross'] and not s['divergences'])
        print(f'  E{ep_num} → {len(sheets_data)} sheets | {total_cells} cells | {tf} canon | {tac} cell-align | {tax} cross-row | {td} push-diverge | {cs}/{len(sheets_data)} clean')
        print(f'  Wrote: {out_path}')
        all_stats.append({
            'ep': ep_num, 'label': ep_label, 'sheets': sheets_data,
            'total_cells': total_cells,
            'total_findings': tf,
            'total_align_cell': tac,
            'total_align_cross': tax,
            'total_divergences': td,
            'clean_sheets': cs,
        })

    if not target_eps and all_stats:
        write_index(all_stats, out_dir, push_log)
        print(f'\n  Wrote: {out_dir / f"audit-{AUDIT_DATE}-comprehensive-INDEX.md"}')


if __name__ == '__main__':
    main()
