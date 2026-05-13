#!/usr/bin/env python3
"""Apply user-approved E2 fixes to local xlsx (then push via push-file.py).

Tom sign-off 2026-05-13:
  - J17 BattleButte: KEEP (Flemish — `Ben ik al niet genoeg verloren?`)
  - J22 BattleButte: apply proposed → `Wat is je reactie?` (cross-sheet unify)
  - J28: fix
  - J8 ChildsHouse: fix
  - J16 ChildsHouse: override → `HELP MIJ ALSTUBLIEFT`
  - J20/J23/J35/J39 World_A1: fix
  - J53/J54 World_A1: fix
  - J15 World_A2: fix
  - J3 World_A3: override → `Triestige, deze regen is teken dat we opnieuw kunnen starten!`
  - J25 World_B1: fix
  - J39 World_A1 + B1: option (b) — UNIFY both to one phrasing

14 cells written, 1 KEPT.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/2_asses.masses_E2Proxy.xlsx'

# Unified J39 phrasing for both A1 and B1 (option b)
UNIFIED_J39 = 'DRIE, de rest haalt water uit de rivier, blust het vuur en overtuigt de Mensen om hun Machines achter te laten...'

APPROVED = {
    # J17 BattleButte KEPT (not in dict)
    ('E2_BattleButte_localization', 22): (
        'Hoe reageer je?',
        'Wat is je reactie?',
    ),
    ('E2_BattleButte_localization', 28): (
        'Je VALT ANN!',
        'Je VALT AAN!',
    ),
    ('E2_ChildsHouse_localization', 8): (
        'Hé! Stop met te lanterfanten!',
        'Hé! Stopt met lanterfanten!',
    ),
    ('E2_ChildsHouse_localization', 16): (
        'Please! HELP MIJ!',
        'HELP MIJ ALSTUBLIEFT',
    ),
    ('E2_World_A1_localization', 20): (
        'Oude Ezel en Trouwe Ezel waren niet geschikt voor het leiderschap. Ze zijn te soft.',
        'Oude Ezel en Trouwe Ezel waren niet geschikt voor het leiderschap. Ze waren te zacht.',
    ),
    ('E2_World_A1_localization', 23): (
        '*snuf snuf* Is er dan een p-p-plan C?',
        '*snuf snuf* Is er dan een P-P-Plan C?',
    ),
    ('E2_World_A1_localization', 35): (
        'En d\'anderen! Ze zullen ze schoon gesjareld zijn als we daar niet op tijd zijn, héé!',
        'En d\'anderen! Ze zullen schoon gesjareld zijn als we daar niet op tijd zijn, héé!',
    ),
    # J39 World_A1 — option (b) unified form
    ('E2_World_A1_localization', 39): (
        'DRIE, de rest moet water uit de rivier scheppen, het vuur blussen en de Mensen overtuigen om de Machines te laten...',
        UNIFIED_J39,
    ),
    ('E2_World_A1_localization', 53): (
        'Oesje, de rook is te dik. Er is geen enkele weg naar de Hoeve.',
        'Oesje, de rook is te dik. Dit is helemaal niet de weg naar de Hoeve.',
    ),
    ('E2_World_A1_localization', 54): (
        'Oesje, de rook is te dik. Er is geen enkele weg naar de Hoeve.',
        'Oesje, de rook is te dik. Dit is helemaal niet de weg naar de Hoeve.',
    ),
    ('E2_World_A2_localization', 15): (
        'The Kudde rekent op mij! Ja, ja — dit is waar ik voor geoefend heb.',
        'De Kudde rekent op mij! Ja, ja — dit is waar ik voor geoefend heb.',
    ),
    ('E2_World_A3_localization', 3): (
        'Triestige, deze regen is teken dat we terug opnieuw kunnen starten!',
        'Triestige, deze regen is teken dat we opnieuw kunnen starten!',
    ),
    ('E2_World_B1_localization', 25): (
        'We gaat er onze Hoeven kuisen?',
        'Wie gaat er onze Hoeven kuisen?',
    ),
    # J39 World_B1 — same unified form as A1
    ('E2_World_B1_localization', 39): (
        'DRIE, de rest gaat water van de rivier halen, blust het vuur en overtuigt de mensen hun Machines zo te laten...',
        UNIFIED_J39,
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    if discrepancies:
        for s, r, msg in discrepancies:
            print(f'  ⚠ {s}/J{r}: {msg}')
        sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
