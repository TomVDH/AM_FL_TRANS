#!/usr/bin/env python3
"""Retcon scan: find cells where local NL starts with apostrophe but remote (xlsx-export) lost it.

Compares excels/*.xlsx vs a fresh-pull dir. Outputs cells that need re-push via RAW.
"""
import sys
from pathlib import Path
import openpyxl

LOCAL_DIR = Path('excels')


def main():
    fresh_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('/tmp/retcon-pull')
    for local_path in sorted(LOCAL_DIR.glob('*.xlsx')):
        fresh_path = fresh_dir / local_path.name
        if not fresh_path.exists():
            continue
        lwb = openpyxl.load_workbook(local_path, data_only=True, read_only=True)
        fwb = openpyxl.load_workbook(fresh_path, data_only=True, read_only=True)
        fresh_sheets = list(fwb.sheetnames)
        issues = []
        for s in lwb.sheetnames:
            fname = None
            for fs in fresh_sheets:
                if fs == s or fs.startswith(s) or s.startswith(fs):
                    fname = fs
                    break
            if not fname:
                continue
            ws_l = lwb[s]
            ws_f = fwb[fname]
            l_rows = {r: (row[9] if len(row) > 9 else None)
                      for r, row in enumerate(ws_l.iter_rows(min_row=2, values_only=True), 2)}
            f_rows = {r: (row[9] if len(row) > 9 else None)
                      for r, row in enumerate(ws_f.iter_rows(min_row=2, values_only=True), 2)}
            for r, lv in l_rows.items():
                lv_s = '' if lv is None else str(lv)
                if not lv_s.startswith("'"):
                    continue
                fv = f_rows.get(r)
                fv_s = '' if fv is None else str(fv)
                if not fv_s.startswith("'") and lv_s[1:] == fv_s:
                    issues.append((s, r, lv_s))
        lwb.close()
        fwb.close()
        if issues:
            print(f'\n{local_path.name}: {len(issues)} apostrophe-stripped cell(s)')
            for s, r, snip in issues:
                print(f'  {s} J{r}: {snip[:80]!r}')


if __name__ == '__main__':
    main()
