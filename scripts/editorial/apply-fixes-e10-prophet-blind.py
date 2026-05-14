#!/usr/bin/env python3
"""Apply E10_ProphetSpeech blind-spot walk fixes — Tom sign-off 2026-05-13.

5 writes / 2 keeps (J4/J5/J25 Golden Ass bare-stem — Tom §5.1 archaic-divine
register exception; J46 already at target).

Tom decisions:
  - J28 — `Volk van Technopolis.` (Tom pick; avoids cross-cell Mensen overload)
  - J66 — drop `, kameraad,` insertion (J56/J28 precedent)
  - J4/J5/J25 KEPT — Golden Ass `archaic divine formal` register exception
    over §5.4 (biblical/solemn imperative cadence)
  - J41 — `Parlementsgebouw` capped (compound proper noun match for J46's caps)
  - J85/J86 — `Het lijkt me dat` → `Het is duidelijk dat` (restores Derriere's
    emphatic-certainty rhetoric)
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/10_asses.masses_E10Proxy.xlsx'

APPROVED = {
    # §13 added-vocative (2 writes)
    ('E10_ProphetSpeech_localization', 28): (
        'Kameraden van Technopolis.',
        'Volk van Technopolis.',
    ),
    ('E10_ProphetSpeech_localization', 66): (
        'Net als gij, kameraad, denken deze Ezels van zichzelf dat ze intelligente dieren zijn.',
        'Net als gij denken deze Ezels van zichzelf dat ze intelligente dieren zijn.',
    ),

    # §6.7 / §7.3.1 Parlement cap (1 write; J46 already has `Parlement` capped)
    ('E10_ProphetSpeech_localization', 41): (
        'Die Ezels bezetten uw parlementsgebouw.',
        'Die Ezels bezetten uw Parlementsgebouw.',
    ),

    # §13 Derriere `Clearly` (2 writes)
    ('E10_ProphetSpeech_localization', 85): (
        'Het lijkt me dat ongure radicalen in mijn Fabriek zijn ingebroken, en de Ezels hebben losgelaten op onze prachtige stad.',
        'Het is duidelijk dat ongure radicalen in mijn Fabriek zijn ingebroken, en de Ezels hebben losgelaten op onze prachtige stad.',
    ),
    ('E10_ProphetSpeech_localization', 86): (
        'Het lijkt me dat ongure radicalen mijn Fabriek hebben OPGEBLAZEN, en de Ezels hebben losgelaten op onze prachtige stad.',
        'Het is duidelijk dat ongure radicalen mijn Fabriek hebben OPGEBLAZEN, en de Ezels hebben losgelaten op onze prachtige stad.',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    for s, r, msg in discrepancies:
        print(f'  ⚠ {s}/J{r}: {msg}')
    if discrepancies: sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:90]!r}')
        print(f'    +  {proposed[:90]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
