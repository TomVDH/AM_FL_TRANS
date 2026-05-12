#!/usr/bin/env python3
"""Per-sheet cell dump for deep-eyeball audit.

Usage:
    python3 scripts/editorial/deep-dump.py <ep_num> <sheet_name>

Prints every NL-non-empty cell on the named sheet, with EN + NL verbatim,
speaker (per workbook convention), and push-status from _PUSH-LOG.md.
Spacer/template rows (Lorem ipsum / SPACER) are skipped.

Output format (markdown-friendly so I can read it directly):

    J<row> | speaker=<speaker> | push=Y/N | push-rule=<rule>
    EN:  <verbatim EN>
    NL:  <verbatim NL>
    ---
"""
import sys
import importlib.util
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')

# Load orchestrator for push-log parsing + speaker convention
spec = importlib.util.spec_from_file_location(
    'audit', REPO / 'scripts/editorial/comprehensive-audit.py')
audit = importlib.util.module_from_spec(spec)
spec.loader.exec_module(audit)

CONVENTION = {n: c for n, _, c in audit.EXCELS}
LABEL = {n: lbl for n, lbl, _ in audit.EXCELS}


def dump(ep_num, sheet_name):
    label = LABEL.get(ep_num)
    if not label:
        print(f'!! unknown episode E{ep_num}')
        return
    xlsx = REPO / 'excels' / f'{ep_num}_asses.masses_{label}.xlsx'
    conv = CONVENTION[ep_num]
    push_log = audit.parse_push_log(audit.PUSH_LOG)

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    # Match sheet by exact name or by prefix
    target = None
    for s in wb.sheetnames:
        if s == sheet_name or s.startswith(sheet_name):
            target = s
            break
    if not target:
        print(f'!! sheet not found in workbook: {sheet_name}')
        print(f'   Available: {wb.sheetnames}')
        return
    ws = wb[target]

    print(f'## E{ep_num} / {target}')
    print(f'_Workbook: {xlsx.name} — speaker col: {conv}_')
    print()

    seen = 0
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) < 10:
            continue
        speaker = audit.get_speaker(row, conv)
        en = (row[2] or '').strip() if row[2] is not None else ''
        nl = (row[9] or '').strip() if row[9] is not None else ''
        if audit.is_template_row(speaker, en, nl):
            continue
        if not nl:
            continue
        history = push_log.get((ep_num, target, r), [])
        push_status = 'Y' if history else 'N'
        push_rule = history[-1].get('rule', '') if history else ''
        push_pushed = history[-1].get('pushed', '') if history else ''
        print(f'**J{r}** | speaker=`{speaker or "∅"}` | push={push_status}'
              + (f' | rule=`{push_rule}`' if push_rule else ''))
        print(f'  EN:  `{en}`')
        print(f'  NL:  `{nl}`')
        if push_status == 'Y' and push_pushed and push_pushed not in nl:
            print(f'  ⚠ pushed-fragment: `{push_pushed[:80]}`')
        print()
        seen += 1
    print(f'_End of {target} — {seen} cells_')
    wb.close()


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python3 scripts/editorial/deep-dump.py <ep_num> <sheet_name>')
        sys.exit(1)
    dump(int(sys.argv[1]), sys.argv[2])
