#!/usr/bin/env python3
"""Diff every DRIFT cell: local excels/ vs fresh remote pull.

For each cell flagged DRIFT in audit-2026-05-12-deep-eyeball.md, this script:
  1. Re-reads the cell from local excels/<workbook>.xlsx
  2. Re-reads the cell from <remote-dir>/<workbook>.xlsx (fresh Google-Sheets pull)
  3. Compares EN+NL+speaker between the two
  4. Reports IN-SYNC vs DIVERGED for each cell

Output: data/editorial/rescan-2026-05-13-local-vs-remote-INDEX.md and per-episode files.

This answers: "are my DRIFT claims still drift in the live remote? Or has Patrick
already fixed some of them?"

Usage:
  python3 scripts/editorial/rescan-drift-local-vs-remote.py [remote_dir]

  Default remote_dir: excels.fresh-pull-2026-05-13-pre-E5-fixes
"""
import re
import sys
import importlib.util
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
DEEP_EYEBALL = REPO / 'data/editorial/audit-2026-05-12-deep-eyeball.md'
OUT_DIR = REPO / 'data/editorial'
DEFAULT_REMOTE_DIR = 'excels.fresh-pull-2026-05-13-pre-E5-fixes'
RESCAN_DATE = '2026-05-13'

# Load the rescan parser (re-use parse_deep_eyeball)
spec = importlib.util.spec_from_file_location(
    'rescan', REPO / 'scripts/editorial/rescan-drift-2026-05-12.py')
rescan = importlib.util.module_from_spec(spec)
spec.loader.exec_module(rescan)

spec2 = importlib.util.spec_from_file_location(
    'audit', REPO / 'scripts/editorial/comprehensive-audit.py')
audit = importlib.util.module_from_spec(spec2)
spec2.loader.exec_module(audit)

EXCELS = {n: (lbl, conv) for n, lbl, conv in audit.EXCELS}


def md_inline(s):
    if not s:
        return '∅'
    return str(s).replace('`', '\\`').replace('|', '\\|').replace('\n', ' ⏎ ')


def read_cell(wb_path, sheet_name, row_num, convention):
    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        # Try prefix match
        match = next((s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if match is None:
            wb.close()
            return None
        sheet_name = match
    ws = wb[sheet_name]
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if r == row_num:
            if len(row) < 10:
                wb.close()
                return None
            speaker = audit.get_speaker(row, convention)
            en = (row[2] or '').strip() if row[2] is not None else ''
            nl = (row[9] or '').strip() if row[9] is not None else ''
            wb.close()
            return {'speaker': speaker, 'en': en, 'nl': nl, 'sheet': sheet_name}
    wb.close()
    return None


def diff_cells(local, remote):
    """Return list of fields that diverge between local and remote."""
    if local is None and remote is None:
        return ['BOTH-MISSING']
    if local is None:
        return ['LOCAL-MISSING']
    if remote is None:
        return ['REMOTE-MISSING']
    diffs = []
    for k in ('speaker', 'en', 'nl'):
        if (local.get(k) or '').strip() != (remote.get(k) or '').strip():
            diffs.append(k)
    return diffs


def write_episode_diff(ep, entries, local_wb_path, remote_wb_path, ep_label, convention, out_path):
    L = [f'# Local-vs-remote diff — E{ep} DRIFT cells']
    L.append('')
    L.append(f'_Generated: {RESCAN_DATE}_')
    L.append(f'_Local:  `{local_wb_path.relative_to(REPO)}`_')
    L.append(f'_Remote: `{remote_wb_path.relative_to(REPO)}` (fresh Google-Sheets pull)_')
    L.append('')

    in_sync = 0
    diverged = 0
    drift_resolved_in_remote = 0

    by_sheet = {}
    for e in entries:
        by_sheet.setdefault(e['sheet'], []).append(e)
    by_sheet = dict(sorted(by_sheet.items()))

    details = []

    for claimed_sheet, sheet_entries in by_sheet.items():
        section = []
        section.append(f'## {claimed_sheet}')
        section.append('')
        for entry in sheet_entries:
            r = entry['row']
            local = read_cell(local_wb_path, claimed_sheet, r, convention)
            remote = read_cell(remote_wb_path, claimed_sheet, r, convention)
            diffs = diff_cells(local, remote)
            section.append(f'### J{r}')
            section.append('')
            section.append(f'- **Claim:** {md_inline(entry["claim_line"])}')
            if not diffs:
                section.append(f'- **Status:** ✓ IN-SYNC — local and remote NL match')
                section.append(f'- **EN:** `{md_inline(local["en"])}`')
                section.append(f'- **NL (both):** `{md_inline(local["nl"])}`')
                in_sync += 1
            else:
                diverged += 1
                section.append(f'- **Status:** ⚠ DIVERGED on field(s): `{", ".join(diffs)}`')
                if local:
                    section.append(f'- **Local EN:**  `{md_inline(local["en"])}`')
                    section.append(f'- **Local NL:**  `{md_inline(local["nl"])}`')
                else:
                    section.append(f'- **Local:** ⚠ row not found')
                if remote:
                    section.append(f'- **Remote EN:** `{md_inline(remote["en"])}`')
                    section.append(f'- **Remote NL:** `{md_inline(remote["nl"])}`')
                else:
                    section.append(f'- **Remote:** ⚠ row not found')
                # Detect "drift resolved" — if local NL matches claim's drifted text
                # but remote NL is different, Patrick may have already fixed it.
                if local and remote and local.get('nl', '') != remote.get('nl', ''):
                    drift_resolved_in_remote += 1
                    section.append(f'  → Remote NL differs — Patrick may have already touched this cell.')
            section.append('')
        details.append(('\n'.join(section), len(sheet_entries)))

    # Summary at top
    L.append('## Summary')
    L.append('')
    L.append(f'- Cells checked: **{len(entries)}**')
    L.append(f'- IN-SYNC (local = remote): **{in_sync}**')
    L.append(f'- DIVERGED (local ≠ remote): **{diverged}**')
    L.append(f'- Cells where remote NL differs from local (Patrick re-edited?): **{drift_resolved_in_remote}**')
    L.append('')

    for section, _ in details:
        L.append(section)

    out_path.write_text('\n'.join(L))
    return {'in_sync': in_sync, 'diverged': diverged, 'drift_resolved': drift_resolved_in_remote}


def main():
    remote_dir = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_REMOTE_DIR
    remote_root = REPO / remote_dir
    if not remote_root.exists():
        print(f'!! remote dir not found: {remote_root}', file=sys.stderr)
        sys.exit(1)

    print(f'Parsing deep-eyeball doc for DRIFT cells…')
    entries = rescan.parse_deep_eyeball()
    print(f'  Found {len(entries)} DRIFT cells.')
    print(f'Comparing local vs remote ({remote_dir})…')

    by_ep = {}
    for e in entries:
        by_ep.setdefault(e['episode'], []).append(e)

    totals = {'in_sync': 0, 'diverged': 0, 'drift_resolved': 0, 'episodes': []}
    for ep in sorted(by_ep.keys()):
        ep_label, conv = EXCELS[ep]
        local_xlsx = REPO / 'excels' / f'{ep}_asses.masses_{ep_label}.xlsx'
        remote_xlsx = remote_root / f'{ep}_asses.masses_{ep_label}.xlsx'
        out_path = OUT_DIR / f'rescan-{RESCAN_DATE}-E{ep}-local-vs-remote.md'
        stats = write_episode_diff(ep, by_ep[ep], local_xlsx, remote_xlsx, ep_label, conv, out_path)
        totals['in_sync'] += stats['in_sync']
        totals['diverged'] += stats['diverged']
        totals['drift_resolved'] += stats['drift_resolved']
        totals['episodes'].append({
            'ep': ep, 'label': ep_label, 'count': len(by_ep[ep]),
            'in_sync': stats['in_sync'], 'diverged': stats['diverged'],
            'drift_resolved': stats['drift_resolved'],
        })
        print(f'  E{ep}: {len(by_ep[ep])} cells → {stats["in_sync"]} in-sync, {stats["diverged"]} diverged, {stats["drift_resolved"]} remote re-edited')

    # INDEX
    L = [f'# Local-vs-remote DRIFT diff — {RESCAN_DATE}']
    L.append('')
    L.append(f'_Local: `excels/`_')
    L.append(f'_Remote: `{remote_dir}/` (fresh Google-Sheets pull)_')
    L.append('')
    L.append('## Totals')
    L.append('')
    L.append(f'- Total DRIFT cells diffed: **{sum(e["count"] for e in totals["episodes"])}**')
    L.append(f'- IN-SYNC: **{totals["in_sync"]}** (local matches remote)')
    L.append(f'- DIVERGED: **{totals["diverged"]}** (local differs from remote)')
    L.append(f'- Cells where remote NL differs from local: **{totals["drift_resolved"]}** (Patrick may have re-edited)')
    L.append('')
    L.append('## Per-episode breakdown')
    L.append('')
    L.append('| Episode | Cells | In-sync | Diverged | Remote re-edited | Report |')
    L.append('|---|---|---|---|---|---|')
    for e in totals['episodes']:
        link = f'rescan-{RESCAN_DATE}-E{e["ep"]}-local-vs-remote.md'
        L.append(f'| E{e["ep"]} | {e["count"]} | {e["in_sync"]} | {e["diverged"]} | {e["drift_resolved"]} | [{link}]({link}) |')
    L.append('')
    L.append('## Interpretation')
    L.append('')
    L.append('- **IN-SYNC** cells: local NL = remote NL. The DRIFT claim is still active on the live remote — fix is needed.')
    L.append('- **DIVERGED** cells: local NL ≠ remote NL. Either the DRIFT was already fixed remotely (Patrick or a prior push), OR local has changes not yet pushed. Inspect before applying any fix.')
    L.append('- **Remote re-edited**: subset of DIVERGED where the NL string itself differs — strong signal that Patrick has touched the cell since last sync.')
    L.append('')

    (OUT_DIR / f'rescan-{RESCAN_DATE}-local-vs-remote-INDEX.md').write_text('\n'.join(L))
    print(f'\n  Wrote: rescan-{RESCAN_DATE}-local-vs-remote-INDEX.md')
    print(f'  Wrote: 11 per-episode files')
    print(f'\nTotals: {sum(e["count"] for e in totals["episodes"])} cells | {totals["in_sync"]} in-sync | {totals["diverged"]} diverged | {totals["drift_resolved"]} remote re-edited')


if __name__ == '__main__':
    main()
