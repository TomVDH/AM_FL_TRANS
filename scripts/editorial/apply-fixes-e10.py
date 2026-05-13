#!/usr/bin/env python3
"""Apply user-approved E10 fixes to local xlsx (then push via push-file.py).

Tom sign-off 2026-05-13:
  - J47/J68/J100/J110/J154 Government: fix as proposed (mechanical)
  - J56 ProphetSpeech: Tom flagged ADDITION canon-violation — NL adds
    `, kameraaden` vocative with no EN equivalent. Drop the tail entirely
    rather than just fix the double-a typo.
  - J81 ProphetSpeech: OVERRIDE → `BUITEN MET ONZE LEIDERS!` (Tom's exact
    form; chant-rhythm variant)
  - J110 ProphetSpeech: Tom's exact form `WEES NIET BANG` (no `!`)
  - J112 ProphetSpeech: KEPT (speaker-differentiated tough-guy register OK)

8 cells written, 1 KEPT.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/10_asses.masses_E10Proxy.xlsx'

APPROVED = {
    # E10_Government (5 cells)
    ('E10_Government_localization', 47): (
        '"Heeft iemand zin om STENEN-SPEL te spelen?"',
        '"Heeft iemand zin om KEIEN-SPEL te spelen?"',
    ),
    ('E10_Government_localization', 68): (
        'Lieve Ezel? Natuurlijk dat die idioten van Muilegem hier zouden zijn...',
        'Lieve Ezel? Natuurlijk dat die idioten van Muilenbeek hier zouden zijn...',
    ),
    ('E10_Government_localization', 100): (
        'Zeg de Mensen dat ze de grenzen van ons territorium in Muilegem moeten respecteren.',
        'Zeg de Mensen dat ze de grenzen van ons territorium in Muilenbeek moeten respecteren.',
    ),
    ('E10_Government_localization', 110): (
        'Zeg tegen de Mensen dat wij onze eigen zaakjes willen runnen, zoals het Poepegaatje, héé!',
        'Zeg tegen de Mensen dat wij onze eigen zaakjes willen runnen, zoals De Zatten Ezel, héé!',
    ),
    ('E10_Government_localization', 154): (
        'Als de mensen ons geen grondgebied afstaan, zullen we hun kinderen doden.',
        'Als de Mensen ons geen grondgebied afstaan, zullen we hun kinderen doden.',
    ),

    # E10_ProphetSpeech (3 cells; J112 KEPT)
    # J56: Tom flagged canon violation — NL adds `, kameraaden` not in EN. Drop.
    ('E10_ProphetSpeech_localization', 56): (
        'Toch zijn wij hier allemaal, zowel Mensen als Ezels, rationele wezens, kameraaden.',
        'Toch zijn wij hier allemaal, zowel Mensen als Ezels, rationele wezens.',
    ),
    # J81: Tom override — chant-rhythm variant
    ('E10_ProphetSpeech_localization', 81): (
        'BUITENGEGOOID MET ONZE LEIDERS—',
        'BUITEN MET ONZE LEIDERS!',
    ),
    # J110: Tom's exact form (no `!`)
    ('E10_ProphetSpeech_localization', 110): (
        'WEES NIET BANG!',
        'WEES NIET BANG',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    if discrepancies:
        for s, r, msg in discrepancies:
            print(f'  ⚠ {s}/J{r}: {msg}')
        sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
