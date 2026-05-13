#!/usr/bin/env python3
"""Apply slogan-retcon batch (Herinneer→Herinner + caps fixes) to E1 cells.

Per Tom 2026-05-13 canon decision:
  - Herinneer (legacy/non-standard) → Herinner (modern Dutch imperative)
  - In long-form dying-words quote: cap Wereld and Belangen
  - Case mirrors EN intent (ALL-CAPS stays ALL-CAPS; mixed stays mixed)
  - Terminal ! mirrors EN

4 cells in this batch, all in E1Proxy:
  - E1_Stable2F   J34   ALL-CAPS  HERINNEER → HERINNER
  - E1_TheProtest J67   mixed     Herinneer→Herinner + wereld→Wereld + belangan→Belangen
  - E1_TheProtest J69   mixed     same as J67
  - E1_TheProtest J115  ALL-CAPS  HERINNEER → HERINNER + add terminal !

E4_AstralPlaneMain J71 already at canonical form from E4 Push 2 — not in batch.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/1_asses.masses_E1Proxy.xlsx'

APPROVED = {
    ('E1_Stable2F_localization', 34): (
        'HERINNEER DE WERELD AAN DE BELANGEN VAN DE EZELS!',
        'HERINNER DE WERELD AAN DE BELANGEN VAN DE EZELS!',
    ),
    ('E1_TheProtest_localization', 67): (
        '"Herinneer de wereld aan de belangan van de Ezels" waren Oude Ezels laatste woorden.',
        '"Herinner de Wereld aan de Belangen van de Ezels" waren Oude Ezels laatste woorden.',
    ),
    ('E1_TheProtest_localization', 69): (
        '"Herinneer de wereld aan de belangan van de Ezels" waren Oude Ezels laatste woorden.',
        '"Herinner de Wereld aan de Belangen van de Ezels" waren Oude Ezels laatste woorden.',
    ),
    ('E1_TheProtest_localization', 115): (
        'HERINNEER DE WERELD AAN DE BELANGEN VAN DE EZELS',
        'HERINNER DE WERELD AAN DE BELANGEN VAN DE EZELS!',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    if discrepancies:
        for s, r, msg in discrepancies:
            print(f'  ⚠ {s}/J{r}: {msg}')
        sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
