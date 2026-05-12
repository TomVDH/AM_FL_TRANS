#!/usr/bin/env python3
"""E9 canon sweep scanner — carries E8 ruleset unchanged.

E9 context (per §19 / §5.2):
- EpisodeTitle: short label cells
- Captcha: input/system cells
- BadCave: cave sequence
- MineEscape: action sequence
- GoldenAss: God-register speaker — 4 cells expected with lowercase u/uw → cap U/Uw

All other patterns identical to E8 scanner.
"""
import re
import sys
from pathlib import Path
import openpyxl

XLSX = Path('excels.fresh-pull-2026-05-11/9_asses.masses_E9Proxy.xlsx')
DIARY_SHEETS = set()

RETIRED_MONIKERS = ['Felle Gast', 'Betweter', 'Bikkeharde', 'Snotezel', 'Stampgast', 'Sloom']

MACHINE_COMPOUNDS_UNHYPHENATED = [
    r'\bBoormachine[ns]?\b',
    r'\bCamion[ ]Machine[ns]?\b',
    r'\bTractor[ ]Machine[ns]?\b',
    r'\bAuto[ ]Machine[ns]?\b',
    r'\bMysterie[ ]Machine[ns]?\b',
    r'\bTank[ ]Machine[ns]?\b',
    r'\bVlieg[ ]Machine[ns]?\b',
]


def extract_speaker(key):
    if not key:
        return ''
    parts = str(key).split('.')
    if len(parts) >= 2:
        return parts[-1].strip()
    return str(key).strip()


def scan_sheet(ws, sheet_name):
    findings = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) < 10:
            continue
        key = (row[0] or '')
        speaker = extract_speaker(key)
        en = (row[2] or '').strip()
        nl = (row[9] or '').strip()
        if not nl:
            continue

        def add(rule, note):
            findings.append((r, rule, note, speaker, en, nl))

        # §2 nie (standalone)
        if re.search(r"\bnie\b", nl, flags=re.IGNORECASE) or re.search(r"\bnie'", nl):
            add('§2', 'nie standalone → niet')

        # §7.1 Ezel cap — lowercase 'ezel' mid-sentence
        for m in re.finditer(r"\bezel(s|innekes|tje|tjes|en|kes)?\b", nl):
            if m.start() == 0:
                continue
            prev = nl[max(0, m.start()-2):m.start()]
            if re.search(r"[.!?]\s$", prev + ' '):
                continue
            add('§7.1', f'lowercase "{m.group(0)}" mid-sentence — cap?')
            break

        # §7.2 Circusdirecteur/Tournee — cap everywhere; flag lowercase
        if re.search(r"\bcircusdirecteur\b", nl):
            add('§7.2', 'lowercase circusdirecteur — should be Circusdirecteur (cap-always per §7.2)')
        if re.search(r"\btournee\b", nl):
            add('§7.2', 'lowercase tournee — should be Tournee (cap-always per §7.2)')

        # §3.6 Boerderij
        if re.search(r"\bboerderij\b", nl, flags=re.IGNORECASE):
            add('§3.6', 'Boerderij → Hoeve')

        # §3.1 Muilebeek
        if re.search(r"\bMuilebeek\b", nl):
            add('§3.1', 'Muilebeek → Muilenbeek')

        # §3.4 Mecha/Mechalen → Technopolis (§3.4.1 Imechelda exception)
        if re.search(r"\bMecha(len)?\b", nl):
            if 'Ziekenhuis' in nl:
                add('§3.4.1', 'Mecha + Ziekenhuis → Imechelda Algemeen Ziekenhuis (exception)')
            else:
                add('§3.4', 'Mecha/Mechalen → Technopolis')
        if 'MECHA' in nl and 'MECHALEN' not in nl:
            add('§3.4', 'MECHA → TECHNOPOLIS (all-caps)')

        # §6.16 Beroep/Beroepen
        if re.search(r"\bBeroep(en)?\b", nl):
            if 'beroep doen op' not in nl.lower():
                add('§6.16', 'Beroep(en) → Job(s) — verify game-system context')

        # §6.16 lowercase job/jobs
        if re.search(r"\bjob[s]?\b", nl) and not re.search(r"\bJob[s]?\b", nl):
            add('§6.16-lc', 'lowercase job/jobs — game-system? cap to Job/Jobs')

        # §6.1 HUDO / Hudo / Sekreet / Privaat
        if re.search(r"\b[Hh]udo\b", nl) or 'HUDO' in nl:
            add('§6.1', 'Hudo → Plee')
        if re.search(r"\bSekreet\b|\bPrivaat\b", nl):
            add('§6.1', 'Sekreet/Privaat → Plee')

        # §6.2 Oom (non-Smart Ass)
        if re.search(r"\bOom\b", nl) and 'Smart' not in speaker:
            add('§6.2', f'Oom (speaker={speaker!r}) → Nonkel?')

        # §6.10 Acte/Actes/Acten
        if re.search(r"\b[Aa]cte[sn]?\b", nl):
            add('§6.10', 'Acte/Actes/Acten → Nummer/Nummers')

        # §6.12 Nijg
        if re.search(r"\b[Nn]ijg(t|en|de)?\b", nl):
            add('§6.12', 'Nijg → fel')

        # §10.1 doekjes rond winden
        if 'doekjes rond winden' in nl.lower():
            add('§10.1', "doekjes rond winden → doekjes om winden")

        # §18 Wereldtournee — cap-always
        if re.search(r"\bwereldtournee\b", nl):
            add('§18', 'lowercase wereldtournee — should be Wereldtournee (cap-always)')

        # §6.11 jansen — extinct
        if re.search(r"\bjansen\b", nl, flags=re.IGNORECASE):
            add('§6.11', f'jansen found (speaker={speaker!r}) — EXTINCT')

        # §7.4 / §5.2 Gods — lowercase u/uw from God-register speakers
        god_speakers = {'Haw', 'Hee', 'Golden Ass', 'THE GODS'}
        if any(g in speaker for g in god_speakers):
            if re.search(r"\bu\b|\buw\b", nl):
                add('§7.4', f'God speaker {speaker!r} uses lowercase u/uw — should be U/Uw')

        # §10.7 slokkie/slokje → slokske
        if re.search(r"\bslokk?[ij]e\b", nl, flags=re.IGNORECASE):
            add('§10.7', 'slokkie/slokje → slokske (Flemish diminutive)')

        # §12.2 Sturdy motto fragment
        if 'Sturdy' in speaker:
            if re.search(r"\bwerk[- ]afpakkende\b|\bkind[- ]dodende\b|\bzielloze\b|\bslechte\b", nl, flags=re.IGNORECASE):
                add('§12.2', 'Sturdy motto fragment — verify full canonical form')

        # §8.1 Machien
        if re.search(r"\bMachien(en)?\b", nl):
            add('§8.1', 'Machien → Machine')

        # §8.2 Machine compounds — unhyphenated
        for pat in MACHINE_COMPOUNDS_UNHYPHENATED:
            if re.search(pat, nl):
                add('§8.2', f'machine compound needs hyphen ({pat})')

        # §4.3 retired monikers
        for mon in RETIRED_MONIKERS:
            if mon in nl:
                add('§4.3', f'retired moniker {mon}')

        # §4.4 Schoon Beest
        if 'Schoon Beest' in nl:
            add('§4.4', f'Schoon Beest (speaker={speaker!r}) — preserve only if Thirsty→Nice')

        # §12.1 , Kameraad
        if re.search(r",\s*Kameraad\b", nl):
            if 'Comrade' not in en:
                add('§12.1', ', Kameraad without EN Comrade — strip?')

        # §14.1 retired slogans
        for slog in ['EZELSKRACHT', 'EZEL MACHT', 'EZELKRACHT', 'EZELS-KRACHT']:
            if slog in nl:
                add('§14.1', f'retired slogan {slog} → EZELS EERST')

        # §5.4 Stop imperative
        if re.search(r"(^|[.!?]\s+)Stop\b", nl):
            add('§5.4', 'Stop imperative — verify register')

        # §9.1 sentence-start apostrophe missing
        if re.search(r"(^|[.!?]\s+)k\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: k → 'k")
        if re.search(r"(^|[.!?]\s+)t\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: t → 't")

        # §9.2 mid-sentence Ik [Cap]
        if re.search(r",\s+Ik\s+[A-Z]", nl):
            add('§9.2', "mid-sentence ', Ik [Cap]' — lowercase verb")

    return findings


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    target_sheets = sys.argv[1:] or wb.sheetnames
    total = 0
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            matches = [s for s in wb.sheetnames if s.startswith(sheet_name)]
            if not matches:
                print(f'!! sheet not found: {sheet_name}')
                continue
            sheet_name = matches[0]
        ws = wb[sheet_name]
        findings = scan_sheet(ws, sheet_name)
        total += len(findings)
        print(f'\n=== {sheet_name}: {len(findings)} finding(s) ===')
        for r, rule, note, speaker, en, nl in findings:
            print(f'  J{r} [{rule}] {note}')
            print(f'    Speaker: {speaker!r}')
            print(f'    EN: {en!r}')
            print(f'    NL: {nl!r}')
    print(f'\n=== TOTAL: {total} finding(s) across {len(target_sheets)} sheet(s) ===')
    wb.close()


if __name__ == '__main__':
    main()
