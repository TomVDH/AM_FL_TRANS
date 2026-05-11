#!/usr/bin/env python3
"""
three-way-diff.py — 3-way reconcile diff over column-J NL cells.

Compares three xlsx tree-roots:
  baseline  = excels/Originals/   (last shared sync, 2026-05-03)
  ours      = excels/             (local edits this cycle)
  theirs    = excels/Patrick-*/   (fresh GS pull)

Per (file, sheet, J-row):
  ours_only    : baseline == theirs, ours differs        → ours wins, push later
  theirs_only  : baseline == ours,   theirs differs      → pull theirs into local
  agreed       : ours == theirs (both differ from baseline OR both equal it; converged)
  conflicts    : ours, theirs, baseline all differ from each other → manual resolution

Outputs JSON to data/editorial/reconcile-<date>.json (or path passed as 2nd arg).

USAGE
  py scripts/editorial/three-way-diff.py <theirs-dir> [<out-json>]
  py scripts/editorial/three-way-diff.py excels/Patrick-2026-05-08 data/editorial/reconcile-2026-05-08.json
"""
import json
import sys
from pathlib import Path
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
BASELINE_DIR = PROJECT_ROOT / 'excels' / 'Originals'
OURS_DIR = PROJECT_ROOT / 'excels'

FILES = [
    '0_asses.masses_Manager+Intermissions+E0Proxy.xlsx',
    '1_asses.masses_E1Proxy.xlsx',
    '2_asses.masses_E2Proxy.xlsx',
    '3_asses.masses_E3Proxy.xlsx',
    '4_asses.masses_E4Proxy.xlsx',
    '5_asses.masses_E5Proxy.xlsx',
    '6_asses.masses_E6Proxy.xlsx',
    '7_asses.masses_E7Proxy.xlsx',
    '8_asses.masses_E8Proxy.xlsx',
    '9_asses.masses_E9Proxy.xlsx',
    '10_asses.masses_E10Proxy.xlsx',
]


def norm(v):
    return '' if v is None else str(v)


def load_j_col(xlsx_path):
    """Return {sheet_name: {row_index: nl_value}} for column J across all sheets."""
    out = {}
    if not xlsx_path.exists():
        return out
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = {}
        max_r = ws.max_row or 0
        for r in range(1, max_r + 1):
            v = ws.cell(r, 10).value  # column J
            if v is not None:
                rows[r] = norm(v)
        out[sn] = rows
    wb.close()
    return out


def load_en_col(xlsx_path):
    """Return {sheet_name: {row_index: en_value}} for column C (EN/Standard)."""
    out = {}
    if not xlsx_path.exists():
        return out
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = {}
        max_r = ws.max_row or 0
        for r in range(1, max_r + 1):
            v = ws.cell(r, 3).value
            if v is not None:
                rows[r] = norm(v)
        out[sn] = rows
    wb.close()
    return out


def col_letter_row(idx):
    return f'J{idx}'


def main():
    if len(sys.argv) < 2:
        print('Usage: three-way-diff.py <theirs-dir> [<out-json>]')
        sys.exit(2)
    theirs_dir = Path(sys.argv[1]).resolve()
    if not theirs_dir.is_absolute():
        theirs_dir = (PROJECT_ROOT / sys.argv[1]).resolve()
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else PROJECT_ROOT / 'data/editorial/reconcile.json'
    if not out_path.is_absolute():
        out_path = (PROJECT_ROOT / sys.argv[2]).resolve()

    ours_only, theirs_only, agreed, conflicts = [], [], [], []
    sheets_missing_in_theirs = []
    sheets_missing_in_baseline = []

    summary = {}

    for fname in FILES:
        baseline = load_j_col(BASELINE_DIR / fname)
        ours = load_j_col(OURS_DIR / fname)
        theirs = load_j_col(theirs_dir / fname)
        en_local = load_en_col(OURS_DIR / fname)

        file_counts = {'ours_only': 0, 'theirs_only': 0, 'agreed': 0, 'conflicts': 0}

        # Sheet-level coverage check
        for sn in ours:
            if sn not in theirs:
                sheets_missing_in_theirs.append((fname, sn))

        # Iterate union of all sheets present in any source
        all_sheets = set(baseline) | set(ours) | set(theirs)
        for sn in all_sheets:
            b_rows = baseline.get(sn, {})
            o_rows = ours.get(sn, {})
            t_rows = theirs.get(sn, {})
            en_rows = en_local.get(sn, {})
            all_rows = set(b_rows) | set(o_rows) | set(t_rows)
            for r in all_rows:
                b = b_rows.get(r, '')
                o = o_rows.get(r, '')
                t = t_rows.get(r, '')
                en = en_rows.get(r, '')

                # Skip header (row 1) — it's "NL" or similar metadata
                if r == 1:
                    continue

                if o == t:
                    if o == b:
                        # All three equal → no diff at all; don't emit
                        continue
                    # ours == theirs but != baseline → both made same edit
                    agreed.append({
                        'file': fname, 'sheet': sn, 'cell': col_letter_row(r),
                        'english': en, 'baseline': b, 'value': o,
                    })
                    file_counts['agreed'] += 1
                else:
                    if o == b and t != b:
                        # Patrick changed it; we didn't
                        theirs_only.append({
                            'file': fname, 'sheet': sn, 'cell': col_letter_row(r),
                            'english': en, 'baseline': b, 'theirs': t,
                        })
                        file_counts['theirs_only'] += 1
                    elif t == b and o != b:
                        # We changed it; Patrick didn't
                        ours_only.append({
                            'file': fname, 'sheet': sn, 'cell': col_letter_row(r),
                            'english': en, 'baseline': b, 'ours': o,
                        })
                        file_counts['ours_only'] += 1
                    else:
                        # Both differ from baseline AND from each other → conflict
                        conflicts.append({
                            'file': fname, 'sheet': sn, 'cell': col_letter_row(r),
                            'english': en, 'baseline': b, 'ours': o, 'theirs': t,
                        })
                        file_counts['conflicts'] += 1

        summary[fname] = file_counts

    # Stable sort for diff-friendly output
    sort_key = lambda d: (d['file'], d['sheet'], int(d['cell'][1:]))
    ours_only.sort(key=sort_key)
    theirs_only.sort(key=sort_key)
    agreed.sort(key=sort_key)
    conflicts.sort(key=sort_key)

    result = {
        '_meta': {
            'baseline_dir': str(BASELINE_DIR.relative_to(PROJECT_ROOT)),
            'ours_dir': str(OURS_DIR.relative_to(PROJECT_ROOT)),
            'theirs_dir': str(theirs_dir.relative_to(PROJECT_ROOT)) if theirs_dir.is_relative_to(PROJECT_ROOT) else str(theirs_dir),
            'totals': {
                'ours_only': len(ours_only),
                'theirs_only': len(theirs_only),
                'agreed': len(agreed),
                'conflicts': len(conflicts),
            },
            'per_file': summary,
            'sheets_missing_in_theirs': [{'file': f, 'sheet': s} for f, s in sheets_missing_in_theirs],
        },
        'ours_only': ours_only,
        'theirs_only': theirs_only,
        'agreed': agreed,
        'conflicts': conflicts,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')

    print(f'Reconcile written: {out_path.relative_to(PROJECT_ROOT)}')
    print(f'  ours_only  : {len(ours_only):4d}')
    print(f'  theirs_only: {len(theirs_only):4d}')
    print(f'  agreed     : {len(agreed):4d}')
    print(f'  conflicts  : {len(conflicts):4d}')
    if sheets_missing_in_theirs:
        print(f'  WARNING: {len(sheets_missing_in_theirs)} sheet(s) present locally but missing in theirs')


if __name__ == '__main__':
    main()
