#!/usr/bin/env python3
"""Apply E4_AstralPlaneMain blind-spot walk fixes — Tom sign-off 2026-05-13.

17 writes / 6 keeps.

Notable Tom decisions:
  - DJ Dope Ass cluster: J207 KEPT ge/gij (`Hebt gij`); J214 je/jij rewrite
    (`Doe me … ga fixen voor mij`); J208/J210/J215 KEPT ge/gij as-is
    — Tom mixes register per-cell rather than blanket-applying canon je/jij.
  - J138 Chafed Ass: Tom present-tense rewrite (`afmaken, villen, maken`
    not proposed past-tense `afgemaakt, gevild, eindigen`).
  - J181 KEPT (`rusting chillen` stays).
  - J140 KEEPS present tense `geven` (only swaps Anglicism
    `representatie` → `vertegenwoordiging`).
  - J142 Tom rewrite: `onder ons` + plural `rijke innerlijke levens`.
  - J179 Tom adds `TOT`, drops trailing `!`.
  - J84 Steppes KEPT.
  - J123 MZK acronym KEPT.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/4_asses.masses_E4Proxy.xlsx'

APPROVED = {
    # DJ Dope Ass cluster (2 writes, 3 keeps)
    ('E4_AstralPlaneMain_localization', 207): (
        'Hebde gij misschien wat van dat Paarse Spul...?',
        'Hebt gij misschien wat van dat Paarse Spul...?',
    ),
    ('E4_AstralPlaneMain_localization', 214): (
        'Doet mij een plezier en fix mij wat, oké?',
        'Doe me een plezier en ga fixen voor mij, oké?',
    ),

    # §13 BUGS (3 writes; J181 KEPT)
    ('E4_AstralPlaneMain_localization', 138): (
        'Nihao. Ik haat Mensen omdat ik mijn hele leven voor heb heb gewerkt, voordat ze me afmaakten en van mij een delicatesse hebben bereid.',
        "Nihao. Ik haat Mensen omdat ik mijn hele leven voor ze heb gewerkt, voordat ze me afmaken, villen en van mij een 'delicatesse' maken.",
    ),
    ('E4_AstralPlaneMain_localization', 125): (
        'Ik haat Mensen omdat ze met een stalen pijp bijeen sloegen!',
        'Ik haat Mensen omdat ze mij met een stalen pijp in elkaar geslagen hebben!',
    ),
    ('E4_AstralPlaneMain_localization', 204): (
        "Buiten, natuurlijk, DJ Lijpe Ezel's Astrale Trap Concert.",
        'Behalve, natuurlijk, voor het Astrale Trap Concert van DJ Lijpe Ezel.',
    ),

    # §13 semantic shifts (9 writes)
    ('E4_AstralPlaneMain_localization', 51): (
        'Ik wist dat ze een gat in haar hand had!',
        'Ik wist dat zij een gokker was!',
    ),
    ('E4_AstralPlaneMain_localization', 140): (
        'UGH! Ik haat Mensen omdat ze ons nooit de representatie in de media geven die we verdienden.',
        'UGH! Ik haat Mensen omdat ze ons nooit de vertegenwoordiging in de media geven die we verdienden.',
    ),
    ('E4_AstralPlaneMain_localization', 142): (
        'De meeste van ons zijn hele mensen met rijke innerlijke belevenissen.',
        'De meesten onder ons zijn complexe individuen met rijke innerlijke levens.',
    ),
    ('E4_AstralPlaneMain_localization', 156): (
        'Sommige wouden betere levensomstandigheden.',
        'Sommigen wilden betere levensomstandigheden.',
    ),
    ('E4_AstralPlaneMain_localization', 159): (
        'Ik heb een paar spontane opwellingen meegemaakt in de jaren 1300. We hadden bijna de regering van de troon gestoten.',
        'Ik heb een paar spontane opstanden geleid in de jaren 1300. We hadden bijna de regering van de troon gestoten.',
    ),
    ('E4_AstralPlaneMain_localization', 160): (
        'Er was niet veel nodig om die Boeren te laten betogen.',
        'Er was niet veel nodig om die Boeren in opstand te laten komen.',
    ),
    ('E4_AstralPlaneMain_localization', 163): (
        'Anderen zeiden dat mijn Hongerstaking generatie op generatie heeft gëinspireerd.',
        'Anderen hier zeggen dat mijn Hongerstaking generaties geïnspireerd heeft.',
    ),
    ('E4_AstralPlaneMain_localization', 179): (
        'NU KAN IK MIJN BEENDEREN VOELEN!',
        'NU VOEL IK HET TOT IN MIJN BOTTEN',
    ),
    ('E4_AstralPlaneMain_localization', 231): (
        "Dan wat sta je hier te doen? Wat met DJ Lijpe Ezel's Concert?!",
        "Wat sta je hier dan te doen? En DJ Lijpe Ezel's Concert dan?!",
    ),

    # §12.4 English bleed (1 write; J84 KEPT)
    ('E4_AstralPlaneMain_localization', 164): (
        'Mijn Kudde heeft een Boycott tegen de lange reizen op de Zijderoute gepleegd.',
        'Mijn Kudde heeft een Boycot gepleegd tegen de lange reizen op de Zijderoute.',
    ),

    # §9.3 terminal punct (2 writes)
    ('E4_AstralPlaneMain_localization', 153): (
        'ACH! Protesten... lijken meestal op een goed idee maar ze werken MEESTAL NIET.',
        'ACHH! Protesten... lijken op een goed idee maar ze WERKEN NIET.',
    ),
    ('E4_AstralPlaneMain_localization', 169): (
        'Onze liefde voor vrede werd ons niet in dank afgenomen.',
        'Onze liefde voor vrede werd ons niet in dank afgenomen...',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    for s, r, msg in discrepancies:
        print(f'  ⚠ {s}/J{r}: {msg}')
    if discrepancies: sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
