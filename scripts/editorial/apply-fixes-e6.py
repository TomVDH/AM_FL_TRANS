#!/usr/bin/env python3
"""Apply user-approved E6 fixes to local xlsx (then push via push-file.py).

Tom sign-off 2026-05-13:
  - E6_BadCave J24/J38/J39: fix as proposed
  - E6_BadCave J40: KEPT (already capped — no change)
  - E6_BattleHard J18: OVERRIDE flip → Verduren → Volharden (unify Endure on Volharden)
  - E6_BattleHard J20: fix → Wonden likken → Wonden Likken (title case match J38)
  - E6_BattleHard J33: KEPT (already Volharden — Tom's pick)
  - E6_BattleHard J38: KEPT (already Wonden Likken)
  - E6_Nightmare J3: fix as proposed
  - E6_World J75: OVERRIDE → `'t IS AL GOED MOEDER` (Tom exact quote, no `!`)
  - E6_World J188: fix as proposed
  - E6_World J196: OVERRIDE → `Café De Zatten Ezel` (move Café before, venue compound name)
  - E6_World J197: OVERRIDE → drop `verdomse` + `DROOM`-caps; keep `gedoeme` + original word order
  - E6_World J206: OVERRIDE → `godver`→`ja`; add `doe` after `open`; CAFÉ accent
  - E6_World J209: KEPT (Flemish dialect ✓ — recommended skip)
  - E6_World J260/J292: fix as proposed

J75 corpus-checked: unique line, no `FINE MOTHER`/`GESTELD MOEDER` repeats elsewhere.

13 cells written, 4 KEPT.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/6_asses.masses_E6Proxy.xlsx'

APPROVED = {
    # E6_BadCave (3 cells; J40 KEPT)
    ('E6_BadCave_localization', 24): (
        "Ze denkt nog steeds dat we hier elke dag 'Stenen-spel' aan het spelen zijn...",
        "Ze denkt nog steeds dat we hier elke dag 'Keien-Spel' aan het spelen zijn...",
    ),
    ('E6_BadCave_localization', 38): (
        'Een plek voor de mensen.',
        'Een plek voor de Mensen.',
    ),
    ('E6_BadCave_localization', 39): (
        'Intelligente mensen.',
        'Intelligente Mensen.',
    ),

    # E6_BattleHard (2 writes; J33/J38 KEPT — already at target values)
    # Tom override: unify Endure on `Volharden` (J18 flip + J33 keep)
    ('E6_BattleHard_localization', 18): (
        'Verduren',
        'Volharden',
    ),
    ('E6_BattleHard_localization', 20): (
        'Wonden likken',
        'Wonden Likken',
    ),

    # E6_Nightmare (1 cell)
    ('E6_Nightmare_localization', 3): (
        'Er word niet op het werk geslapen, nutteloze Ezel!',
        'Er wordt niet op het werk geslapen, nutteloze Ezel!',
    ),

    # E6_World (7 writes; J209 KEPT)
    ('E6_World_localization', 75): (
        'Ik ben GESTELD MOEDER!',
        "'t IS AL GOED MOEDER",
    ),
    ('E6_World_localization', 188): (
        'Ik probeer nog wat mooi tijd door te brengen met {$NewName}, maar ze groeit veel te snel op.',
        'Ik probeer nog wat mooie tijd door te brengen met {$NewName}, maar ze groeit veel te snel op.',
    ),
    ('E6_World_localization', 196): (
        'Vanavond gaat De Zatten Ezel Cafe eindelijk open, héé!',
        'Vanavond gaat Café De Zatten Ezel eindelijk open, héé!',
    ),
    ('E6_World_localization', 197): (
        "'t Is altijd al mijn verdomse DROOM geweest om een plekske waar Ezels efkes kunnen pauzeren open te doen, gedoeme.",
        "'t Is altijd al mijn droom geweest om een plekske waar Ezels efkes kunnen pauzeren open te doen, gedoeme.",
    ),
    ('E6_World_localization', 206): (
        "Awel godver Kameraad, da's precies daarom dadde ik mijn CAFE open, héé!",
        "Awel ja Kameraad, da's precies daarom dadde ik mijn CAFÉ open doe, héé!",
    ),
    ('E6_World_localization', 260): (
        'Slome Ezel? Waarom kom je hier terug aangesjokd?',
        'Slome Ezel? Waarom kom je hier terug aangesjokt?',
    ),
    ('E6_World_localization', 292): (
        'Onze B-B-Baas, Dierendokter D-D-Dina, heeft ons B-B-VERRADEN!',
        'Onze B-B-Baas, Dierendokter D-D-Dina, heeft ons V-V-VERRADEN!',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    if discrepancies:
        for s, r, msg in discrepancies:
            print(f'  ⚠ {s}/J{r}: {msg}')
        sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
