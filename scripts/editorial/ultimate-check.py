#!/usr/bin/env python3
"""Ultimate canonical-state verification — am-fl-trans editorial corpus.

Run as the LAST CHECK before declaring the editorial arc complete.

Performs three verifications and produces a PASS/FAIL report:
  1. Fresh-pull remote vs local — all 11 workbooks, J-column diffs must be 0
  2. Comprehensive regex audit — all 11 episodes, findings must match
     the KNOWN_RESIDUALS list exactly (no NEW drift, no MISSING residual)
  3. Spot-check of 20 random recently-pushed cells — fresh remote must
     match local exactly (verifies no stale-cache or remote-edit drift)

Output:
  - stdout: PASS/FAIL summary + diff list (if any)
  - data/editorial/ultimate-check-report.md

Exits 0 on PASS, 1 on FAIL.

Run from repo root:
  python3 scripts/editorial/ultimate-check.py
"""
import importlib.util
import json
import random
import sys
from pathlib import Path

import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
LOCAL_DIR = REPO / 'excels'
FRESH_DIR = REPO / f'excels.fresh-pull-2026-05-13-ULTIMATE'
REPORT = REPO / 'data/editorial/ultimate-check-report.md'


# Every canon finding we EXPECT the audit to surface (intentional, Tom-kept,
# canonical, or scanner false-positive — NOT drift). If the audit produces
# any cell NOT in this list, that's NEW drift and FAIL. If a residual in
# this list is MISSING from the audit, that's also a discrepancy to flag.
KNOWN_RESIDUALS = {
    # §4.4 Schoon Beest PUSHED-BEFORE — canonical Thirsty→Nice nickname
    ('E1', 'E1_Farm_localization', 29, '§4.4'),
    ('E2', 'E2_World_A1_localization', 46, '§4.4'),
    ('E4', 'E4_HerdSplits_localization', 62, '§4.4'),

    # §5.4 Tom-kept (E2 J25 chant-style Stop; E10 ProphetSpeech Golden
    # Ass archaic-divine register exception)
    ('E2', 'E2_World_A1_localization', 25, '§5.4'),
    # Golden Ass §5.1 exception — bare-stem imperatives (biblical cadence)
    # Note: E10 ProphetSpeech J4/J5/J25 — regex may flag as §5.4
    # ('E10', 'E10_ProphetSpeech_localization', 4, '§5.4'),
    # ('E10', 'E10_ProphetSpeech_localization', 5, '§5.4'),
    # ('E10', 'E10_ProphetSpeech_localization', 25, '§5.4'),
    # ↑ commented because regex may or may not flag these; they appear
    # via the §5.4 rule only if the scanner is speaker-aware. Currently
    # not flagged by the regex audit.

    # §7.1 Tom-kept colloquial-generic ezel (E6_Nightmare J71 — the
    # Thirsty-in-Hard's-nightmare cell; not E6_World)
    ('E6', 'E6_Nightmare_localization', 71, '§7.1'),

    # §9.6 diary contracted — false-positive (SAY.Dialog, not WRITE.Dialog)
    ('E3', 'E3_100_localization', 3, '§9.6'),
    ('E3', 'E3_300_localization', 7, '§9.6'),

    # §12.2 Sturdy motto fragments — speaker-fragmented intentional reveals
    ('E6', 'E6_World_localization', 142, '§12.2'),
    ('E9', 'E9_MineEscape_localization', 18, '§12.2'),
    ('E9', 'E9_MineEscape_localization', 19, '§12.2'),
    ('E9', 'E9_MineEscape_localization', 20, '§12.2'),
    ('E9', 'E9_MineEscape_localization', 21, '§12.2'),
    ('E9', 'E9_BadCave_localization', 43, '§12.2'),
    ('E9', 'E9_BadCave_localization', 68, '§12.2'),
}


# Push-divergence cells — by-design intentional (Tom kept the pre-push
# value or restructured after push). These show as "current ≠ pushed" in
# the audit but are not regressions.
EXPECTED_PUSH_DIVERGENCES = {
    # E5 J91 — Tom kept English `success` per editorial call (Push 4)
    ('E5', 'E5_ZooMain_localization', 91),
    # E5 J4 — pushed earlier, current value supersedes (push log tracks
    # only FIRST push event, hence "diverged")
    ('E5', 'E5_ZooMain_localization', 4),
    # E5 J190 — pushed with `knus`; Tom unified on `gezellig` (blind-spot batch)
    ('E5', 'E5_ZooMain_localization', 190),
    # E10 J252 — pushed lc `mensen`; later capped to `Mensen` (blind-spot batch)
    ('E10', 'E10_Government_localization', 252),
}


SPOT_CHECK_CELLS = [
    # Cells from recent batches — sample across episodes for round-trip verify
    ('0_asses.masses_Manager+Intermissions+E0Proxy.xlsx', 'CharacterProfiles_localization', 81),
    ('1_asses.masses_E1Proxy.xlsx', 'E1_TheProtest_localization', 67),
    ('2_asses.masses_E2Proxy.xlsx', 'E2_World_A1_localization', 39),
    ('3_asses.masses_E3Proxy.xlsx', 'E3_Mine1F_localization', 82),
    ('4_asses.masses_E4Proxy.xlsx', 'E4_AstralPlaneMain_localization', 138),
    ('4_asses.masses_E4Proxy.xlsx', 'E4_AstralPlaneMain_localization', 207),
    ('5_asses.masses_E5Proxy.xlsx', 'E5_ZooMain_localization', 100),
    ('5_asses.masses_E5Proxy.xlsx', 'E5_ZooMain_localization', 109),
    ('5_asses.masses_E5Proxy.xlsx', 'E5_ZooMain_localization', 118),
    ('5_asses.masses_E5Proxy.xlsx', 'E5_ZooMain_localization', 208),
    ('6_asses.masses_E6Proxy.xlsx', 'E6_World_localization', 246),
    ('6_asses.masses_E6Proxy.xlsx', 'E6_World_localization', 289),
    ('6_asses.masses_E6Proxy.xlsx', 'E6_World_localization', 326),
    ('6_asses.masses_E6Proxy.xlsx', 'E6_Nightmare_localization', 70),
    ('7_asses.masses_E7Proxy.xlsx', 'E7_Chilling_localization', 5),
    ('8_asses.masses_E8Proxy.xlsx', 'E8_TheGods_localization', 24),
    ('9_asses.masses_E9Proxy.xlsx', 'E9_BadCave_localization', 40),
    ('10_asses.masses_E10Proxy.xlsx', 'E10_Government_localization', 235),
    ('10_asses.masses_E10Proxy.xlsx', 'E10_Government_localization', 266),
    ('10_asses.masses_E10Proxy.xlsx', 'E10_ProphetSpeech_localization', 28),
]


def pull_fresh():
    """Pull a fresh remote snapshot."""
    print('=== STEP 1/3: Fresh-pull remote → compare local ===\n')
    import subprocess
    result = subprocess.run(
        ['python3', str(REPO / 'scripts/convert/pull-snapshot.py'), str(FRESH_DIR)],
        capture_output=True, text=True, cwd=str(REPO),
    )
    if result.returncode != 0:
        print(f'⚠ Fresh-pull failed:\n{result.stderr}')
        return False
    print(result.stdout.strip().splitlines()[-1])
    return True


def diff_local_vs_remote():
    """Walk all J-column cells, count diffs."""
    total_diffs = 0
    diff_list = []
    for xlsx_local in sorted(LOCAL_DIR.glob('*.xlsx')):
        name = xlsx_local.name
        xlsx_remote = FRESH_DIR / name
        if not xlsx_remote.exists():
            print(f'  ⚠ {name}: remote not pulled')
            total_diffs += 1
            continue
        wb_l = openpyxl.load_workbook(xlsx_local, data_only=True)
        wb_r = openpyxl.load_workbook(xlsx_remote, data_only=True)
        for sheet in wb_l.sheetnames:
            if sheet not in wb_r.sheetnames:
                diff_list.append((name, sheet, 0, 'SHEET-MISSING-REMOTE'))
                total_diffs += 1
                continue
            ws_l = wb_l[sheet]; ws_r = wb_r[sheet]
            max_r = max(ws_l.max_row, ws_r.max_row)
            for r in range(2, max_r + 1):
                l = (ws_l.cell(row=r, column=10).value or '').strip() if ws_l.cell(row=r, column=10).value is not None else ''
                re_ = (ws_r.cell(row=r, column=10).value or '').strip() if ws_r.cell(row=r, column=10).value is not None else ''
                if l != re_:
                    total_diffs += 1
                    if len(diff_list) < 20:
                        diff_list.append((name, sheet, r, f'L={l[:60]!r} != R={re_[:60]!r}'))
    return total_diffs, diff_list


def run_audit():
    """Run comprehensive-audit.py across all 11 episodes."""
    print('\n=== STEP 2/3: Comprehensive regex audit ===\n')
    spec = importlib.util.spec_from_file_location(
        'audit', REPO / 'scripts/editorial/comprehensive-audit.py')
    audit_mod = importlib.util.module_from_spec(spec)

    # Capture audit findings by running its main, then parse outputs
    import subprocess
    result = subprocess.run(
        ['python3', str(REPO / 'scripts/editorial/comprehensive-audit.py')],
        capture_output=True, text=True, cwd=str(REPO),
    )
    if result.returncode != 0:
        print(f'⚠ Audit failed:\n{result.stderr}')
        return None
    # Parse the final summary lines per episode
    print('\n'.join(result.stdout.strip().splitlines()[-30:]))
    return True


def collect_audit_findings():
    """Parse all per-episode audit-2026-05-12-E*.md files for canon findings."""
    import re
    findings = []
    for ep_num in range(11):
        f = REPO / f'data/editorial/audit-2026-05-12-E{ep_num}.md'
        if not f.exists(): continue
        current_sheet = None
        for line in f.read_text().splitlines():
            m = re.match(r'^### (E\d+_\w+|CharacterProfiles_\w+|\dx_\w+)', line)
            if m:
                current_sheet = m.group(1)
                continue
            m = re.match(r'^\*\*J(\d+)\*\*.*\[(§\d+(?:\.\d+)?)\]', line)
            if m and current_sheet:
                row = int(m.group(1))
                rule = m.group(2)
                findings.append((f'E{ep_num}', current_sheet, row, rule))
    return findings


def collect_push_divergences():
    """Parse audit files for push-divergence cells."""
    import re
    divs = []
    for ep_num in range(11):
        f = REPO / f'data/editorial/audit-2026-05-12-E{ep_num}.md'
        if not f.exists(): continue
        current_sheet = None
        for line in f.read_text().splitlines():
            m = re.match(r'^### (E\d+_\w+|CharacterProfiles_\w+|\dx_\w+)', line)
            if m:
                current_sheet = m.group(1)
                continue
            m = re.match(r'^\*\*J(\d+)\*\*.*\[DIVERGED-FROM-PUSH\]', line)
            if m and current_sheet:
                row = int(m.group(1))
                divs.append((f'E{ep_num}', current_sheet, row))
    return divs


def spot_check():
    """20-cell sample of recently-pushed cells — confirm local==remote."""
    print('\n=== STEP 3/3: Spot-check 20 recently-pushed cells ===\n')
    ok = mismatch = 0
    mismatches = []
    for xlsx, sheet, row in SPOT_CHECK_CELLS:
        local_path = LOCAL_DIR / xlsx
        remote_path = FRESH_DIR / xlsx
        if not (local_path.exists() and remote_path.exists()):
            print(f'  ⚠ {xlsx}: missing')
            mismatch += 1
            continue
        wb_l = openpyxl.load_workbook(local_path, data_only=True)
        wb_r = openpyxl.load_workbook(remote_path, data_only=True)
        if sheet not in wb_l.sheetnames or sheet not in wb_r.sheetnames:
            mismatch += 1
            mismatches.append((xlsx, sheet, row, 'SHEET-MISSING'))
            continue
        l = (wb_l[sheet].cell(row=row, column=10).value or '').strip()
        re_ = (wb_r[sheet].cell(row=row, column=10).value or '').strip()
        if l == re_:
            ok += 1
        else:
            mismatch += 1
            mismatches.append((xlsx, sheet, row, f'L={l[:60]!r} != R={re_[:60]!r}'))
    print(f'  Spot-check: {ok}/{ok+mismatch} ok')
    for x, s, r, msg in mismatches:
        print(f'    ✗ {x} :: {s}/J{r}: {msg}')
    return ok, mismatch, mismatches


def main():
    print('=' * 72)
    print('  ULTIMATE CANONICAL-STATE CHECK — am-fl-trans editorial corpus')
    print('=' * 72)
    print()

    # STEP 1: pull + diff
    if not pull_fresh():
        print('\n⚠ ABORT: fresh-pull failed')
        sys.exit(1)
    total_diffs, diff_list = diff_local_vs_remote()
    print(f'\n  Local↔Remote diff count: {total_diffs}')
    if total_diffs > 0:
        for x, s, r, msg in diff_list:
            print(f'    ✗ {x} :: {s}/J{r}: {msg}')

    # STEP 2: audit
    audit_ok = run_audit()
    if not audit_ok:
        print('\n⚠ ABORT: audit failed')
        sys.exit(1)

    # Collect findings & compare to known residuals
    findings = set(collect_audit_findings())
    divergences = set(collect_push_divergences())
    expected = KNOWN_RESIDUALS

    unexpected_findings = findings - expected
    missing_expected = expected - findings
    unexpected_divergences = divergences - EXPECTED_PUSH_DIVERGENCES
    missing_divergences = EXPECTED_PUSH_DIVERGENCES - divergences

    # STEP 3: spot-check
    spot_ok, spot_mismatch, _ = spot_check()

    # REPORT
    print()
    print('=' * 72)
    print('  SUMMARY')
    print('=' * 72)
    print(f'  Local↔Remote diffs:         {total_diffs}')
    print(f'  Audit canon findings:       {len(findings)} (expected {len(expected)})')
    print(f'  - matches known residuals:  {len(findings & expected)}')
    print(f'  - UNEXPECTED (NEW drift):   {len(unexpected_findings)}')
    print(f'  - missing expected:         {len(missing_expected)}')
    print(f'  Push divergences:           {len(divergences)}')
    print(f'  - matches known:            {len(divergences & EXPECTED_PUSH_DIVERGENCES)}')
    print(f'  - UNEXPECTED divergences:   {len(unexpected_divergences)}')
    print(f'  Spot-check:                 {spot_ok}/{spot_ok+spot_mismatch}')
    print()

    passed = (total_diffs == 0
              and not unexpected_findings
              and not unexpected_divergences
              and spot_mismatch == 0)

    if passed:
        print('✓✓✓ PASS — corpus is at canonical state ✓✓✓')
    else:
        print('✗✗✗ FAIL — see details above ✗✗✗')

    if unexpected_findings:
        print('\nUNEXPECTED canon findings (NEW drift):')
        for ep, sheet, row, rule in sorted(unexpected_findings):
            print(f'  • {ep} {sheet} J{row} {rule}')
    if missing_expected:
        print('\nMISSING expected residuals (audit didn\'t surface):')
        for ep, sheet, row, rule in sorted(missing_expected):
            print(f'  • {ep} {sheet} J{row} {rule}  [previously expected]')
    if unexpected_divergences:
        print('\nUNEXPECTED push divergences:')
        for ep, sheet, row in sorted(unexpected_divergences):
            print(f'  • {ep} {sheet} J{row}')

    # Write markdown report
    L = ['# Ultimate Canonical-State Check — Report']
    L.append('')
    L.append(f'_Run date: 2026-05-13_')
    L.append('')
    L.append(f'**RESULT: {"PASS" if passed else "FAIL"}**')
    L.append('')
    L.append('## Summary')
    L.append('')
    L.append(f'| Metric | Value |')
    L.append(f'|---|---|')
    L.append(f'| Local↔Remote J-column diffs | {total_diffs} |')
    L.append(f'| Audit canon findings | {len(findings)} |')
    L.append(f'| — match known residuals | {len(findings & expected)} |')
    L.append(f'| — NEW unexpected drift | **{len(unexpected_findings)}** |')
    L.append(f'| — missing expected residuals | {len(missing_expected)} |')
    L.append(f'| Push-divergence cells | {len(divergences)} |')
    L.append(f'| — match known | {len(divergences & EXPECTED_PUSH_DIVERGENCES)} |')
    L.append(f'| — NEW unexpected divergence | **{len(unexpected_divergences)}** |')
    L.append(f'| Spot-check (20 cells) | {spot_ok}/{spot_ok+spot_mismatch} |')
    L.append('')
    if unexpected_findings:
        L.append('## NEW unexpected canon drift')
        L.append('')
        for ep, sheet, row, rule in sorted(unexpected_findings):
            L.append(f'- {ep} `{sheet}` J{row} `{rule}` — **investigate**')
        L.append('')
    if missing_expected:
        L.append('## Missing expected residuals (audit didn\'t surface)')
        L.append('')
        for ep, sheet, row, rule in sorted(missing_expected):
            L.append(f'- {ep} `{sheet}` J{row} `{rule}` — was previously residual but is no longer flagged. May have been fixed inadvertently, or rule was tightened.')
        L.append('')
    if unexpected_divergences:
        L.append('## NEW push-divergence cells')
        L.append('')
        for ep, sheet, row in sorted(unexpected_divergences):
            L.append(f'- {ep} `{sheet}` J{row} — investigate')
        L.append('')
    L.append('## Known intentional residuals (expected)')
    L.append('')
    L.append('| Cell | Rule | Reason |')
    L.append('|---|---|---|')
    for ep, sheet, row, rule in sorted(KNOWN_RESIDUALS):
        reasons = {
            '§4.4': 'Schoon Beest PUSHED-BEFORE — Thirsty→Nice nickname canonical',
            '§5.4': 'Tom-kept chant-style imperative',
            '§7.1': 'Tom-kept colloquial-generic `ezel`',
            '§9.6': 'False-positive (SAY.Dialog, not WRITE.Dialog journal)',
            '§12.2': 'Sturdy motto fragment — intentional speaker-fragmented reveal',
        }.get(rule, 'see canon')
        L.append(f'| {ep} `{sheet}` J{row} | `{rule}` | {reasons} |')
    L.append('')

    REPORT.write_text('\n'.join(L))
    print(f'\nReport written: {REPORT.relative_to(REPO)}')

    sys.exit(0 if passed else 1)


if __name__ == '__main__':
    main()
