#!/usr/bin/env python3
"""E4 canon sweep scanner — flags candidates per sheet.

Adds to E3 scanner:
  §3.4   Mecha / Mechalen → Technopolis (with §3.4.1 Imechelda exception)
  §3.4.1 watch for "Mecha [X] Ziekenhuis" pattern
  §8.2   Machine-compound hyphen (Boormachine, Camion Machine, etc.)
"""
import re
import sys
from pathlib import Path
import openpyxl

XLSX = Path('excels/4_asses.masses_E4Proxy.xlsx')
DIARY_SHEETS = set()  # no diary sheets in E4

RETIRED_MONIKERS = ['Felle Gast', 'Betweter', 'Bikkeharde', 'Snotezel', 'Stampgast', 'Sloom']

# §8.2: Machine compounds that should be hyphenated. Match ONLY the unhyphenated form:
#   - solid (Boormachine) — wrong
#   - space-separated (Camion Machine) — wrong
# Hyphenated form (Camion-Machine) is correct and NOT flagged.
MACHINE_COMPOUNDS_UNHYPHENATED = [
    r'\bBoormachine[ns]?\b',           # solid; should be Boor-Machine
    r'\bCamion[ ]Machine[ns]?\b',      # space; should be Camion-Machine
    r'\bTractor[ ]Machine[ns]?\b',
    r'\bAuto[ ]Machine[ns]?\b',
    r'\bMysterie[ ]Machine[ns]?\b',
    r'\bTank[ ]Machine[ns]?\b',
    r'\bVlieg[ ]Machine[ns]?\b',
]


def scan_sheet(ws, sheet_name):
    findings = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) < 10:
            continue
        speaker = (row[1] or '').strip()
        en = (row[2] or '').strip()
        nl = (row[9] or '').strip()
        if not nl:
            continue

        def add(rule, note):
            findings.append((r, rule, note, speaker, en, nl))

        # §2 nie (standalone)
        if re.search(r"\bnie\b", nl, flags=re.IGNORECASE) or re.search(r"\bnie'", nl):
            add('§2', 'nie standalone → niet')

        # §7.1 Ezel cap — lowercase 'ezel' mid-sentence
        for m in re.finditer(r"\bezel(s|innekes|tje|tjes|en|kes)?\b", nl):
            if m.start() == 0:
                continue
            prev = nl[max(0, m.start()-2):m.start()]
            if re.search(r"[.!?]\s$", prev + ' '):
                continue
            add('§7.1', f'lowercase "{m.group(0)}" mid-sentence — cap?')
            break

        # §3.6 Boerderij
        if re.search(r"\bboerderij\b", nl, flags=re.IGNORECASE):
            add('§3.6', 'Boerderij → Hoeve')

        # §3.1 Muilebeek (no extra n)
        if re.search(r"\bMuilebeek\b", nl):
            add('§3.1', 'Muilebeek → Muilenbeek')

        # §3.4 Mecha / Mechalen → Technopolis
        # Watch for §3.4.1 exception: "Mecha [X] Ziekenhuis" → Imechelda variant
        if re.search(r"\bMecha(len)?\b", nl):
            if 'Ziekenhuis' in nl:
                add('§3.4.1', 'Mecha + Ziekenhuis → Imechelda Algemeen Ziekenhuis (exception)')
            else:
                add('§3.4', 'Mecha/Mechalen → Technopolis')
        if 'MECHA' in nl and 'MECHALEN' not in nl:
            # all-caps form
            add('§3.4', 'MECHA → TECHNOPOLIS (all-caps)')

        # §6.16 Beroep / Beroepen
        if re.search(r"\bBeroep(en)?\b", nl):
            if 'beroep doen op' in nl.lower():
                pass
            else:
                add('§6.16', 'Beroep(en) → Job(s) — verify game-system context')

        # §6.16 lowercase job/jobs
        if re.search(r"\bjob[s]?\b", nl) and not re.search(r"\bJob[s]?\b", nl):
            add('§6.16-lc', 'lowercase job/jobs — game-system? cap to Job/Jobs')

        # §6.1 HUDO / Hudo / Sekreet / Privaat
        if re.search(r"\b[Hh]udo\b", nl) or 'HUDO' in nl:
            add('§6.1', 'Hudo → Plee')
        if re.search(r"\bSekreet\b|\bPrivaat\b", nl):
            add('§6.1', 'Sekreet/Privaat → Plee')

        # §6.2 Oom (non-Smart Ass)
        if re.search(r"\bOom\b", nl) and 'Smart' not in speaker:
            add('§6.2', f'Oom (speaker={speaker!r}) → Nonkel?')

        # §8.1 Machien
        if re.search(r"\bMachien(en)?\b", nl):
            add('§8.1', 'Machien → Machine')

        # §8.2 Machine compounds — unhyphenated
        for pat in MACHINE_COMPOUNDS_UNHYPHENATED:
            if re.search(pat, nl):
                add('§8.2', f'machine compound needs hyphen ({pat})')

        # §4.3 retired monikers
        for mon in RETIRED_MONIKERS:
            if mon in nl:
                add('§4.3', f'retired moniker {mon}')

        # §4.4 Schoon Beest — flag (preserve only if Thirsty→Nice)
        if 'Schoon Beest' in nl:
            add('§4.4', f'Schoon Beest (speaker={speaker!r}) — preserve only if Thirsty→Nice')

        # §12.1 , Kameraad
        if re.search(r",\s*Kameraad\b", nl):
            if 'Comrade' not in en:
                add('§12.1', ', Kameraad without EN Comrade — strip?')

        # §14.1 retired slogans
        for slog in ['EZELSKRACHT', 'EZEL MACHT', 'EZELKRACHT', 'EZELS-KRACHT']:
            if slog in nl:
                add('§14.1', f'retired slogan {slog} → EZELS EERST')

        # §5.4 Stop imperative
        if re.search(r"(^|[.!?]\s+)Stop\b", nl):
            add('§5.4', 'Stop imperative — verify register')

        # §9.1 sentence-start apostrophe missing
        if re.search(r"(^|[.!?]\s+)k\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: k → 'k")
        if re.search(r"(^|[.!?]\s+)t\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: t → 't")

        # §9.6 diary — N/A for E4

        # §9.2 mid-sentence Ik [Cap]
        if re.search(r",\s+Ik\s+[A-Z]", nl):
            add('§9.2', "mid-sentence ', Ik [Cap]' — lowercase verb")

    return findings


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    target_sheets = sys.argv[1:] or wb.sheetnames
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            matches = [s for s in wb.sheetnames if s.startswith(sheet_name)]
            if not matches:
                print(f'!! sheet not found: {sheet_name}')
                continue
            sheet_name = matches[0]
        ws = wb[sheet_name]
        findings = scan_sheet(ws, sheet_name)
        print(f'\n=== {sheet_name}: {len(findings)} finding(s) ===')
        for r, rule, note, speaker, en, nl in findings:
            print(f'  J{r} [{rule}] {note}')
            print(f'    Speaker: {speaker!r}')
            print(f'    EN: {en!r}')
            print(f'    NL: {nl!r}')
    wb.close()


if __name__ == '__main__':
    main()
