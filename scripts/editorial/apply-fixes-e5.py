#!/usr/bin/env python3
"""Apply user-approved E5 fixes to local xlsx (and optionally push to remote).

User sign-off (2026-05-13):
  ✅ 23 mechanical fixes per propose-fixes-e5.py
  ✅ J22 (CircusMain): SKIP — keep "kabinet" for flavor (user decision)
  ✅ J91 (CircusMain): user override — full replacement string provided
  ✅ J176 + J190 + J22 (CircusMain) — user-approved despite [VERIFY] tier

Total: 24 cells fixed, 1 skipped.

Safety:
  - Each cell read pre-edit: current NL must match the expected pre-image
    OR equal the proposed post-image (idempotent — already-applied is fine).
    Aborts with a hard error otherwise.
  - Writes locally only; the push step is invoked separately via
    scripts/convert/push-file.py.

Usage:
  python3 scripts/editorial/apply-fixes-e5.py             # dry-run (default)
  python3 scripts/editorial/apply-fixes-e5.py --apply     # write to local xlsx
"""
import sys
from pathlib import Path

import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/5_asses.masses_E5Proxy.xlsx'

# Each entry: (sheet, row): (expected_current_nl, proposed_nl)
# Skipped cells excluded from this dict entirely.
# J91 is the user-override string.
APPROVED = {
    # E5_CircusMain_localization
    ('E5_CircusMain_localization', 9): (
        'IEDEREEN WELKOM in het GROTE stadscircus!',
        'IEDEREEN WELKOM in het GROTE STADSCIRCUS!',
    ),
    # J22 SKIPPED per user — keep kabinet for flavor
    ('E5_CircusMain_localization', 36): (
        '...wie zijn best gedaan heeft, maar toch net niet het podium tot leven gebracht heeft.',
        '...die zijn best gedaan heeft, maar toch net niet het podium tot leven gebracht heeft.',
    ),
    ('E5_CircusMain_localization', 64): (
        'Haar DNA is 95% mens en 5% MAGIE... en ze volgt feedback als de beste!',
        'Haar DNA is 95% MENS en 5% MAGIE... en ze volgt feedback als de beste!',
    ),
    ('E5_CircusMain_localization', 88): (
        'Denkt ge dat er presentatoren in het publiek zetten?',
        'Denkt ge dat er presentatoren in het publiek zaten?',
    ),
    ('E5_CircusMain_localization', 91): (
        'Ha ha! Veel success met mijn act te overtreffen!',
        'Ha ha! Veel success met het overtreffen van mijn nummer!',  # USER OVERRIDE
    ),
    ('E5_CircusMain_localization', 123): (
        'Ik kan niet wachten om op toernee te gaan! Groter publiek, mooiere kleedkamers...',
        'Ik kan niet wachten om op Tournee te gaan! Groter publiek, mooiere kleedkamers...',
    ),
    ('E5_CircusMain_localization', 153): (
        'Ik hoor nauwelijk iets — geef me nog een oorverdovend applaus voor CHEKOV!',
        'Ik hoor nauwelijks iets — geef me nog een oorverdovend applaus voor CHEKOV!',
    ),
    ('E5_CircusMain_localization', 161): (
        'Ik denk dat we nu wel MEER dan gewoon kameraden zijn!',
        'Ik denk dat we nu wel MEER dan gewoon Kameraden zijn!',
    ),

    # E5_Highway_localization
    ('E5_Highway_localization', 4): (
        'Oef... mijn knieëen.',
        'Oef... mijn knieën.',
    ),
    ('E5_Highway_localization', 11): (
        '*zwaar gëadem*',
        '*zwaar geademd*',
    ),

    # E5_ZooMain_localization
    ('E5_ZooMain_localization', 4): (
        # NB: xlsx stores literal backslash-n (not actual newlines)
        'Leuke weetjes over de Equus Asinus: \\n\\n- Oorspronkelijk van Afrika\\n- Worden tot 50 jaar oud\\n- Zeer wendbaar op rotsachtig terrein\\n- Word sinds 7,000 jaar als werkdier gebruikt\\n- Vindt de regen niet zo fijn',
        'Leuke weetjes over de Equus Asinus: \\n\\n- Oorspronkelijk van Afrika\\n- Worden tot 50 jaar oud\\n- Zeer wendbaar op rotsachtig terrein\\n- Worden sinds 7,000 jaar als werkdier gebruikt\\n- Vinden de regen niet zo fijn',
    ),
    ('E5_ZooMain_localization', 18): (
        'Het volk stroomt binnen en het word net zo druk als gisteren!',
        'Het volk stroomt binnen en het wordt net zo druk als gisteren!',
    ),
    ('E5_ZooMain_localization', 47): (
        'Wees maar niet zo zelfingenomen. Jouw gelukt komt zo op zijn eind!',
        'Wees maar niet zo zelfingenomen. Jouw geluk komt zo op zijn eind!',
    ),
    ('E5_ZooMain_localization', 119): (
        'Ik dacht dat de Boederij en Mijn achterwege laten me misschien een kans zou geven mijn gebroken hart te helen...',
        'Ik dacht dat de Hoeve en Mijn achterwege laten me misschien een kans zou geven mijn gebroken hart te helen...',
    ),
    ('E5_ZooMain_localization', 125): (
        'Deze Dierentuin heeft verse hooi, lieve Mensen, geen lelijke Machines en rigoreus dagdagelijks labeur!',
        'Deze Dierentuin heeft vers hooi, lieve Mensen, geen lelijke Machines en rigoureus dagdagelijks labeur!',
    ),
    ('E5_ZooMain_localization', 154): (
        'Ik weet niet wat mij vreugde bengt...',
        'Ik weet niet wat mij vreugde brengt...',
    ),
    ('E5_ZooMain_localization', 165): (
        'Het lijkt er op dat onze Iejoor van deze Dierentuin ONTSNAPT is...',
        'Het lijkt er op dat onze Iejoor uit deze Dierentuin ONTSNAPT is...',
    ),
    ('E5_ZooMain_localization', 169): (
        '...De Raad van Bestuur van de Dierentuin komt om de Ezel van de Week ontmoeten?',
        '...De Raad van Bestuur van de Dierentuin komt om de Ezel van de Week te ontmoeten?',
    ),
    ('E5_ZooMain_localization', 176): (
        'De RAAD VAN BESTUUR wilt dat jullie gezellig hun komt Knuffelen!',
        'De RAAD VAN BESTUUR wil dat jullie ze gezellig komen Knuffelen!',
    ),
    ('E5_ZooMain_localization', 190): (
        'De RAAD VAN BESTUUR wilt dat jullie knus hun komt Knuffelen!',
        'De RAAD VAN BESTUUR wil dat jullie ze knus komen Knuffelen!',
    ),
    ('E5_ZooMain_localization', 216): (
        'WAT?! Ge bent zo lomp als een Machine.',
        'WAT?! Gij zijt zo lomp als een Machine.',
    ),

    # E5_Zoo_Introduction_localization
    ('E5_Zoo_Introduction_localization', 12): (
        'Ik noem het Ezel van de Week.',
        'Ik noem het EZEL VAN DE WEEK.',
    ),
    ('E5_Zoo_Introduction_localization', 22): (
        'Ik heb een plan!',
        'Ik heb een Plan!',
    ),
    ('E5_Zoo_Introduction_localization', 33): (
        'En dan is natuurlijk dat oude dametje die het gewoon leuk vind als je naar haar luistert.',
        'En dan is natuurlijk dat oude dametje die het gewoon leuk vindt als je naar haar luistert.',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    print(f'  Tabs: {wb.sheetnames}\n')

    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        # Resolve truncated tab names (Excel 31-char limit)
        if sheet_name not in wb.sheetnames:
            match = next((s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
            if match is None:
                discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
                continue
            actual_sheet = match
        else:
            actual_sheet = sheet_name

        ws = wb[actual_sheet]
        cell = ws.cell(row=row, column=10)  # column J
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED-CURRENT: {current!r} (expected {expected!r})'))

    print(f'Status:')
    print(f'  Will write:       {len(will_write)} cell(s)')
    print(f'  Already applied:  {len(already_applied)} cell(s)')
    print(f'  Discrepancies:    {len(discrepancies)} cell(s)')
    print()

    if discrepancies:
        print('⚠ DISCREPANCIES — aborting before write:')
        for sheet, row, msg in discrepancies:
            print(f'  {sheet} / J{row}: {msg}')
        sys.exit(1)

    if not will_write:
        print('All cells already at proposed values — nothing to write.')
        return

    print('Proposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet} / J{row}:')
        print(f'    -  {current[:80]!r}')
        print(f'    +  {proposed[:80]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply to write to xlsx.')
        return

    print('\nApplying writes…')
    for sheet, row, _, proposed in will_write:
        ws = wb[sheet]
        ws.cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells to {XLSX.name}')


if __name__ == '__main__':
    main()
