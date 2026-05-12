#!/usr/bin/env python3
"""E3 canon sweep scanner — flags candidates per sheet."""
import re
import sys
from pathlib import Path
import openpyxl

XLSX = Path('excels/3_asses.masses_E3Proxy.xlsx')
DIARY_SHEETS = {'E3_100_localization', 'E3_200_localization', 'E3_300_localization'}

RETIRED_MONIKERS = ['Felle Gast', 'Betweter', 'Bikkeharde', 'Snotezel', 'Stampgast', 'Sloom']
GAME_TERMS_LOWER_CHECK = {  # lowercase mid-sentence might need cap if game-system context
    # not auto-flagging — too many false positives. Will be manual.
}


def scan_sheet(ws, sheet_name):
    findings = []
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) < 10:
            continue
        speaker = (row[1] or '').strip()
        en = (row[2] or '').strip()
        nl = (row[9] or '').strip()
        if not nl:
            continue

        def add(rule, note):
            findings.append((r, rule, note, speaker, en, nl))

        # §2 nie (standalone)
        if re.search(r"\bnie\b", nl, flags=re.IGNORECASE) or re.search(r"\bnie'", nl):
            # avoid niet, niets, nieuw, etc.
            add('§2', 'nie standalone → niet')

        # §7.1 Ezel cap — lowercase 'ezel' mid-sentence (not at sentence start, not after . ! ?)
        # crude: any lowercase 'ezel' is suspect
        for m in re.finditer(r"\bezel(s|innekes|tje|tjes|en|kes)?\b", nl):
            # exclude if at very start of nl
            if m.start() == 0:
                continue
            # exclude if preceded by sentence-ending punct
            prev = nl[max(0, m.start()-2):m.start()]
            if re.search(r"[.!?]\s$", prev + ' '):
                continue
            add('§7.1', f'lowercase "{m.group(0)}" mid-sentence — cap?')
            break  # one flag per cell

        # §3.6 Boerderij
        if re.search(r"\bboerderij\b", nl, flags=re.IGNORECASE):
            add('§3.6', 'Boerderij → Hoeve')

        # §3.1 Muilebeek (no extra n) — should be Muilenbeek
        if re.search(r"\bMuilebeek\b", nl):
            add('§3.1', 'Muilebeek → Muilenbeek')

        # §6.16 Beroep / Beroepen (game-system) — but exempt "beroep doen op"
        if re.search(r"\bBeroep(en)?\b", nl):
            if 'beroep doen op' in nl.lower():
                pass
            else:
                add('§6.16', 'Beroep(en) → Job(s) — verify game-system context')

        # §6.16 lowercase job/jobs — flag for manual game-system context check
        if re.search(r"\bjob[s]?\b", nl) and not re.search(r"\bJob[s]?\b", nl):
            add('§6.16-lc', 'lowercase job/jobs — game-system? cap to Job/Jobs')

        # §6.1 HUDO / Hudo / Sekreet / Privaat
        if re.search(r"\b[Hh]udo\b", nl) or 'HUDO' in nl:
            add('§6.1', 'Hudo → Plee')
        if re.search(r"\bSekreet\b|\bPrivaat\b", nl):
            add('§6.1', 'Sekreet/Privaat → Plee')

        # §6.2 Oom (non-Smart Ass)
        if re.search(r"\bOom\b", nl) and 'Smart' not in speaker:
            add('§6.2', f'Oom (speaker={speaker!r}) → Nonkel?')

        # §8.1 Machien
        if re.search(r"\bMachien(en)?\b", nl):
            add('§8.1', 'Machien → Machine')

        # §4.3 retired monikers
        for mon in RETIRED_MONIKERS:
            if mon in nl:
                add('§4.3', f'retired moniker {mon}')
        # Schoon Beest — flag (§4.4 exception: Thirsty→Nice keep)
        if 'Schoon Beest' in nl:
            add('§4.4', f'Schoon Beest (speaker={speaker!r}) — preserve only if Thirsty→Nice')

        # §12.1 , Kameraad — flag if EN lacks Comrade
        if re.search(r",\s*Kameraad\b", nl):
            if 'Comrade' not in en:
                add('§12.1', ', Kameraad without EN Comrade — strip?')

        # §14.1 retired slogans
        for slog in ['EZELSKRACHT', 'EZEL MACHT', 'EZELKRACHT', 'EZELS-KRACHT']:
            if slog in nl:
                add('§14.1', f'retired slogan {slog} → EZELS EERST')

        # §5.4 Stop/Stopt imperative — narrow to imperative-like contexts
        # Flag bare 'Stop' at sentence start or after !/. — manual check needed
        if re.search(r"(^|[.!?]\s+)Stop\b", nl):
            add('§5.4', 'Stop imperative — verify register')

        # §9.1 sentence-start apostrophe missing (k /'t /'n ) — at very start or after . ! ?
        if re.search(r"(^|[.!?]\s+)k\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: k → 'k")
        if re.search(r"(^|[.!?]\s+)t\s+[A-Z]", nl):
            add('§9.1', "missing apostrophe: t → 't")

        # §9.6 diary uncontracted — flag 'k or 't in diary sheets
        if sheet_name in DIARY_SHEETS:
            if re.search(r"'[kt]\b", nl):
                add('§9.6', "diary: contracted 'k/'t → uncontracted Ik/Het")

        # §9.2 mid-sentence Ik [Cap] (Patrick artifact)
        if re.search(r",\s+Ik\s+[A-Z]", nl):
            add('§9.2', "mid-sentence ', Ik [Cap]' — lowercase verb")

    return findings


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    target_sheets = sys.argv[1:] or wb.sheetnames
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            # prefix match
            matches = [s for s in wb.sheetnames if s.startswith(sheet_name)]
            if not matches:
                print(f'!! sheet not found: {sheet_name}')
                continue
            sheet_name = matches[0]
        ws = wb[sheet_name]
        findings = scan_sheet(ws, sheet_name)
        print(f'\n=== {sheet_name}: {len(findings)} finding(s) ===')
        for r, rule, note, speaker, en, nl in findings:
            print(f'  J{r} [{rule}] {note}')
            print(f'    Speaker: {speaker!r}')
            print(f'    EN: {en!r}')
            print(f'    NL: {nl!r}')
    wb.close()


if __name__ == '__main__':
    main()
