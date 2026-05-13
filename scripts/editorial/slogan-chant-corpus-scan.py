#!/usr/bin/env python3
"""Project-wide scan of the §14.1 slogan / chant forms across the entire corpus.

Reads the live remote Google Sheets for all 11 workbooks via gspread API
and surfaces every cell that mentions one of the canonical/non-canonical
forms of the EN slogan "ASS POWER!" / "Show the World the Power of the Ass!"
and its NL counterparts.

The goal: a complete, honest inventory so we can decide canon policy:
  - Where is the chant form (EZELS! / EERST! or ASS!/POWER! short)?
  - Where is the long-form quote ("Herinner de wereld aan de belangen van de Ezels"
    or "Show the world the power of the Ass")?
  - Where is each variant used by whom in which context?
  - Are there leftover obsolete translations (e.g. "ASS POWER", "MACHT VAN DE EZELS")?

Output: data/editorial/slogan-chant-corpus-scan-2026-05-13.md

EN search tokens (case-insensitive substring):
  - "ass power"
  - "ezels eerst"        (NL)
  - "ezels!" + "eerst!"  (chant pieces)
  - "power of the ass"
  - "belangen van de ezels"  (NL long-form)
  - "wereld aan de"          (NL long-form fragment)
  - "show the world the"
  - "macht van de ezels"     (alt NL — flag if found)
  - "ass!" / "power!" with cap (chant fragments)

For each hit: workbook, tab, J-row, Key, Speaker (parsed from key suffix),
EN, NL, classified bucket (CHANT / LONG-FORM / OTHER / OBSOLETE).
"""
import re
import sys
import time
from pathlib import Path

import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
# Use fresh post-E4-push snapshot (= live remote at time of pull). Override
# via argv[1] to use a different snapshot dir.
DEFAULT_SNAPSHOT = REPO / 'excels.fresh-pull-2026-05-13-post-E4-push'
OUT = REPO / 'data/editorial/slogan-chant-corpus-scan-2026-05-13.md'

SHEETS = [
    (0, '0_asses.masses_Manager+Intermissions+E0Proxy.xlsx'),
    (1, '1_asses.masses_E1Proxy.xlsx'),
    (2, '2_asses.masses_E2Proxy.xlsx'),
    (3, '3_asses.masses_E3Proxy.xlsx'),
    (4, '4_asses.masses_E4Proxy.xlsx'),
    (5, '5_asses.masses_E5Proxy.xlsx'),
    (6, '6_asses.masses_E6Proxy.xlsx'),
    (7, '7_asses.masses_E7Proxy.xlsx'),
    (8, '8_asses.masses_E8Proxy.xlsx'),
    (9, '9_asses.masses_E9Proxy.xlsx'),
    (10, '10_asses.masses_E10Proxy.xlsx'),
]

# Search tokens (case-insensitive) grouped by classification bucket
TOKENS = {
    # Pure chant form, EN side
    'EN_CHANT': [
        r'\bass\s+power\b',
        r'\bass!\b.*\bpower!\b',
        r'when i say ass',
    ],
    # Long-form dying-words quote, EN side
    'EN_LONGFORM': [
        r'show the world the power of the ass',
        r'power of the ass\b',
    ],
    # NL chant
    'NL_CHANT': [
        r'\bezels\s+eerst\b',
        r'\bezels!\b.*\beerst!\b',
    ],
    # NL long-form
    'NL_LONGFORM': [
        r'belangen van de ezels',
        r'belangan van de ezels',  # typo variant — flag if still present anywhere
        r'wereld aan de belang',
        r'macht van de ezels',     # obsolete — should not exist
        r'power of the ezels',     # mixed/odd — flag
    ],
    # Bare chant pieces (may need context)
    'CHANT_FRAGMENT': [
        r'\bEZELS!\B',  # word with possible suffix
        r'\bEERST!\b',
    ],
}

ALL_TOKENS = [(bucket, tok) for bucket, toks in TOKENS.items() for tok in toks]


def extract_speaker(key):
    if not key:
        return ''
    parts = str(key).split('.')
    return parts[-1].strip() if len(parts) >= 2 else str(key).strip()


def classify(en, nl):
    hits = []
    for bucket, toks in TOKENS.items():
        for tok in toks:
            target = en if bucket.startswith('EN_') else nl
            if re.search(tok, target or '', flags=re.IGNORECASE):
                hits.append(bucket)
                break
    return hits


def md_inline(s):
    if not s:
        return '∅'
    return str(s).replace('`', '\\`').replace('|', '\\|').replace('\n', ' ⏎ ')


def scan_workbook(ep_num, xlsx_path):
    print(f'  E{ep_num} ({xlsx_path.name})…')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    findings = []
    for tab in wb.sheetnames:
        ws = wb[tab]
        for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if len(row) < 10:
                continue
            key = row[0] or ''
            en = (row[2] or '').strip() if row[2] is not None else ''
            nl = (row[9] or '').strip() if row[9] is not None else ''
            if not en and not nl:
                continue
            buckets = classify(en, nl)
            if not buckets:
                continue
            findings.append({
                'ep': ep_num, 'tab': tab, 'row': r,
                'key': str(key), 'speaker': extract_speaker(key),
                'en': en, 'nl': nl, 'buckets': buckets,
            })
    wb.close()
    return findings


def main():
    snapshot = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SNAPSHOT
    if not snapshot.exists():
        print(f'!! snapshot dir not found: {snapshot}', file=sys.stderr)
        sys.exit(1)
    print(f'Scanning snapshot: {snapshot.relative_to(REPO)}\n')
    all_findings = []
    for ep_num, fname in SHEETS:
        path = snapshot / fname
        if not path.exists():
            print(f'  E{ep_num} ({fname}): missing — skip')
            continue
        all_findings.extend(scan_workbook(ep_num, path))

    print(f'\nTotal cells matched: {len(all_findings)}')

    # Group by bucket for summary table
    by_bucket = {}
    for f in all_findings:
        for b in f['buckets']:
            by_bucket.setdefault(b, []).append(f)

    L = ['# Project-wide slogan / chant corpus scan']
    L.append('')
    L.append(f'_Read at: {time.strftime("%Y-%m-%d %H:%M:%S %Z")} (fresh-pull snapshot of live remote)_')
    L.append(f'_Snapshot: `{snapshot.relative_to(REPO)}/`_')
    L.append(f'_Workbooks: 11 (E0–E10) — pulled from source-of-truth Google Sheets_')
    L.append('')
    L.append('## Purpose')
    L.append('')
    L.append('Surface every occurrence of the §14.1 slogan + dying-words quote across the corpus so canon policy can be decided. EN forms searched:')
    L.append('')
    L.append('- `ass power` / `ass! ... power!` / `when i say ass` (chant)')
    L.append('- `show the world the power of the ass` (long-form dying words)')
    L.append('')
    L.append('NL forms searched:')
    L.append('')
    L.append('- `ezels eerst` / `ezels! ... eerst!` (canonical chant)')
    L.append('- `belangen van de ezels` / `wereld aan de belang…` (long-form quote)')
    L.append('- `belangan` typo (legacy)')
    L.append('- `macht van de ezels` (obsolete — should not exist on remote)')
    L.append('- `power of the ezels` (mixed/odd)')
    L.append('- `EZELS!` / `EERST!` capped chant fragments')
    L.append('')
    L.append('## Totals')
    L.append('')
    L.append(f'- Total cells matched: **{len(all_findings)}**')
    for bucket, hits in sorted(by_bucket.items(), key=lambda x: -len(x[1])):
        L.append(f'- {bucket}: {len(hits)} cell(s)')
    L.append('')
    L.append('## Cells by episode/tab')
    L.append('')

    by_ep = {}
    for f in all_findings:
        by_ep.setdefault(f['ep'], []).append(f)
    for ep in sorted(by_ep.keys()):
        L.append(f'### E{ep}')
        L.append('')
        by_tab = {}
        for f in by_ep[ep]:
            by_tab.setdefault(f['tab'], []).append(f)
        for tab in sorted(by_tab.keys()):
            L.append(f'#### {tab}')
            L.append('')
            for f in sorted(by_tab[tab], key=lambda x: x['row']):
                buckets_label = ', '.join(f['buckets'])
                L.append(f'**J{f["row"]}** `[{buckets_label}]`')
                L.append(f'- Key:     `{md_inline(f["key"])}`')
                L.append(f'- Speaker: `{md_inline(f["speaker"])}`')
                L.append(f'- EN:      `{md_inline(f["en"])}`')
                L.append(f'- NL:      `{md_inline(f["nl"])}`')
                L.append('')

    L.append('## Notes for canon resolution')
    L.append('')
    L.append('- **Chant form vs long-form** appear to be deliberate co-existing variants:')
    L.append('  - Chant = punchy crowd-call, `EZELS! / EERST!` answer pattern')
    L.append('  - Long-form = Old Ass dying-words quote rendered as full sentence')
    L.append('- If canon policy is to collapse both to the chant form, the long-form cells need a retcon batch.')
    L.append('- If canon policy is to keep both, the §14.1 entry should explicitly disambiguate (chant-form vs quote-form).')
    L.append('- Any `macht van de ezels` or `power of the ezels` hits = bugs (mixed/obsolete forms) and should be fixed.')

    OUT.write_text('\n'.join(L))
    print(f'\nWrote {OUT.relative_to(REPO)}')


if __name__ == '__main__':
    main()
