#!/usr/bin/env python3
"""Apply E6_World blind-spot walk fixes — Tom sign-off 2026-05-13.

12 writes / 1 keep (J334). J67 already at target `OKÉ DAN...` from cross-cell
unification target (no write needed).

Tom decisions:
  - J304 alternate `wij moeten` (more formal than `we`)
  - J326 Tom override `Is 't niet, Kameraad Bikkelharde?`
  - J251 `GODVERDOEMS` (Thirsty -oe- family per codex)
  - J246 fix with Tom tweak `me oewe`→`met u` (Flemish colloquial possessive)
  - J289 stutter §12.3 fix `b-b-geloven`→`g-g-geloven`
  - J184 single-assertion form
  - J257/J275 strip leading whitespace
  - J191 Option B `jezelf` (je/jij agreement)
  - J84/J88 cross-cell unify on `OKÉ DAN...` (J67 already at target)
  - J319 Option A `?!`→`?` (revert to EN match)
  - J124 alternate `'t Veulentje` (Sturdy contraction)
  - J334 KEPT (`vergeten` register)
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/6_asses.masses_E6Proxy.xlsx'

APPROVED = {
    # HIGH §13/codex/stutter (5 writes)
    ('E6_World_localization', 304): (
        'Kameraden, ik denk dat je moet onthouden dat—',
        'Kameraden, ik denk dat wij moeten onthouden dat—',
    ),
    ('E6_World_localization', 326): (
        'Ist dat niet waar, Kameraad Bikkelharde?',
        "Is 't niet, Kameraad Bikkelharde?",
    ),
    ('E6_World_localization', 251): (
        'EEN GODVERDOMS FEESTJE, HÉÉ!',
        'EEN GODVERDOEMS FEESTJE, HÉÉ!',
    ),
    ('E6_World_localization', 246): (
        'TRIESTIGE EZEL! Ik heb nen emmer me jouwe naam erop, héé!',
        'TRIESTIGE EZEL! Ik heb nen emmer met u naam erop, héé!',
    ),
    ('E6_World_localization', 289): (
        'Als je het kunt b-b-geloven—ík, tot mijn eigen verb-b-bazing, heb-b Ezel van de Week gewonnen!',
        'Als je het kunt g-g-geloven—ík, tot mijn eigen verb-b-bazing, heb-b Ezel van de Week gewonnen!',
    ),

    # MEDIUM (3 writes)
    ('E6_World_localization', 184): (
        'Blijkt dat ik goed ben in het zorgen voor de herinneringen aan onze verloren geliefden... dat is iets waar ik echt goed in ben...',
        'Blijkt dat het zorgen voor de herinneringen aan onze verloren geliefden iets is waar ik echt goed in ben...',
    ),
    ('E6_World_localization', 257): (
        ' Ik b-b-ben eigenlijk een Leeuw!',
        'Ik b-b-ben eigenlijk een Leeuw!',
    ),
    ('E6_World_localization', 275): (
        ' Je zou je ogen niet geloofd heb-b-ben!',
        'Je zou je ogen niet geloofd heb-b-ben!',
    ),

    # LOW / VERIFY (5 writes; J334 KEPT)
    ('E6_World_localization', 191): (
        'Pas goed op uzelf, Kameraad.',
        'Pas goed op jezelf, Kameraad.',
    ),
    # J67 already at `OKÉ DAN...` — no write (idempotent in script logic)
    ('E6_World_localization', 84): (
        'GOED DAN...',
        'OKÉ DAN...',
    ),
    ('E6_World_localization', 88): (
        'PRIMA...',
        'OKÉ DAN...',
    ),
    ('E6_World_localization', 319): (
        'Wie heeft er hier nog vuile geheimen zitten verzwijgen?!',
        'Wie heeft er hier nog vuile geheimen zitten verzwijgen?',
    ),
    ('E6_World_localization', 124): (
        'Veulentje kan soms echt een hoefvol zijn.',
        "'t Veulentje kan soms echt een hoefvol zijn.",
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        # For J257/J275: don't strip leading whitespace from current for comparison
        # because the leading space IS the drift we're fixing.
        if row in (257, 275):
            current_raw = cell.value or ''
            if current_raw == expected:
                will_write.append((actual_sheet, row, current_raw, proposed))
            elif current_raw == proposed:
                already_applied.append((actual_sheet, row, current_raw))
            else:
                discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current_raw!r}'))
            continue
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    for s, r, msg in discrepancies:
        print(f'  ⚠ {s}/J{r}: {msg}')
    if discrepancies: sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
