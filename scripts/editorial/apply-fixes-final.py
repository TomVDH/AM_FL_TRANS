#!/usr/bin/env python3
"""Apply user-approved FINAL batch fixes across 6 workbooks.

Tom sign-off 2026-05-13:

E0 CharacterProfiles (2 writes):
  - J81 fix → `Grove Ezel`
  - J95 override → `Exceem Ezel` (Tom's exact form; preserves X from original)

E2 World_A1 (0 writes):
  - J25 KEPT per Tom "keep current?" — `Stop. Sinds wanneer zijt GIJ` stays.
    Canon §5.4 would prefer `Stopt` but Tom calls; verify-skip.

E5 ZooMain (1 write):
  - J208 OVERRIDE → `'t Beste aan hem om een job zonder Machines te zoeken!`
    (Tom: voor→aan, vinden→zoeken, plus apostrophe fix)

E6 Nightmare (3 writes):
  - J63 OVERRIDE → Tom rewrites `Kom d'r maar naartoe gesjokt en neem...`
    to `Komt maar gauw naar hier en pakt...` (ge/gij imperative stem+t for
    both verbs, replaces gesjokt+drankspecialiteit)
  - J70 MATCH → Tom: "my edit there rules" — re-use Tom's E6_World J197
    truncated override exact string
  - J71 fix as proposed (slokkie→slokske, ezel lc kept)

E7 (4 writes; J11 skipped — already at J9's unify target):
  - J5 Chilling: fix Hemelvaarts-zang-der-Ezel-zielen
  - J22 Holding1: fix zij→zijt
  - J9 MeatProcessing: fix toont→laat zien (unify with J11)
  - J8 Skinning: fix electriciteit→elektriciteit

E8 TheGods (2 writes):
  - J7 fix Ezel-Zielen (Gods names checked: Hee/Haw/H'ii all canonical)
  - J24 OVERRIDE → Tom rewrite preserving archaic `der dragen` for divine
    register: `Gezien onze heilige last der dragen van het Universum,...`

E9 BadCave (1 write):
  - J40 fix STENEN-SPEL → KEIEN-SPEL (closes §8 game-name pattern)

Total: 13 cells across 6 workbooks. 2 cells KEPT/SKIPPED.

Run with --apply to commit local writes. Then push each workbook
separately via push-file.py.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')

# Match E6_World J197 final value exactly (set by apply-fixes-e6.py earlier)
WORLD_J197 = "'t Is altijd al mijn droom geweest om een plekske waar Ezels efkes kunnen pauzeren open te doen, gedoeme."

# (workbook_filename, sheet, row) → (expected_pre_image, proposed_post_image)
APPROVED = {
    # =========================================================================
    # E0 — CharacterProfiles
    # =========================================================================
    ('0_asses.masses_Manager+Intermissions+E0Proxy.xlsx', 'CharacterProfiles_localization', 81): (
        'Groffe Ezel',
        'Grove Ezel',
    ),
    ('0_asses.masses_Manager+Intermissions+E0Proxy.xlsx', 'CharacterProfiles_localization', 95): (
        'Excema Ezel',
        'Exceem Ezel',  # Tom exact form (preserves X from original; not standard Dutch but Tom-chosen)
    ),

    # =========================================================================
    # E5 — ZooMain
    # =========================================================================
    ('5_asses.masses_E5Proxy.xlsx', 'E5_ZooMain_localization', 208): (
        't Beste voor hem om een job zonder Machines te vinden!',
        "'t Beste aan hem om een job zonder Machines te zoeken!",
    ),

    # =========================================================================
    # E6 — Nightmare
    # =========================================================================
    # J63 Tom override
    ('6_asses.masses_E6Proxy.xlsx', 'E6_Nightmare_localization', 63): (
        "Kom d'r maar naartoe gesjokt en neem een slokje van m'n nieuwe drankspecialiteit: de Emmer Ale!",
        "Komt maar gauw naar hier en pakt een slokske van m'n nieuwe drankspecial: de Emmer Ale!",
    ),
    # J70: Tom: "my edit there rules" — match E6_World J197 final value
    ('6_asses.masses_E6Proxy.xlsx', 'E6_Nightmare_localization', 70): (
        "t Is altijd al m'n verdomde DROOM geweest om een plek te hebben waar wij Ezels even op adem kunnen komen en samen lekker een slokkie water kunnen delen.",
        WORLD_J197,
    ),
    # J71: proposed fix (slokkie→slokske, ezel lc kept)
    ('6_asses.masses_E6Proxy.xlsx', 'E6_Nightmare_localization', 71): (
        "Ik kan zeker wel een slokkie gebruiken en een ezel om z'n oor te kauwen.",
        "Ik kan zeker wel een slokske gebruiken en een ezel om z'n oor te kauwen.",
    ),

    # =========================================================================
    # E7
    # =========================================================================
    ('7_asses.masses_E7Proxy.xlsx', 'E7_Chilling_localization', 5): (
        'Wacht. Voor we vertrekken moeten we de ezelenhemelvaartszangderzielen zingen voor onze Kameraden.',
        'Wacht. Voor we vertrekken moeten we de Hemelvaarts-zang-der-Ezel-zielen zingen voor onze Kameraden.',
    ),
    ('7_asses.masses_E7Proxy.xlsx', 'E7_Holding1_localization', 22): (
        '@#$%&! Wie zij GIJ?!',
        '@#$%&! Wie zijt GIJ?!',
    ),
    ('7_asses.masses_E7Proxy.xlsx', 'E7_MeatProcessing_localization', 9): (
        "Tijd dat de Meesterstrateeg iedereen toont hoe 't moet!",
        "Tijd dat de Meesterstrateeg iedereen laat zien hoe 't moet!",
    ),
    ('7_asses.masses_E7Proxy.xlsx', 'E7_Skinning_localization', 8): (
        'Stamp en Triestige—electriciteit UIT!',
        'Stamp en Triestige—elektriciteit UIT!',
    ),

    # =========================================================================
    # E8 — TheGods
    # =========================================================================
    ('8_asses.masses_E8Proxy.xlsx', 'E8_TheGods_localization', 7): (
        "Dit is het Heiligdom van H'ii waar alle Ezel Zielen Heraanstelling ondergaan.",
        "Dit is het Heiligdom van H'ii waar alle Ezel-Zielen Heraanstelling ondergaan.",
    ),
    # J24: Tom override — preserves archaic `der dragen` for divine register
    ('8_asses.masses_E8Proxy.xlsx', 'E8_TheGods_localization', 24): (
        'Gezien onze goddelijk last der dragen van het Universum op onze rug, kunnen wij niet ingrijpen in de Materiële Wereld zonder al het leven in gevaar te brengen.',
        'Gezien onze heilige last der dragen van het Universum, kunnen wij niet in de Materiële Wereld ingrijpen zonder alle leven in gevaar te brengen.',
    ),

    # =========================================================================
    # E9 — BadCave
    # =========================================================================
    ('9_asses.masses_E9Proxy.xlsx', 'E9_BadCave_localization', 40): (
        "Ik dacht dat je STENEN-SPEL aan 't spelen was met je NONKEL!",
        "Ik dacht dat je KEIEN-SPEL aan 't spelen was met je NONKEL!",
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    # Group by workbook
    by_xlsx = {}
    for (xlsx, sheet, row), (exp, prop) in APPROVED.items():
        by_xlsx.setdefault(xlsx, []).append((sheet, row, exp, prop))

    total_will = 0
    total_already = 0
    total_disc = 0
    plans = {}  # xlsx -> list of (sheet, row, current, proposed)

    for xlsx, items in by_xlsx.items():
        print(f'\nLoading {xlsx}…')
        wb = openpyxl.load_workbook(REPO / 'excels' / xlsx)
        will_write = []
        already_applied = []
        discrepancies = []

        for sheet_name, row, expected, proposed in items:
            actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
                (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
            if actual_sheet is None:
                discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
                continue
            cell = wb[actual_sheet].cell(row=row, column=10)
            current = (cell.value or '').strip() if cell.value is not None else ''
            if current == expected.strip():
                will_write.append((actual_sheet, row, current, proposed))
            elif current == proposed.strip():
                already_applied.append((actual_sheet, row, current))
            else:
                discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

        plans[xlsx] = (wb, will_write, already_applied, discrepancies)
        total_will += len(will_write)
        total_already += len(already_applied)
        total_disc += len(discrepancies)

        print(f'  Status: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
        for s, r, msg in discrepancies:
            print(f'    ⚠ {s}/J{r}: {msg}')

    print(f'\n=== TOTAL: {total_will} write / {total_already} already / {total_disc} discrepant ===')
    if total_disc:
        sys.exit(1)
    if not total_will:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for xlsx, (wb, will_write, _, _) in plans.items():
        for sheet, row, current, proposed in will_write:
            print(f'  {xlsx} :: {sheet}/J{row}:')
            print(f'    -  {current[:100]!r}')
            print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for xlsx, (wb, will_write, _, _) in plans.items():
        for sheet, row, _, proposed in will_write:
            wb[sheet].cell(row=row, column=10).value = proposed
        wb.save(REPO / 'excels' / xlsx)
        print(f'✓ {xlsx}: wrote {len(will_write)} cells')


if __name__ == '__main__':
    main()
