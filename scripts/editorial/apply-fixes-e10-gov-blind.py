#!/usr/bin/env python3
"""Apply E10_Government blind-spot walk fixes — Tom sign-off 2026-05-13.

29 writes / 0 keeps. Cells across §5.4 ge/gij imperative sweep, §12.4 EN
bleed, §13 mistranslations, §7.1 Mensen cap, §9.3 punct, §6.7 cross-cell.

Tom decisions:
  - Group A (§5.4 sweep): all 8 approved — bare-stem Zeg/Ga → stem+t
    Zegt/Gaat for ge/gij speakers (Helpful, Edgy, Kick, Big)
  - Group B (§12.4): J11 default, J51 default `Wauw`, J213 `Toestemming`
  - Group C: J18 Tom override `Halt! Staan blijven!`,
    J19/J50/J83 defaults, J82 alternate `manier van leven`
  - Group D: J235 Tom override (full restructure, keeps `Zeg`),
    J252/J3 defaults
  - Group E pair decisions:
    - J52/J84 → Tom new form `Geheel zoals de dagen van weleer` (no `in`)
    - J59/J91 → unify on J91's `Gedaan met de ritjes!!` (Tom flipped)
    - J60/J92 → unify on J92's `Gij hebt allen`
    - J63/J95 → unify on `beknopt`
    - J65/J97 → unify on `te allen tijde`
    - J164/J192 → `Wilde` form (J173/J183 already at target)
    - J266/J267 → keep J266 archaic, only drop stray `te`
    - J130/J136 → unify on `zou gewild hebben`
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/10_asses.masses_E10Proxy.xlsx'

# Shared form for J52/J84 (Tom new)
J52_J84_FORM = 'Een wonderbaarlijke samenwerking! Geheel zoals de dagen van weleer.'
# Shared form for J60/J92 (J92's form)
J60_J92_FORM = 'STILTE. Gij hebt allen veel te zeggen, maar wij zullen dit Manifest op ordelijke wijze samenstellen.'

APPROVED = {
    # =================================================================
    # Group A — §5.4 ge/gij imperative sweep (8 writes)
    # =================================================================
    ('E10_Government_localization', 45): (
        "Zeg ne keer iets in 't Mens!",
        "Zegt ne keer iets in 't Mens!",
    ),
    ('E10_Government_localization', 57): (
        'Zeg de Mensen dat ze ons steunen—',
        'Zegt de Mensen dat ze ons steunen—',
    ),
    ('E10_Government_localization', 88): (
        'Zeg tegen de Mensen da we zóveel geleden hebben—',
        'Zegt tegen de Mensen da we zóveel geleden hebben—',
    ),
    ('E10_Government_localization', 89): (
        'Zeg de Mensen dat ze moeten steunen—',
        'Zegt de Mensen dat ze moeten steunen—',
    ),
    ('E10_Government_localization', 129): (
        'Zeg tegen de Mensen dat we NIET MEER WILLEN VECHTEN!',
        'Zegt tegen de Mensen dat we NIET MEER WILLEN VECHTEN!',
    ),
    ('E10_Government_localization', 134): (
        '@#$%&! Ga met de ANDEREN praten!',
        '@#$%&! Gaat met de ANDEREN praten!',
    ),
    ('E10_Government_localization', 142): (
        'Zeg tegen de Mensen dat ze overal hun Fabrieken sluiten...',
        'Zegt tegen de Mensen dat ze overal hun Fabrieken sluiten...',
    ),
    ('E10_Government_localization', 240): (
        'Zeg tegen de Mensen dat ze onze creatieve inspanningen moeten steunen en waarderen.',
        'Zegt tegen de Mensen dat ze onze creatieve inspanningen moeten steunen en waarderen.',
    ),

    # =================================================================
    # Group B — §12.4 English bleed (3 writes)
    # =================================================================
    ('E10_Government_localization', 11): (
        'Meer over dit verhaal aan de top van het uur.',
        'Meer over dit verhaal op het hele uur.',
    ),
    ('E10_Government_localization', 51): (
        'Wowee... oké...',
        'Wauw... oké...',
    ),
    ('E10_Government_localization', 213): (
        'Consent is wat elke Ezel wil.',
        'Toestemming is wat elke Ezel wil.',
    ),

    # =================================================================
    # Group C — §13 mistranslation (5 writes)
    # =================================================================
    # J18 Tom override
    ('E10_Government_localization', 18): (
        'Halt, blijf staan!',
        'Halt! Staan blijven!',
    ),
    ('E10_Government_localization', 19): (
        'Geen Mensen toegelaten.',
        'Geen Mensen binnen toegelaten.',
    ),
    ('E10_Government_localization', 50): (
        'Jonge Profetes, staat u ons toe te vragen — wilt u onze eisen vertolken aan de Mensen?',
        'Jonge Profeet, wilt u onze eisen vertolken aan de Mensen?',
    ),
    # J82 alternate `manier van leven`
    ('E10_Government_localization', 82): (
        'Zodra ze de wereld leren zien door jullie ogen, zullen ze samen met ons de mensen wereldwijd tot nadenken brengen.',
        'Zodra ze de wereld leren zien door jullie ogen, zullen ze samen met ons de manier van leven van de Mensen wereldwijd veranderen.',
    ),
    ('E10_Government_localization', 83): (
        "Als ge mij uw eisen vertelt, kameraden, dan geef 'k die door aan de burgers van Technopolis — in hun eigen taal.",
        "Als ge mij uw eisen vertelt, dan geef 'k die door aan de burgers van Technopolis — in hun eigen taal.",
    ),

    # =================================================================
    # Group D — §7.1 + §9.3 (3 writes)
    # =================================================================
    # J235 Tom full override (keeps `Zeg`, restructures whole sentence)
    ('E10_Government_localization', 235): (
        'Zeg tegen de mensen dat ze alle schulden moeten kwijtschelden waarvan ze denken dat wij die bij hen hebben.',
        'Zeg tegen de Mensen dat ze al de schulden, die ze denken dat we bij ze hebben, kwijt te schelden.',
    ),
    ('E10_Government_localization', 252): (
        'Bent gij bereid om U tot de mensen te wenden?',
        'Bent gij bereid om U tot de Mensen te wenden?',
    ),
    ('E10_Government_localization', 3): (
        'Geen Mens komt erin en geen Ezel gaat eruit tot wij onze volgende stap hebben bepaald!',
        'Geen Mens komt erin en geen Ezel gaat eruit tot wij onze volgende stap hebben bepaald.',
    ),

    # =================================================================
    # Group E — §6.7 cross-cell pairs (10 writes)
    # =================================================================
    # J52/J84 → Tom new shared form
    ('E10_Government_localization', 52): (
        'Een wonderbaarlijke samenwerking! Geheel zoals in de dagen van weleer.',
        J52_J84_FORM,
    ),
    ('E10_Government_localization', 84): (
        'Een wonderbaarlijke samenwerking! Net als in de goede oude tijd.',
        J52_J84_FORM,
    ),
    # J59 → J91's form `Gedaan met de ritjes!!`
    ('E10_Government_localization', 59): (
        'Geen ritten meer!!',
        'Gedaan met de ritjes!!',
    ),
    # J60 → J92's `Gij hebt allen`
    ('E10_Government_localization', 60): (
        'STILTE. Gij allen hebt veel te zeggen, maar wij zullen dit Manifest op ordelijke wijze samenstellen.',
        J60_J92_FORM,
    ),
    # J95 → J63's `beknopt`
    ('E10_Government_localization', 95): (
        'Profeet, wij dienen onze eisen bondig te houden.',
        'Profeet, wij dienen onze eisen beknopt te houden.',
    ),
    # J97 → J65's `te allen tijde`
    ('E10_Government_localization', 97): (
        'Nadat gij Uw Manifest hebt gevuld, kunt gij steeds een vorig idee vervangen door een nieuw te kiezen.',
        'Nadat gij Uw Manifest hebt gevuld, kunt gij te allen tijde een vorig idee vervangen door een nieuw te kiezen.',
    ),
    # J164 → Wilde
    ('E10_Government_localization', 164): (
        "Wil de 't weten?",
        "Wilde 't weten?",
    ),
    # J192 → Wilde
    ('E10_Government_localization', 192): (
        "Wil de 't nog ne keer horen, mijn verhaal?",
        "Wilde 't nog ne keer horen, mijn verhaal?",
    ),
    # J266 → keep archaic, drop stray `te`
    ('E10_Government_localization', 266): (
        'Uw zuster heeft haar Ziel aan ons geschonken, opdat wij tot deze Wereld mochten wederkeren en haar te redden.',
        'Uw zuster heeft haar Ziel aan ons geschonken, opdat wij tot deze Wereld mochten wederkeren en haar redden.',
    ),
    # J136 → J130's `zou gewild hebben`
    ('E10_Government_localization', 136): (
        'Dat is wat Mega Ezel gewild zou hebben.',
        'Dat is wat Mega Ezel zou gewild hebben.',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    for s, r, msg in discrepancies:
        print(f'  ⚠ {s}/J{r}: {msg}')
    if discrepancies: sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:80]!r}')
        print(f'    +  {proposed[:80]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
