#!/usr/bin/env python3
"""Rescan every DRIFT cell I claimed in audit-2026-05-12-deep-eyeball.md.

Parses the deep-eyeball doc for cell references tagged DRIFT, re-reads the
xlsx verbatim for each cell, and writes per-episode rescan reports plus an
INDEX. Output goes to data/editorial/rescan-2026-05-12-E{N}-drift.md and
data/editorial/rescan-2026-05-12-INDEX.md.

For each rescanned cell:
  - row number
  - speaker (verbatim from column A key suffix for E5–E10, column B for E0–E4)
  - EN verbatim from column C
  - NL verbatim from column J (10th column)
  - DRIFT claim from deep-eyeball doc (the bullet header line)
  - VERIFIED / DISCREPANCY status (whether the cell still exists and has content)
  - PUSH status from _PUSH-LOG.md

LOGGING ONLY — no xlsx edits, no canon edits.
"""
import re
import sys
import importlib.util
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
DEEP_EYEBALL = REPO / 'data/editorial/audit-2026-05-12-deep-eyeball.md'
OUT_DIR = REPO / 'data/editorial'
RESCAN_DATE = '2026-05-12'

# Load orchestrator (push log + speaker resolution)
spec = importlib.util.spec_from_file_location(
    'audit', REPO / 'scripts/editorial/comprehensive-audit.py')
audit = importlib.util.module_from_spec(spec)
spec.loader.exec_module(audit)

EXCELS = {n: (lbl, conv) for n, lbl, conv in audit.EXCELS}


def parse_deep_eyeball():
    """Parse deep-eyeball.md and return DRIFT entries.

    Returns list of dicts: {episode, sheet, row, claim_line, tier, source_line_num}.
    Tier is filtered to DRIFT only.
    """
    text = DEEP_EYEBALL.read_text().splitlines()
    entries = []
    current_episode = None
    current_sheet = None

    # Patterns
    re_episode = re.compile(r'^## E(\d+)')
    re_sheet = re.compile(r'^### E(\d+)\s*/\s*([A-Za-z0-9_]+_localization|[A-Za-z0-9_]+_localiza|[A-Za-z0-9_]+_localizatio|[0-9]x_Intermission_localization|ManagerScene_localization|CharacterProfiles_localization)')
    re_sheet_loose = re.compile(r'^### E(\d+)\s*/\s*(\S+)')
    re_cell_simple = re.compile(r'^\*\*J(\d+)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`')
    re_cell_pair = re.compile(r'^\*\*J(\d+)\s+vs\s+J(\d+)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`')
    re_cell_range = re.compile(r'^\*\*J(\d+)\s*[–-]\s*J(\d+)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`')
    re_cell_slash = re.compile(r'^\*\*J(\d+)\s*/\s*J(\d+)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`')
    re_cell_triple = re.compile(r'^\*\*J(\d+)\s*/\s*J(\d+)\s*/\s*J(\d+)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`')
    # Loose cell pattern for full-walk addendum (no DRIFT tag inline)
    re_cell_addendum = re.compile(r'^\*\*J(\d+)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`')

    # Track context in the "Full-Walk Findings" appendix (different sheet-header format)
    re_full_sheet = re.compile(r'^### E(\d+)\s*/\s*(\S+)')

    for i, line in enumerate(text, 1):
        m = re_episode.match(line)
        if m:
            current_episode = int(m.group(1))
            current_sheet = None
            continue
        m = re_sheet_loose.match(line)
        if m:
            # Sheet headers always update both episode + sheet (the Full-Walk addendum
            # mixes episodes in one section, so we can't trust current_episode here)
            current_episode = int(m.group(1))
            current_sheet = m.group(2).rstrip()
            continue

        # Inline format (sheet in cell-header) — check FIRST since it doesn't need current_sheet
        # **E7_Holding1 / J22** — `DRIFT` — ...
        # **E6_World / J39 vs J46** — `DRIFT` — ...
        # **E10_Government / J68 / J100** — `DRIFT` — ...
        # Negative lookahead `(?!J\d)` prevents J21/J53/J68 patterns from being treated
        # as sheet names — those are bare cell-range cells where current_sheet should drive.
        m_inline = re.match(
            r'^\*\*(?:E\d+\s*/\s*)?(?!J\d)([A-Za-z][A-Za-z0-9_]+)\s*/\s*(J\d+(?:\s*(?:[/+]|vs|–|-)\s*J\d+)*)\*\*\s*(?:—|--|-)\s*`(DRIFT|VERIFY|NOTE)`',
            line)
        if m_inline and m_inline.group(3) == 'DRIFT' and current_episode is not None:
            sheet_name = m_inline.group(1)
            row_spec = m_inline.group(2)
            row_nums = [int(x) for x in re.findall(r'J(\d+)', row_spec)]
            if '–' in row_spec or '-' in row_spec:
                if len(row_nums) == 2:
                    row_nums = list(range(row_nums[0], row_nums[1] + 1))
            sheet_resolved = sheet_name if sheet_name.endswith('_localization') else sheet_name + '_localization'
            for r in row_nums:
                entries.append({
                    'episode': current_episode,
                    'sheet': sheet_resolved,
                    'row': r,
                    'claim_line': line.strip(),
                    'tier': 'DRIFT',
                    'source_line_num': i,
                })
            continue

        if current_episode is None or current_sheet is None:
            continue

        for regex in (re_cell_triple, re_cell_range, re_cell_pair, re_cell_slash):
            m = regex.match(line)
            if m:
                groups = m.groups()
                tier = groups[-1]
                if tier != 'DRIFT':
                    break
                rows = []
                if regex is re_cell_range:
                    rows = list(range(int(groups[0]), int(groups[1]) + 1))
                else:
                    rows = [int(g) for g in groups[:-1]]
                for r in rows:
                    entries.append({
                        'episode': current_episode,
                        'sheet': current_sheet,
                        'row': r,
                        'claim_line': line.strip(),
                        'tier': tier,
                        'source_line_num': i,
                    })
                break
        else:
            m = re_cell_simple.match(line)
            if m and m.group(2) == 'DRIFT':
                entries.append({
                    'episode': current_episode,
                    'sheet': current_sheet,
                    'row': int(m.group(1)),
                    'claim_line': line.strip(),
                    'tier': 'DRIFT',
                    'source_line_num': i,
                })

    # Dedupe by (episode, sheet, row) — keep the LATEST claim (later in doc)
    deduped = {}
    for e in entries:
        key = (e['episode'], e['sheet'], e['row'])
        deduped[key] = e
    return sorted(deduped.values(), key=lambda x: (x['episode'], x['sheet'], x['row']))


def normalize_sheet_name(ep, claimed_name, workbook_sheets):
    """Resolve a sheet name from the claim into the actual workbook sheet name.

    Handles truncated names (Excel 31-char limit) and per-episode prefix variants.
    """
    # Try exact match
    if claimed_name in workbook_sheets:
        return claimed_name
    # Try prefix match
    for s in workbook_sheets:
        if s.startswith(claimed_name) or claimed_name.startswith(s):
            return s
    # Try with _localization suffix variants
    base = claimed_name.rstrip('_').rstrip()
    for s in workbook_sheets:
        s_base = s.rstrip('_').rstrip()
        if s_base.startswith(base[:31]) or base.startswith(s_base[:31]):
            return s
    return None


def read_cell(wb, sheet_name, row_num, convention):
    """Read EN+NL+speaker verbatim from xlsx row."""
    ws = wb[sheet_name]
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if r == row_num:
            if len(row) < 10:
                return None
            speaker = audit.get_speaker(row, convention)
            en = (row[2] or '').strip() if row[2] is not None else ''
            nl = (row[9] or '').strip() if row[9] is not None else ''
            return {'speaker': speaker, 'en': en, 'nl': nl}
    return None


def md_inline(s):
    if not s:
        return '∅'
    return str(s).replace('`', '\\`').replace('|', '\\|').replace('\n', ' ⏎ ')


def write_episode_rescan(ep, entries, push_log, wb, ep_label, convention, out_path):
    L = [f'# Rescan — E{ep} DRIFT cells (verbatim re-read)']
    L.append('')
    L.append(f'_Rescan date: {RESCAN_DATE} — source: `audit-2026-05-12-deep-eyeball.md` — workbook: `excels/{ep}_asses.masses_{ep_label}.xlsx`_')
    L.append('')
    L.append(f'**{len(entries)} DRIFT cells claimed** for this episode. Each rescanned verbatim below.')
    L.append('')

    # Group by sheet
    by_sheet = {}
    for e in entries:
        by_sheet.setdefault(e['sheet'], []).append(e)
    by_sheet = dict(sorted(by_sheet.items()))

    workbook_sheets = wb.sheetnames

    for claimed_sheet, sheet_entries in by_sheet.items():
        resolved_sheet = normalize_sheet_name(ep, claimed_sheet, workbook_sheets)
        L.append(f'## {claimed_sheet}')
        L.append('')
        if resolved_sheet is None:
            L.append(f'⚠ **Sheet not found in workbook.** Claimed: `{claimed_sheet}`. Available: {workbook_sheets}')
            L.append('')
            continue
        if resolved_sheet != claimed_sheet:
            L.append(f'_(Resolved to actual workbook sheet: `{resolved_sheet}`)_')
            L.append('')
        L.append(f'_{len(sheet_entries)} DRIFT cell(s) in this sheet._')
        L.append('')

        for entry in sheet_entries:
            r = entry['row']
            current = read_cell(wb, resolved_sheet, r, convention)
            history = push_log.get((ep, resolved_sheet, r), [])
            push_status = 'Y' if history else 'N'
            push_rule = history[-1].get('rule', '') if history else ''

            L.append(f'### J{r}')
            L.append('')
            L.append(f'- **Claim** (deep-eyeball L{entry["source_line_num"]}): {md_inline(entry["claim_line"])}')
            if current is None:
                L.append(f'- **Status:** ⚠ DISCREPANCY — row J{r} not found in current xlsx')
            else:
                L.append(f'- **Status:** ✓ VERIFIED — re-read from xlsx')
                L.append(f'- **Speaker:** `{md_inline(current["speaker"])}`')
                L.append(f'- **EN:** `{md_inline(current["en"])}`')
                L.append(f'- **NL:** `{md_inline(current["nl"])}`')
            if push_status == 'Y':
                L.append(f'- **Push:** `[{history[-1]["id"]}]` — rule: `{md_inline(push_rule)}`')
                pushed_val = history[-1].get('pushed', '')
                if pushed_val:
                    L.append(f'- **Pushed value:** `{md_inline(pushed_val)}`')
            else:
                L.append(f'- **Push:** never pushed (no entry in `_PUSH-LOG.md`)')
            L.append('')

    out_path.write_text('\n'.join(L))


def write_index(ep_stats, out_dir):
    L = [f'# Rescan INDEX — DRIFT cells from deep-eyeball audit ({RESCAN_DATE})']
    L.append('')
    L.append(f'_Rescan source: `data/editorial/audit-2026-05-12-deep-eyeball.md`. Every cell tagged \\`DRIFT\\` has been verbatim re-read from xlsx and written to per-episode rescan files. No edits performed._')
    L.append('')

    total = sum(e['count'] for e in ep_stats)
    L.append(f'## Totals')
    L.append('')
    L.append(f'- Episodes with DRIFT: **{len(ep_stats)}/11**')
    L.append(f'- Total DRIFT cells rescanned: **{total}**')
    L.append('')
    L.append('## Per-episode breakdown')
    L.append('')
    L.append('| Episode | DRIFT cells | Sheets touched | Rescan report |')
    L.append('|---|---|---|---|')
    for e in ep_stats:
        link = f'rescan-{RESCAN_DATE}-E{e["ep"]}-drift.md'
        L.append(f'| E{e["ep"]} ({e["label"]}) | {e["count"]} | {e["sheets"]} | [{link}]({link}) |')
    L.append('')

    L.append('## Reading order')
    L.append('')
    L.append('Each per-episode rescan file is grouped by sheet, then by J-row. For each cell:')
    L.append('')
    L.append('- **Claim** — the original deep-eyeball line (with source line number for traceability)')
    L.append('- **Status** — `VERIFIED` (cell re-read from xlsx) or `DISCREPANCY` (row not found)')
    L.append('- **Speaker / EN / NL** — verbatim from `openpyxl` re-read')
    L.append('- **Push** — push-log history if the cell was previously pushed; pushed value if available')
    L.append('')
    L.append('## How to use this')
    L.append('')
    L.append('1. Open the per-episode rescan file for the episode you\'re editing.')
    L.append('2. Each DRIFT cell shows verbatim current EN+NL — you can edit the xlsx cell directly and the claim tells you what to fix.')
    L.append('3. For cells with push history, double-check the push rule before editing — the cell may be in canonical-exception state.')
    L.append('4. After fixes land, replay the regex audit: `python3 scripts/editorial/comprehensive-audit.py`.')
    L.append('')

    (out_dir / f'rescan-{RESCAN_DATE}-INDEX.md').write_text('\n'.join(L))


def main():
    print('Parsing deep-eyeball doc for DRIFT cells…')
    entries = parse_deep_eyeball()
    print(f'  Found {len(entries)} DRIFT cells.')

    # Group by episode
    by_ep = {}
    for e in entries:
        by_ep.setdefault(e['episode'], []).append(e)

    push_log = audit.parse_push_log(audit.PUSH_LOG)

    ep_stats = []
    for ep in sorted(by_ep.keys()):
        ep_label, conv = EXCELS[ep]
        xlsx = REPO / 'excels' / f'{ep}_asses.masses_{ep_label}.xlsx'
        wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
        out_path = OUT_DIR / f'rescan-{RESCAN_DATE}-E{ep}-drift.md'
        write_episode_rescan(ep, by_ep[ep], push_log, wb, ep_label, conv, out_path)
        sheets_touched = len(set(e['sheet'] for e in by_ep[ep]))
        ep_stats.append({
            'ep': ep, 'label': ep_label,
            'count': len(by_ep[ep]),
            'sheets': sheets_touched,
        })
        print(f'  E{ep}: {len(by_ep[ep])} cells across {sheets_touched} sheet(s) → {out_path.name}')
        wb.close()

    write_index(ep_stats, OUT_DIR)
    print(f'  Wrote: rescan-{RESCAN_DATE}-INDEX.md')


if __name__ == '__main__':
    main()
