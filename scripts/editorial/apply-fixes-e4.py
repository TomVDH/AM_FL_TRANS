#!/usr/bin/env python3
"""Apply user-approved E4 fixes to local xlsx (then push via push-file.py).

User sign-off 2026-05-13: all 23 cells approved, with overrides on J46/J71/J106/J235
(Astral), J6 (KicksConfession option c), J6 (Mine1F_Exit).

Total: 23 cells fixed, 0 skipped.

Safety: pre-image check — current cell must match expected OR proposed (idempotent)
before write. Aborts on unexpected current.

Usage:
  python3 scripts/editorial/apply-fixes-e4.py             # dry-run
  python3 scripts/editorial/apply-fixes-e4.py --apply     # write to local xlsx
"""
import sys
from pathlib import Path

import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/4_asses.masses_E4Proxy.xlsx'

# (sheet, row): (expected_current_nl, proposed_nl)
APPROVED = {
    # E4_AstralPlaneMain_localization (14)
    ('E4_AstralPlaneMain_localization', 46): (
        'Hé Niewelingetjes, we wedden op hoeveel Ezels er vandaag de lucht uit gaan vallen. Doen jullie mee?',
        'Hé Nieuwelingetjes, we wedden op hoeveel Ezels er vandaag uit de lucht gaan vallen. Doen jullie mee?',
    ),
    ('E4_AstralPlaneMain_localization', 49): (
        'En jij? Wilde gokken?',
        'En jij? Wil je gokken?',
    ),
    ('E4_AstralPlaneMain_localization', 71): (
        '...En mijn laatste woorden, nadat ik de Spuit was gegeven, waren "Herinneer de wereld aan de belangan van de Ezels!"',
        '...En mijn laatste woorden, nadat ik die Spuit kreeg, waren "Herinner de Wereld aan de Belangen van de Ezels!"',
    ),
    ('E4_AstralPlaneMain_localization', 94): (
        'Ik wou dat Trouwe wat meer geduld had wanneer ze aankwam.',
        'Ik wou dat Trouwe wat meer geduld had toen ze aankwam.',
    ),
    ('E4_AstralPlaneMain_localization', 106): (
        "Let's go let's go let's go!",
        'Komaan komaan komaan!',
    ),
    ('E4_AstralPlaneMain_localization', 117): (
        "Let's go!",
        'Vooruit!',
    ),
    ('E4_AstralPlaneMain_localization', 167): (
        'Wij waren simpele dorpsezels wanneer de Mensen onze Kudde gebruikte om voorraden in hun oorlogen te zeulen.',
        'Wij waren simpele dorpsezels toen de Mensen onze Kudde gebruikten om voorraden in hun oorlogen te zeulen.',
    ),
    ('E4_AstralPlaneMain_localization', 178): (
        'WHAT DE FUCK, IK KAN MIJN EIGEN LICHAAM ZIEN.',
        'WAT DE FUCK, IK KAN MIJN EIGEN LICHAAM ZIEN.',
    ),
    ('E4_AstralPlaneMain_localization', 183): (
        'Hey Nieuwkomertje, wilde poepen? Ge riekt lekker.',
        'Hey Nieuwkomertje, wilt ge poepen? Ge riekt lekker.',
    ),
    ('E4_AstralPlaneMain_localization', 201): (
        'Wow wow wow! Waar denk jij wel niet waar je heen gaat?',
        'Wow wow wow! Waar denk jij wel niet dat je heen gaat?',
    ),
    ('E4_AstralPlaneMain_localization', 220): (
        'GOEDEAVOND MEDE-EZELS!!!!',
        'GOEIENAVOND MEDE-EZELS!!!!',
    ),
    ('E4_AstralPlaneMain_localization', 228): (
        'Lijkt dat we allebeide het Concert gaan missen dan!',
        'Lijkt dat we allebei het Concert gaan missen dan!',
    ),
    ('E4_AstralPlaneMain_localization', 234): (
        'Als je met Wise Ass wil spreken, onthoud gewoon dat ze hier al véél langer dan ons allemaal zit.',
        'Als je met Wijze Ezel wil spreken, onthoud gewoon dat ze hier al véél langer dan ons allemaal zit.',
    ),
    ('E4_AstralPlaneMain_localization', 235): (
        'Wees respectvol. Wise Ass heeft een antwoord op alles.',
        'Toon respect. Wijze Ezel heeft op alles een antwoord.',
    ),

    # E4_HerdSplits_localization (6)
    ('E4_HerdSplits_localization', 29): (
        "'t Is genoeg geweest — ik ben hier KLAAR mee. Tijd voor een nieuw plan.",
        "'t Is genoeg geweest — ik ben hier KLAAR mee. Tijd voor een nieuw Plan.",
    ),
    ('E4_HerdSplits_localization', 32): (
        'DRIE. We overtuigen de Mensen om hun Machines op te geven, we sterven niet in dit hol, en we pakken onze jobs TERUG.',
        'DRIE. We overtuigen de Mensen om hun Machines op te geven, we sterven niet in dit hol, en we pakken onze JOBS TERUG.',
    ),
    ('E4_HerdSplits_localization', 35): (
        'Kameraden, ik blijf hier met mijn veulentje. Waar we samen horen.',
        'Kameraden, ik blijf hier met mijn Veulentje. Waar we samen horen.',
    ),
    ('E4_HerdSplits_localization', 59): (
        'Ik ZWEER op het lijk van mijn ZUSTER — ik raak nooit nog een mens aan.',
        'Ik ZWEER op het lijk van mijn ZUSTER — ik raak nooit nog een Mens aan.',
    ),
    ('E4_HerdSplits_localization', 77): (
        "VIER! Let's GO!",
        'VIER! VOORUIT!',
    ),
    ('E4_HerdSplits_localization', 81): (
        'Een revolutie vergt offers.',
        'Een Revolutie vergt offers.',
    ),

    # E4_KicksConfession_localization (1) — user picked option (c)
    ('E4_KicksConfession_localization', 6): (
        'MOEST ik MEEHELPEN terwijl Snotje op uitkijk stond.',
        'IK MOEST VAN HEM MEEHELPEN terwijl Snotje op uitkijk stond.',
    ),

    # E4_Mine1F_Exit_localization (1) — user override (lowercase mensen, korten with -n)
    ('E4_Mine1F_Exit_localization', 6): (
        'Ge overdrijft. De Mensen maken korte metten met u vanaf de eerste keer ze de kans hebben.',
        'Ge overdrijft. De mensen maken korten metten met u op de eerste kans dat ze krijgen.',
    ),

    # E4_Mine1F_localization (1)
    ('E4_Mine1F_localization', 22): (
        'Heb geen vrees, Kameraad, Ik vertrouw dat de Goden over jou zullen waken.',
        'Heb geen vrees, Kameraad, ik vertrouw dat de Goden over jou zullen waken.',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    print(f'  Tabs: {wb.sheetnames}\n')

    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        if sheet_name not in wb.sheetnames:
            match = next((s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
            if match is None:
                discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
                continue
            actual_sheet = match
        else:
            actual_sheet = sheet_name

        ws = wb[actual_sheet]
        cell = ws.cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED-CURRENT: {current!r} (expected {expected!r})'))

    print(f'Status:')
    print(f'  Will write:       {len(will_write)} cell(s)')
    print(f'  Already applied:  {len(already_applied)} cell(s)')
    print(f'  Discrepancies:    {len(discrepancies)} cell(s)')
    print()

    if discrepancies:
        print('⚠ DISCREPANCIES — aborting before write:')
        for sheet, row, msg in discrepancies:
            print(f'  {sheet} / J{row}: {msg}')
        sys.exit(1)

    if not will_write:
        print('All cells already at proposed values — nothing to write.')
        return

    print('Proposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet} / J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply to write to xlsx.')
        return

    print('\nApplying writes…')
    for sheet, row, _, proposed in will_write:
        ws = wb[sheet]
        ws.cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells to {XLSX.name}')


if __name__ == '__main__':
    main()
