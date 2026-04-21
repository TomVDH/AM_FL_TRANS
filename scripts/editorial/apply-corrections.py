#!/usr/bin/env python3
"""
apply-corrections.py — apply a character corrections JSON file to the source xlsx sheets.

Corrections JSON format:
{
  "character": "Trusty Ass",
  "corrections": [
    {
      "id": "T-1",
      "file": "1_asses.masses_E1Proxy.xlsx",
      "sheet": "E1_Plowing_localization",
      "cell": "J5",
      "english": "...source EN for safety check...",
      "current_nl": "...expected current value...",
      "proposed_nl": "...new value..."
    },
    ...
  ]
}

USAGE
  python3 scripts/editorial/apply-corrections.py data/editorial/trusty-corrections.json          # dry-run
  python3 scripts/editorial/apply-corrections.py data/editorial/trusty-corrections.json --apply  # write

SAFETY
  - Dry-run is the default. --apply must be explicit.
  - Before touching a cell, verifies CURRENT value matches the correction's `current_nl` field.
    If mismatch → skips that row and reports, so stale corrections don't overwrite fresh edits.
  - If EN in sheet doesn't match correction.english → skipped (row might have shifted).
  - Opens each xlsx once, applies all its corrections, saves once (atomic-ish per file).
  - Uses openpyxl with keep_vba=False; styles preserved.
"""

import json
import sys
from pathlib import Path
import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_DIR = PROJECT_ROOT / 'excels'


def cell_col_row(cell_ref):
    """'J5' -> ('J', 5)."""
    i = 0
    while i < len(cell_ref) and cell_ref[i].isalpha():
        i += 1
    return cell_ref[:i], int(cell_ref[i:])


def main():
    if len(sys.argv) < 2:
        print('Usage: apply-corrections.py <corrections.json> [--apply]')
        sys.exit(2)
    corrections_path = Path(sys.argv[1])
    apply = '--apply' in sys.argv[2:]
    if not corrections_path.exists():
        print(f'ERROR: {corrections_path} not found')
        sys.exit(1)

    data = json.loads(corrections_path.read_text(encoding='utf-8'))
    character = data.get('character', '?')
    corrections = data.get('corrections', [])
    print(f'Character: {character}')
    print(f'Corrections: {len(corrections)}')
    print(f'Mode: {"APPLY" if apply else "DRY-RUN (no writes)"}')
    print('-' * 78)

    # Group by file
    by_file = {}
    for c in corrections:
        by_file.setdefault(c['file'], []).append(c)

    total_ok = 0
    total_skip = 0
    total_err = 0

    for fname, items in by_file.items():
        xlsx_path = XLSX_DIR / fname
        if not xlsx_path.exists():
            print(f'  ❌ FILE MISSING: {xlsx_path}')
            total_err += len(items)
            continue

        wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
        file_changes = []

        for c in items:
            cid = c.get('id', '?')
            sheet_name = c['sheet']
            cell_ref = c['cell']
            col_letter, row = cell_col_row(cell_ref)

            if sheet_name not in wb.sheetnames:
                print(f'  ❌ {cid}  sheet missing: {sheet_name}')
                total_err += 1
                continue
            ws = wb[sheet_name]

            # Safety: verify EN in sheet matches correction.english
            # EN is in Column C (Standard) per project convention
            actual_en = ws[f'C{row}'].value
            actual_nl = ws[f'{col_letter}{row}'].value

            exp_en = c.get('english')
            exp_cur_nl = c.get('current_nl')
            new_nl = c.get('proposed_nl')

            en_norm = (actual_en or '').strip()
            exp_en_norm = (exp_en or '').strip()

            if exp_en and en_norm != exp_en_norm:
                print(f'  ⚠️  {cid}  EN mismatch — skipping')
                print(f'      sheet:       {sheet_name} {cell_ref}')
                print(f'      expected EN: {exp_en_norm[:100]}')
                print(f'      actual EN:   {en_norm[:100]}')
                total_skip += 1
                continue

            nl_norm = (actual_nl or '').strip()
            exp_cur_norm = (exp_cur_nl or '').strip()

            if exp_cur_nl is not None and nl_norm != exp_cur_norm:
                print(f'  ⚠️  {cid}  current NL mismatch — skipping to avoid overwriting newer edit')
                print(f'      sheet:       {sheet_name} {cell_ref}')
                print(f'      expected NL: {exp_cur_norm[:100]}')
                print(f'      actual NL:   {nl_norm[:100]}')
                total_skip += 1
                continue

            if nl_norm == (new_nl or '').strip():
                print(f'  ✓ {cid}  already compliant — skipping [{sheet_name} {cell_ref}]')
                total_skip += 1
                continue

            # Plan the change
            print(f'  {"→" if apply else "…"} {cid}  [{sheet_name} {cell_ref}]')
            print(f'      BEFORE: {nl_norm[:140]}')
            print(f'      AFTER:  {new_nl[:140]}')
            file_changes.append((ws, f'{col_letter}{row}', new_nl))

        if apply and file_changes:
            for ws, addr, val in file_changes:
                ws[addr].value = val
            wb.save(str(xlsx_path))
            print(f'  💾 saved {len(file_changes)} change(s) → {xlsx_path.name}')
            total_ok += len(file_changes)
        elif file_changes and not apply:
            total_ok += len(file_changes)
            print(f'  (dry-run: would save {len(file_changes)} change(s) → {xlsx_path.name})')

    print('-' * 78)
    print(f'Summary: {total_ok} would-apply · {total_skip} skipped · {total_err} errors')
    if not apply and total_ok:
        print('\nRe-run with --apply to write.')


if __name__ == '__main__':
    main()
