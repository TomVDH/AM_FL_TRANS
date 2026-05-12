#!/usr/bin/env python3
"""E10 row-alignment scan — flags potential translation bugs.

Heuristics:
  1. LENGTH-RATIO: NL/EN char ratio outside [0.4, 2.5] (where EN >= 8 chars).
     Catches: very short NL for long EN, or bloated NL.
  2. CROSS-ROW PASTE: two cells in same sheet have NL similarity >= 0.80
     while EN similarity is <= 0.30. Catches: paraphrase-paste bugs like
     J74/J75 (Cole→Resentful) — duplicate NL content across different ENs.
  3. EMPTY NL with non-empty EN.

Output: per-sheet findings. Manual review required — heuristics produce
false positives (short standardized utterances cluster on similarity;
templated UI strings cluster on length).
"""
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path
import openpyxl

XLSX = Path(sys.argv[1] if len(sys.argv) > 1 and sys.argv[1].endswith('.xlsx')
            else 'excels/10_asses.masses_E10Proxy.xlsx')

LEN_RATIO_LO = 0.4
LEN_RATIO_HI = 2.5
MIN_EN_CHARS = 8
NL_SIM_THRESHOLD = 0.75
EN_DIFF_THRESHOLD = 0.55
MIN_LEN_FOR_CROSS_ROW = 25  # ignore very short cells for cross-row check


def extract_speaker(key):
    if not key:
        return ''
    parts = str(key).split('.')
    return parts[-1].strip() if len(parts) >= 2 else str(key).strip()


def norm(s):
    return re.sub(r'\s+', ' ', (s or '').strip().lower())


def sim(a, b):
    return SequenceMatcher(None, a, b).ratio()


def scan_sheet(ws, sheet_name):
    rows = []  # (row_index, speaker, en, nl)
    for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) < 10:
            continue
        speaker = extract_speaker(row[0] or '')
        en = (row[2] or '').strip()
        nl = (row[9] or '').strip()
        if not en and not nl:
            continue
        rows.append((r, speaker, en, nl))

    # Heuristic 1 + 3: per-cell length/empty
    cell_flags = []
    for r, sp, en, nl in rows:
        if en and not nl:
            cell_flags.append((r, 'EMPTY-NL', sp, en, nl, 'EN non-empty but NL empty'))
            continue
        if not en and nl:
            cell_flags.append((r, 'EMPTY-EN', sp, en, nl, 'NL non-empty but EN empty'))
            continue
        if len(en) >= MIN_EN_CHARS:
            ratio = len(nl) / len(en) if len(en) else 0
            if ratio < LEN_RATIO_LO or ratio > LEN_RATIO_HI:
                cell_flags.append((r, 'LEN-RATIO',
                                   sp, en, nl,
                                   f'NL/EN char-ratio {ratio:.2f} outside [{LEN_RATIO_LO},{LEN_RATIO_HI}]'))

    # Heuristic 2: cross-row paraphrase detection
    cross_flags = []
    nl_rows = [(r, sp, en, nl) for (r, sp, en, nl) in rows
               if len(nl) >= MIN_LEN_FOR_CROSS_ROW and en]
    nl_norm = {r: norm(nl) for (r, _, _, nl) in nl_rows}
    en_norm = {r: norm(en) for (r, _, en, _) in nl_rows}
    flagged_pairs = set()
    for i, (r_i, sp_i, en_i, nl_i) in enumerate(nl_rows):
        for r_j, sp_j, en_j, nl_j in nl_rows[i+1:]:
            pair = (r_i, r_j)
            if pair in flagged_pairs:
                continue
            nl_s = sim(nl_norm[r_i], nl_norm[r_j])
            if nl_s < NL_SIM_THRESHOLD:
                continue
            en_s = sim(en_norm[r_i], en_norm[r_j])
            if en_s > EN_DIFF_THRESHOLD:
                continue
            cross_flags.append((r_i, r_j, sp_i, sp_j, en_i, en_j, nl_i, nl_j, nl_s, en_s))
            flagged_pairs.add(pair)

    return cell_flags, cross_flags


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    extra_args = sys.argv[1:]
    if extra_args and extra_args[0].endswith('.xlsx'):
        extra_args = extra_args[1:]
    target_sheets = extra_args or wb.sheetnames
    total_cell = 0
    total_cross = 0
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f'!! sheet not found: {sheet_name}')
            continue
        ws = wb[sheet_name]
        cell_flags, cross_flags = scan_sheet(ws, sheet_name)
        if not cell_flags and not cross_flags:
            print(f'\n=== {sheet_name}: clean ===')
            continue
        print(f'\n=== {sheet_name} ===')
        if cell_flags:
            print(f'  --- per-cell flags ({len(cell_flags)}) ---')
            for r, kind, sp, en, nl, note in cell_flags:
                print(f'  J{r} [{kind}] {note}')
                print(f'    Speaker: {sp!r}')
                print(f'    EN: {en!r}')
                print(f'    NL: {nl!r}')
        if cross_flags:
            print(f'  --- cross-row paste candidates ({len(cross_flags)}) ---')
            for r_i, r_j, sp_i, sp_j, en_i, en_j, nl_i, nl_j, nl_s, en_s in cross_flags:
                print(f'  J{r_i} ↔ J{r_j}  (NL-sim={nl_s:.2f}, EN-sim={en_s:.2f})')
                print(f'    J{r_i} Speaker: {sp_i!r}')
                print(f'      EN: {en_i!r}')
                print(f'      NL: {nl_i!r}')
                print(f'    J{r_j} Speaker: {sp_j!r}')
                print(f'      EN: {en_j!r}')
                print(f'      NL: {nl_j!r}')
        total_cell += len(cell_flags)
        total_cross += len(cross_flags)
    print(f'\n=== TOTAL: {total_cell} per-cell + {total_cross} cross-row across {len(target_sheets)} sheet(s) ===')
    wb.close()


if __name__ == '__main__':
    main()
