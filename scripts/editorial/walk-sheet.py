#!/usr/bin/env python3
"""Walk a single sheet — emit per-row EN+NL+key+speaker for agent eyeballing.

Usage: walk-sheet.py <xlsx_filename> <sheet_name> [--start ROW] [--end ROW]

Example:
  walk-sheet.py 6_asses.masses_E6Proxy.xlsx E6_World_localization
  walk-sheet.py 10_asses.masses_E10Proxy.xlsx E10_Government_localization --start 1 --end 100
"""
import re
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')


def extract_speaker_from_key(key: str) -> str:
    """E5-E10: speaker is last '.'-separated segment of key."""
    if not key:
        return ''
    parts = str(key).rsplit('.', 1)
    return parts[-1].strip() if len(parts) > 1 else ''


def main():
    if len(sys.argv) < 3:
        print('Usage: walk-sheet.py <xlsx_filename> <sheet_name> [--start ROW] [--end ROW]')
        sys.exit(2)

    xlsx_name = sys.argv[1]
    sheet_name = sys.argv[2]
    start = 1
    end = None
    args = sys.argv[3:]
    while args:
        a = args.pop(0)
        if a == '--start': start = int(args.pop(0))
        elif a == '--end': end = int(args.pop(0))

    xlsx_path = REPO / 'excels' / xlsx_name
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
        (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
    if actual_sheet is None:
        print(f'ERROR: sheet {sheet_name!r} not found in {xlsx_name}')
        print(f'Available: {wb.sheetnames}')
        sys.exit(1)

    ws = wb[actual_sheet]
    max_row = ws.max_row if end is None else min(ws.max_row, end)

    print(f'# {xlsx_name} :: {actual_sheet}')
    print(f'# rows {start}..{max_row} (total in sheet: {ws.max_row})')
    print()

    # E5-E10 use col-A keys for speaker; E0-E4 use col-B descriptive
    is_a_speaker = any(actual_sheet.startswith(f'E{n}_') for n in range(5, 11))

    for r in range(max(start, 2), max_row + 1):
        key = ws.cell(row=r, column=1).value or ''
        desc = ws.cell(row=r, column=2).value or ''
        en = ws.cell(row=r, column=3).value or ''
        nl = ws.cell(row=r, column=10).value or ''
        if not any([key, desc, en, nl]):
            continue
        speaker = extract_speaker_from_key(key) if is_a_speaker else desc
        print(f'## J{r}')
        print(f'  KEY:     {key!r}')
        print(f'  SPEAKER: {speaker!r}')
        if not is_a_speaker:
            pass  # desc already shown via speaker
        else:
            if desc:
                print(f'  DESC:    {desc!r}')
        print(f'  EN:      {en!r}')
        print(f'  NL:      {nl!r}')
        print()


if __name__ == '__main__':
    main()
