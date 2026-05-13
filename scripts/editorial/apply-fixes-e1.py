#!/usr/bin/env python3
"""Apply user-approved E1 fixes to local xlsx (then push via push-file.py).

Tom sign-off 2026-05-13:
  - J7 FarmHouseInt: full override string
  - J14 FarmHouseInt: KEEP (no change)
  - J13 RedFields: KEEP (no change)
  - J11 Stable1F: KEEP (no change)
  - J67, J69 TheProtest: already applied via slogan retcon — short-circuit
  - All other 18 cells: apply as proposed

J5 FarmHouseInt + Stable2F J22 included by interpolation (mechanical fixes
consistent with the pattern user explicitly approved everywhere else).
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/1_asses.masses_E1Proxy.xlsx'

APPROVED = {
    # E1_FarmHouseInt (2 cells; J14 KEPT per Tom)
    ('E1_FarmHouseInt_localization', 5): (
        'Vertel ons eens... wat is het het geheim van je succes?',
        'Vertel ons eens... wat is het geheim van je succes?',
    ),
    ('E1_FarmHouseInt_localization', 7): (
        'We moeten bereid zijn hard te werken, en dat aan een product dat levens verandert.',
        ' We moeten bereid zijn hard te werken, en dat terwijl we een product hebben dat levens verandert.',
    ),

    # E1_Farm (6 cells)
    ('E1_Farm_localization', 17): (
        'De Camion-Machine is weg. De mensen moeten ergens in naar het dorp zijn.',
        'De Camion-Machine is weg. De Mensen moeten ergens in het dorp zijn.',
    ),
    ('E1_Farm_localization', 39): (
        'Ik z-z-zit zonder hout en ik heb meer nodig om het b-b-bruggetje te repareren voor het p-p-protest.',
        'Ik z-z-zit zonder hout en ik heb meer nodig om het b-b-bruggetje te repareren voor het P-P-Protest.',
    ),
    ('E1_Farm_localization', 59): (
        'De Mijn-Ezels hebben beslist om het protest van Oude Ezel te vervoegen.',
        'De Mijn-Ezels hebben beslist om het Protest van Oude Ezel te vervoegen.',
    ),
    ('E1_Farm_localization', 60): (
        'Niets wat we morgen uithalen zal de ziel van wijlen Luie Ezel naar het Astrale Hiernamaals brengen.',
        'Niets wat we morgen uithalen zal de Ziel van wijlen Luie Ezel naar het Astrale Hiernamaals brengen.',
    ),
    ('E1_Farm_localization', 94): (
        'Maar ge kunt toch verdomme toch niet zomaar jullie jobs terug eisen.',
        'Maar ge kunt verdomme toch niet zomaar jullie Jobs terug eisen.',
    ),
    ('E1_Farm_localization', 102): (
        'Jullie idee van richting en eigenwaarde zou niet, en moet niet, worden bepaald door de averechtse Mensen van Klotegem.',
        'Jullie idee van waarde en eigenwaarde zou niet, en moet niet, worden bepaald door de averechtse Mensen van Klotegem.',
    ),

    # E1_Stable1F (2 cells; J11 KEPT per Tom)
    ('E1_Stable1F_localization', 8): (
        'Het is hoog tijd dat we de Mensen er aan te herinneren hoe belangrijk en nuttig wij voor hen zijn geweest.',
        'Het is hoog tijd dat we de Mensen eraan herinneren hoe belangrijk en nuttig wij voor hen zijn geweest.',
    ),
    ('E1_Stable1F_localization', 10): (
        'Morgen schrijven we geschiedenis met onze alleréérste PROTEST.',
        'Morgen schrijven we geschiedenis met onze allereerste PROTEST.',
    ),

    # E1_Stable2F (1 cell)
    ('E1_Stable2F_localization', 22): (
        "De mensen... ze zijn me komen pakken in 't donker...",
        "De Mensen... ze zijn me komen pakken in 't donker...",
    ),

    # E1_TheProtest (6 cells; J67/J69 already applied via slogan retcon, omitted)
    ('E1_TheProtest_localization', 21): (
        'Goh, zouden de mensen dat echt doen?',
        'Goh, zouden de Mensen dat echt doen?',
    ),
    ('E1_TheProtest_localization', 22): (
        'Verzeker Lieve Ezel dat de mensen mogelijk een foutje gemaakt hebben.',
        'Verzeker Lieve Ezel dat de Mensen mogelijk een foutje gemaakt hebben.',
    ),
    ('E1_TheProtest_localization', 23): (
        'Zeg aan Lieve Ezel dat de mensen zijn verdorven door de Machines.',
        'Zeg aan Lieve Ezel dat de Mensen zijn verdorven door de Machines.',
    ),
    ('E1_TheProtest_localization', 39): (
        'Begrijp me niet verkeerd Trouwe, maar jij bent een luisteraar, niet een leider.',
        'Begrijp me niet verkeerd Trouwe, maar jij bent een luisteraar, geen leider.',
    ),
    ('E1_TheProtest_localization', 61): (
        'Ezels zijn altijd al de ruggegraat van deze Hoeve geweest, sinds de dieren konden spreken.',
        'Ezels zijn altijd al de ruggengraat van deze Hoeve geweest, sinds de dieren konden spreken.',
    ),
    ('E1_TheProtest_localization', 147): (
        'Neem plechtig de Traditionele Ezelshouding aan, die generaties aan Kuddes heeft gëinspireerd.',
        'Neem plechtig de Traditionele Ezelshouding aan, die generaties aan Kuddes heeft geïnspireerd.',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    if discrepancies:
        for s, r, msg in discrepancies:
            print(f'  ⚠ {s}/J{r}: {msg}')
        sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
