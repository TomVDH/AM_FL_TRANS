#!/usr/bin/env python3
"""Apply user-approved E3 fixes to local xlsx (then push via push-file.py).

Tom sign-off 2026-05-13:
  - J3 EpisodeTitle: fix (AFLVERING → AFLEVERING)
  - J3 Mine1FOpening: fix (kloteweeer → kloteweer)
  - J33 Mine1FOpening: fix (ratioenen → rantsoenen)
  - J53 Mine1F: OVERRIDE → Tom drops diminutive AND switches `dit` → `deze`
    `We zijn deze rotsblok aan 't verleggen, voor als de vloed nog dichter komt.`
  - J82 Mine1F: OVERRIDE → Café-prefix venue + verb-final word order
    `Ik zou 't... Café De Zatten Ezel noemen!`
  - J6 E3_100: SKIPPED — Tom called this a parser false-positive after
    clean re-read confirmed `455 ,` literal in cell. No write.
  - J28 LazysGrave: KEPT (Tom: `petje` loose-colloquial OK; J29 helm
    consistency not enforced)
  - J4 E3_200: fix (`Ik Had` → `Ik had`)

6 cells written, 2 KEPT/SKIPPED.
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/3_asses.masses_E3Proxy.xlsx'

APPROVED = {
    ('E3_EpisodeTitle_localization', 3): (
        'AFLVERING DRIE',
        'AFLEVERING DRIE',
    ),
    ('E3_Mine1FOpening_localization', 3): (
        'Kameraad Moeder, gaat dit kloteweeer ooit nog stoppen?',
        'Kameraad Moeder, gaat dit kloteweer ooit nog stoppen?',
    ),
    ('E3_Mine1FOpening_localization', 33): (
        'Slimme Ezel, verdeel de ratioenen.',
        'Slimme Ezel, verdeel de rantsoenen.',
    ),
    # Tom override — drops diminutive + switches `dit` (het-word) → `deze` (de-word)
    ('E3_Mine1F_localization', 53): (
        "We zijn dit rotsbloki aan 't verleggen, voor als de vloed nog dichter komt.",
        "We zijn deze rotsblok aan 't verleggen, voor als de vloed nog dichter komt.",
    ),
    # Tom override — Café-prefix venue name + verb-final word order
    ('E3_Mine1F_localization', 82): (
        "Ik zou 't noemen... De Zatten Ezel Cafe!",
        "Ik zou 't... Café De Zatten Ezel noemen!",
    ),
    ('E3_200_localization', 4): (
        '24 juli\\nVanscheetvelde Kolen & Industrie staat officieel te koop. Ik hoop dat iemand van Technopolis het zal kopen. Ik Had altijd gedacht dat Piet mijn plaats zou innemen, maar...\\n\\nDeze plek moet overleven, op één of andere manier, maar het is moeilijk om los te laten wat ge graag ziet.',
        '24 juli\\nVanscheetvelde Kolen & Industrie staat officieel te koop. Ik hoop dat iemand van Technopolis het zal kopen. Ik had altijd gedacht dat Piet mijn plaats zou innemen, maar...\\n\\nDeze plek moet overleven, op één of andere manier, maar het is moeilijk om los te laten wat ge graag ziet.',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    print(f'Loading {XLSX.name}…')
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    if discrepancies:
        for s, r, msg in discrepancies:
            print(f'  ⚠ {s}/J{r}: {msg}')
        sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
