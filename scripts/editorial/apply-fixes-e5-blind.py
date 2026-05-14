#!/usr/bin/env python3
"""Apply E5_ZooMain blind-spot walk fixes — Tom sign-off 2026-05-13.

21 writes / 9 keeps. Cells across §11.1 codex-lock, §5.1 register,
§12.4 English bleed, §7.3.1 cap, §13 BUGS, §6.7 cross-cell, §9.3 punct,
§12.3 stutter, §13 minor notes.

Tom decisions:
  - J109 — fix + Tom tweak: add "de" between "van" and "eerste"
  - J96 LOSERS KEPT — stylistic loanword punch
  - J183 Tom override: `komt vandaag on-site`
  - J100 Tom idiom override: `Nu zijn we van haar dikke nek af!`
  - J199/J208 Tom new form (both unified): `Moge hij het beste geluk hebben om een Job zonder Machines te vinden!` — this OVERRIDES J208's earlier FINAL push value
  - J207 Tom tweak: `jong` → `joh`
  - J31 Tom override: `Oe-oe-oef!` (preserves vowel cluster + stutter)
  - J108 Tom tweak: `plek` → `plaats`
  - J38/J197/J172 KEPT
  - J51 vs J217 KEPT distinct (Day-3 banter vs Day-7 anger — register-appropriate)
  - J176 KEEPS `gezellig`; J190 changes to `gezellig`
"""
import sys
from pathlib import Path
import openpyxl

REPO = Path('/Users/tomlinson/Projects/VIBE CODING/am-fl-trans')
XLSX = REPO / 'excels/5_asses.masses_E5Proxy.xlsx'

# Shared form for J199/J208 (Tom's new unified value)
J199_J208_FORM = 'Moge hij het beste geluk hebben om een Job zonder Machines te vinden!'

APPROVED = {
    # §11.1 codex (2 writes)
    ('E5_ZooMain_localization', 118): (
        '*boe-hoe-hoe-hoe*',
        '*boe-hoe-hoe*',
    ),
    ('E5_ZooMain_localization', 123): (
        '*boe-hoe-hoe-hoe*',
        '*boe-hoe-hoe*',
    ),

    # §5.1 register (1 write) — Tom adds "de"
    ('E5_ZooMain_localization', 109): (
        'Platvoet! Kleine stap verwijderd van eerste plaats! Pakt ze!',
        'Platvoet! Kleine stap verwijderd van de eerste plaats! Pak ze!',
    ),

    # §12.4 English bleed (4 writes; J96 KEPT)
    ('E5_ZooMain_localization', 48): (
        'Och j-j-jongens.',
        'O j-j-jee.',
    ),
    ('E5_ZooMain_localization', 81): (
        'Ach j-j-jongens. Vandaag was b-b-bijzonder h-h-hard...',
        'O j-j-jee. Vandaag was b-b-bijzonder h-h-hard...',
    ),
    # J183 Tom override
    ('E5_ZooMain_localization', 183): (
        'De Raad van Bestuur van de Dierentuin gaat vandaag op site om de Ezel van de Week te ontmoeten?',
        'De Raad van Bestuur van de Dierentuin komt vandaag on-site om de Ezel van de Week te ontmoeten?',
    ),
    # J97 fix (context-confirmed)
    ('E5_ZooMain_localization', 97): (
        'H-h-haat mij alsjeblief n-n-niet maar...',
        'H-h-haat mij n-n-niet...',
    ),

    # §7.3.1 (1 write)
    ('E5_ZooMain_localization', 184): (
        'En ze overwegen meer geld en middelen naar ons deel van de dierentuin te sturen?',
        'En ze overwegen meer geld en middelen naar ons deel van de Dierentuin te sturen?',
    ),

    # §13 BUGS (3 writes; J38/J197/J172 KEPT)
    ('E5_ZooMain_localization', 155): (
        'Maar ik weet dat dit niet met ieder van jullie hier te vinden valt.',
        'Maar ik weet dat ik het hier bij jullie niet ga vinden.',
    ),
    # J100 Tom override
    ('E5_ZooMain_localization', 100): (
        'Nu moeten we niet meer met haar opgeblazen ego te maken hebben!',
        'Nu zijn we van haar dikke nek af!',
    ),
    ('E5_ZooMain_localization', 122): (
        'En jullie allemaal zonder mij zijn verdergegaan.',
        'En jullie zijn allemaal zonder mij verdergegaan.',
    ),

    # §6.7 cross-cell — Tom new unified form for J199 + J208 (J51/J217 KEPT, J176 keeps gezellig)
    ('E5_ZooMain_localization', 199): (
        "En 't beste voor hem om een job zonder Machines te vinden!",
        J199_J208_FORM,
    ),
    # J208 was Tom's earlier override `'t Beste aan hem om een job zonder Machines te zoeken!`
    # Tom now re-overrides both to the new shared form.
    ('E5_ZooMain_localization', 208): (
        "'t Beste aan hem om een job zonder Machines te zoeken!",
        J199_J208_FORM,
    ),
    # J190 changes to gezellig (match J176)
    ('E5_ZooMain_localization', 190): (
        'De RAAD VAN BESTUUR wil dat jullie ze knus komen Knuffelen!',
        'De RAAD VAN BESTUUR wil dat jullie ze gezellig komen Knuffelen!',
    ),

    # §9.3 punct (4 writes)
    ('E5_ZooMain_localization', 13): (
        'Sofie! Iedereen ziet hoe pienter ze wel niet is.',
        'Sofie! Iedereen ziet hoe pienter ze wel niet is!',
    ),
    ('E5_ZooMain_localization', 67): (
        'Er komen vandaag KINDEREN langs en die vinden het zo fijn wanneer je BALKT.',
        'Er komen vandaag KINDEREN langs en die vinden het zo fijn wanneer je BALKT!',
    ),
    ('E5_ZooMain_localization', 204): (
        '...niet eens s-s-salut gezegd.',
        '...niet eens s-s-salut gezegd!',
    ),
    # J207 with Tom tweak jong → joh
    ('E5_ZooMain_localization', 207): (
        'Niets, jong. Absoluut niet.',
        'Geen sprake van, joh! Absoluut niet.',
    ),

    # §12.3 stutter (2 writes)
    ('E5_ZooMain_localization', 29): (
        'Je kan niet in a-a-alles goed z-z-zijn, Slim—',
        'Je kan niet in alles goed z-z-zijn, Slim—',
    ),
    # J31 Tom override (Oe-oe-oef preserves vowel cluster)
    ('E5_ZooMain_localization', 31): (
        'O-o-oef! Ik ben uitged-d-droogd!',
        'Oe-oe-oef! Ik ben uitged-d-droogd!',
    ),

    # §13 minor (1 write — J108; J85/J189/J95/J193 KEPT)
    ('E5_ZooMain_localization', 108): (
        'Ik heb het gevoel dat jij het nog tot de eerste plek gaat schoppen!',
        'Ik heb het gevoel dat jij het nog tot de eerste plaats gaat schoppen!',
    ),
}


def main():
    apply_mode = '--apply' in sys.argv
    wb = openpyxl.load_workbook(XLSX)
    will_write = []
    already_applied = []
    discrepancies = []

    for (sheet_name, row), (expected, proposed) in APPROVED.items():
        actual_sheet = sheet_name if sheet_name in wb.sheetnames else next(
            (s for s in wb.sheetnames if s.startswith(sheet_name) or sheet_name.startswith(s)), None)
        if actual_sheet is None:
            discrepancies.append((sheet_name, row, 'TAB-NOT-FOUND'))
            continue
        cell = wb[actual_sheet].cell(row=row, column=10)
        current = (cell.value or '').strip() if cell.value is not None else ''
        if current == expected.strip():
            will_write.append((actual_sheet, row, current, proposed))
        elif current == proposed.strip():
            already_applied.append((actual_sheet, row, current))
        else:
            discrepancies.append((actual_sheet, row, f'UNEXPECTED: {current!r}'))

    print(f'\nStatus: {len(will_write)} write / {len(already_applied)} already / {len(discrepancies)} discrepant')
    for s, r, msg in discrepancies:
        print(f'  ⚠ {s}/J{r}: {msg}')
    if discrepancies: sys.exit(1)
    if not will_write:
        print('All cells already at proposed values.')
        return

    print('\nProposed writes:')
    for sheet, row, current, proposed in will_write:
        print(f'  {sheet}/J{row}:')
        print(f'    -  {current[:100]!r}')
        print(f'    +  {proposed[:100]!r}')

    if not apply_mode:
        print('\nDRY-RUN. Re-run with --apply.')
        return

    for sheet, row, _, proposed in will_write:
        wb[sheet].cell(row=row, column=10).value = proposed
    wb.save(XLSX)
    print(f'\n✓ Wrote {len(will_write)} cells.')


if __name__ == '__main__':
    main()
