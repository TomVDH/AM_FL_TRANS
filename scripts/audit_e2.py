import openpyxl
import os

file_path = 'excels/2_asses.masses_E2Proxy.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

rules = ['Machine', 'Machines', 'Oom', 'Nonkel', 'Boerderij', 'ezel', 'hoe-hoe', 'U ', ' u ', 'Uw ', ' uw ']
sheets_to_scan = [sn for sn in wb.sheetnames if 'Title' not in sn]

audit_results = []

for sn in sheets_to_scan:
    ws = wb[sn]
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        # Col C is index 2, Col D is index 3, Col J is index 9
        en = str(row[2]) if len(row) > 2 and row[2] else ''
        speaker = str(row[3]) if len(row) > 3 and row[3] else ''
        nl = str(row[9]) if len(row) > 9 and row[9] else ''
        
        if not nl:
            continue
            
        found_rule = False
        for r in rules:
            if r.lower() in nl.lower():
                found_rule = True
                break
        
        if found_rule:
            audit_results.append({
                "sheet": sn,
                "row": idx,
                "speaker": speaker,
                "en": en,
                "nl": nl
            })

import json
print(json.dumps(audit_results))
