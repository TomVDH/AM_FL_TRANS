#!/usr/bin/env python3
"""
push-tab.py - interactive line-by-line push of column J from local xlsx to remote Google Sheet.

USAGE
  py scripts/convert/push-tab.py <file.xlsx> <tab>

For each row where local J differs from remote J, prints EN (col C) + WAS (remote) + NOW (local)
and waits for input:
  Enter = push this cell    s = skip    q = quit
"""
import sys
from pathlib import Path
import openpyxl
import gspread
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_DIR = PROJECT_ROOT / 'excels'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
PUSHED_GREEN = {'backgroundColor': {'red': 0.851, 'green': 0.918, 'blue': 0.827}}

SHEETS = {
    '0_asses.masses_Manager+Intermissions+E0Proxy.xlsx': '1ScGhva1rUcwqup5rWjePK43gnvxME4mVgZvPZlIfWHQ',
    '1_asses.masses_E1Proxy.xlsx':  '1TTUN6f_DACWw2VFxPP5sBbzKSMvpMZrG8XTsqcpw0SU',
    '2_asses.masses_E2Proxy.xlsx':  '14LxAqOh5lBlDo6hqcF86Q2IJHwTAfrO36xBKUDjk5ow',
    '3_asses.masses_E3Proxy.xlsx':  '1dPFzn_FCcmgLx8HuuMcxBa4X29ixYnlbNqNkcVN95yY',
    '4_asses.masses_E4Proxy.xlsx':  '1KzgaFaoUWE2Jv5eBfTryh8kiw944Q3MBNwtlaofpOn0',
    '5_asses.masses_E5Proxy.xlsx':  '1yxDuFb6NDDYHytA_oAi9EgprHxdzE8zYNu1KgA7Pp2E',
    '6_asses.masses_E6Proxy.xlsx':  '1_56yhltRDywIj1WpCP4nJAMjaBmFVC42fSwXMj2W0uU',
    '7_asses.masses_E7Proxy.xlsx':  '1evCSy0b20NNb4mfndqDVtc-xtfZ_a7bmTm2lxlpSuYA',
    '8_asses.masses_E8Proxy.xlsx':  '1mJjvrManB-mwN-2bawLltEeRoUTq-6ch9OW5_Zyu6v0',
    '9_asses.masses_E9Proxy.xlsx':  '1zG-g9HCyECCTzY3Yc-Un-K5I1DYAjvwD9GdHj7mle_c',
    '10_asses.masses_E10Proxy.xlsx': '1Q1WSEJSGm9ZPEO6bfDNpX8wqXolFhyGad-17rJzc0nw',
}


def get_client():
    token_path = PROJECT_ROOT / 'token.json'
    creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        token_path.write_text(creds.to_json())
    return gspread.authorize(creds)


def norm(v):
    return '' if v is None else str(v)


def main():
    if len(sys.argv) != 3:
        print('Usage: push-tab.py <file.xlsx> <tab>')
        sys.exit(2)
    fname, tab = sys.argv[1], sys.argv[2]
    if fname not in SHEETS:
        print(f'ERROR: {fname} not in SHEETS map')
        sys.exit(1)
    xlsx_path = XLSX_DIR / fname
    if not xlsx_path.exists():
        print(f'ERROR: {xlsx_path} not found')
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if tab not in wb.sheetnames:
        print(f'ERROR: tab {tab!r} not in {fname}. Available: {wb.sheetnames}')
        sys.exit(1)
    ws = wb[tab]

    local_en, local_nl = {}, {}
    for r, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        en = row[2] if len(row) > 2 else None
        nl = row[9] if len(row) > 9 else None
        local_en[r] = norm(en)
        local_nl[r] = norm(nl)
    wb.close()

    # Header sanity: J1 must say NL (otherwise column layout is off, abort)
    local_j1 = local_nl.get(1, '').strip().upper()
    if local_j1 != 'NL':
        print(f'ABORT: local {fname} / {tab} J1 header is {local_j1!r}, expected "NL".')
        print('Column J does not appear to be the NL translation column. Refusing to push.')
        sys.exit(1)

    print(f'Connecting to remote {fname} / {tab}...')
    gs = get_client().open_by_key(SHEETS[fname]).worksheet(tab)
    remote_J = gs.col_values(10)
    remote_nl = {i + 1: norm(v) for i, v in enumerate(remote_J)}

    remote_j1 = remote_nl.get(1, '').strip().upper()
    if remote_j1 != 'NL':
        print(f'ABORT: remote {fname} / {tab} J1 header is {remote_j1!r}, expected "NL".')
        print('Column J does not appear to be the NL translation column. Refusing to push.')
        sys.exit(1)

    all_rows = sorted(set(local_nl) | set(remote_nl))
    diffs = [r for r in all_rows if local_nl.get(r, '') != remote_nl.get(r, '')]

    if not diffs:
        print(f'No diffs in {tab}. Nothing to push.')
        return

    print(f'\n{len(diffs)} diff(s) in {tab}.\n')

    pushed = skipped = 0
    for i, r in enumerate(diffs, 1):
        en = local_en.get(r, '')
        was = remote_nl.get(r, '')
        now = local_nl.get(r, '')
        print(f'[ {i} / {len(diffs)} ]  J{r}')
        print(f'  EN:   {en}')
        print(f'  WAS:  {was}')
        print(f'  NOW:  {now}')
        ans = input('  [Enter] push   [s] skip   [q] quit  > ').strip().lower()
        if ans == 'q':
            print('Quit.')
            break
        if ans == 's':
            skipped += 1
            print('  ~ skipped\n')
            continue
        gs.update(values=[[now]], range_name=f'J{r}', value_input_option='USER_ENTERED')
        gs.format(f'J{r}', PUSHED_GREEN)
        pushed += 1
        print('  + pushed (tinted)\n')

    print(f'\nDone. pushed={pushed}  skipped={skipped}  remaining={len(diffs) - pushed - skipped}')


if __name__ == '__main__':
    main()
