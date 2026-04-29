#!/usr/bin/env python3
"""
push-file.py - batch push all column-J diffs for one xlsx file to its remote Google Sheet.

Reads all tabs, diffs local J vs remote J, then per tab:
  - 1 API call: batch write all changed cells
  - 1 API call: batch tint all changed cells light green
= 2 calls per tab total (vs 2 per cell in push-tab.py)

USAGE
  py scripts/convert/push-file.py <file.xlsx>            # dry-run
  py scripts/convert/push-file.py <file.xlsx> --apply    # push
"""
import sys
import time
from pathlib import Path
import openpyxl
import gspread
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

PROJECT_ROOT = Path(__file__).resolve().parents[2]
XLSX_DIR = PROJECT_ROOT / 'excels'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
GREEN = {'red': 0.851, 'green': 0.918, 'blue': 0.827}
CALL_GAP = 1.1  # seconds between API calls

SHEETS = {
    '0_asses.masses_Manager+Intermissions+E0Proxy.xlsx': '1ScGhva1rUcwqup5rWjePK43gnvxME4mVgZvPZlIfWHQ',
    '1_asses.masses_E1Proxy.xlsx':  '1TTUN6f_DACWw2VFxPP5sBbzKSMvpMZrG8XTsqcpw0SU',
    '2_asses.masses_E2Proxy.xlsx':  '14LxAqOh5lBlDo6hqcF86Q2IJHwTAfrO36xBKUDjk5ow',
    '3_asses.masses_E3Proxy.xlsx':  '1dPFzn_FCcmgLx8HuuMcxBa4X29ixYnlbNqNkcVN95yY',
    '4_asses.masses_E4Proxy.xlsx':  '1KzgaFaoUWE2Jv5eBfTryh8kiw944Q3MBNwtlaofpOn0',
    '5_asses.masses_E5Proxy.xlsx':  '1yxDuFb6NDDYHytA_oAi9EgprHxdzE8zYNu1KgA7Pp2E',
    '6_asses.masses_E6Proxy.xlsx':  '1_56yhltRDywIj1WpCP4nJAMjaBmFVC42fSwXMj2W0uU',
    '7_asses.masses_E7Proxy.xlsx':  '1evCSy0b20NNb4mfndqDVtc-xtfZ_a7bmTm2lxlpSuYA',
    '8_asses.masses_E8Proxy.xlsx':  '1mJjvrManB-mwN-2bawLltEeRoUTq-6ch9OW5_Zyu6v0',
    '9_asses.masses_E9Proxy.xlsx':  '1zG-g9HCyECCTzY3Yc-Un-K5I1DYAjvwD9GdHj7mle_c',
    '10_asses.masses_E10Proxy.xlsx': '1Q1WSEJSGm9ZPEO6bfDNpX8wqXolFhyGad-17rJzc0nw',
}


def get_client():
    token_path = PROJECT_ROOT / 'token.json'
    creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        token_path.write_text(creds.to_json())
    return gspread.authorize(creds)


def norm(v):
    return '' if v is None else str(v)


def main():
    if len(sys.argv) < 2:
        print('Usage: push-file.py <file.xlsx> [--apply]')
        sys.exit(2)
    fname = sys.argv[1]
    apply = '--apply' in sys.argv
    if fname not in SHEETS:
        print(f'ERROR: {fname} not in SHEETS map')
        sys.exit(1)
    xlsx_path = XLSX_DIR / fname
    if not xlsx_path.exists():
        print(f'ERROR: {xlsx_path} not found')
        sys.exit(1)

    # Load local
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    local_tabs = {}
    for tab in wb.sheetnames:
        ws = wb[tab]
        rows = {}
        for r, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            en = row[2] if len(row) > 2 else None
            nl = row[9] if len(row) > 9 else None
            rows[r] = (norm(en), norm(nl))
        local_tabs[tab] = rows
    wb.close()

    gc = get_client()
    sh = gc.open_by_key(SHEETS[fname])

    total_pushed = 0
    tabs_with_diffs = 0

    for tab in wb.sheetnames:
        local = local_tabs[tab]

        # Header check
        j1 = local.get(1, ('', ''))[1].strip().upper()
        if j1 != 'NL':
            print(f'  SKIP {tab}: J1={j1!r} (not NL)')
            continue

        # Excel caps names at 31 chars; remote may have full name — match by prefix
        ws_remote = None
        for ws_candidate in sh.worksheets():
            if ws_candidate.title == tab or ws_candidate.title.startswith(tab):
                ws_remote = ws_candidate
                if ws_candidate.title != tab:
                    print(f'  NOTE {tab} -> {ws_candidate.title} (Excel truncation)')
                break
        if ws_remote is None:
            print(f'  SKIP {tab}: not found in remote')
            continue

        time.sleep(CALL_GAP)
        remote_J = ws_remote.col_values(10)
        remote_nl = {i + 1: norm(v) for i, v in enumerate(remote_J)}

        all_rows = sorted(set(local) | set(remote_nl))
        diffs = []
        for r in all_rows:
            en, lnl = local.get(r, ('', ''))
            rnl = remote_nl.get(r, '')
            if lnl != rnl:
                diffs.append((r, lnl))

        if not diffs:
            continue

        tabs_with_diffs += 1
        print(f'\n  {tab}: {len(diffs)} diff(s)')
        for r, val in diffs:
            snippet = val[:70].encode('ascii', errors='replace').decode()
            print(f'    J{r}: {snippet}')

        if not apply:
            continue

        # Batch write: 1 API call
        batch_data = [{'range': f'J{r}', 'values': [[val]]} for r, val in diffs]
        time.sleep(CALL_GAP)
        ws_remote.batch_update(batch_data, value_input_option='USER_ENTERED')

        # Batch tint: 1 API call
        requests = [{
            'repeatCell': {
                'range': {
                    'sheetId': ws_remote.id,
                    'startRowIndex': r - 1,
                    'endRowIndex': r,
                    'startColumnIndex': 9,
                    'endColumnIndex': 10,
                },
                'cell': {'userEnteredFormat': {'backgroundColor': GREEN}},
                'fields': 'userEnteredFormat.backgroundColor',
            }
        } for r, _ in diffs]
        time.sleep(CALL_GAP)
        sh.batch_update({'requests': requests})

        total_pushed += len(diffs)
        print(f'    -> pushed + tinted {len(diffs)} cells')

    mode = 'APPLIED' if apply else 'DRY RUN'
    print(f'\n[{mode}] {fname}: {total_pushed} cells across {tabs_with_diffs} tabs')
    if not apply:
        print('Re-run with --apply to write.')


if __name__ == '__main__':
    main()
